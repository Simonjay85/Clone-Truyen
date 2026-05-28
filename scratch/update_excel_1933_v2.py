import openpyxl

wb = openpyxl.load_workbook("danh_sach_truyen_doctieuthuyet.xlsx")
ws = wb.active

# Row 192 = story 1933 (header row 4, data starts row 5)
row = 192

# Update Col 3 (Tên Truyện) - already done by previous script
ws.cell(row=row, column=3).value = "Cô Gái Ngồi Cạnh Máy In, Phòng Kế Toán Gọi Tôi Là Đứa Đánh Máy"

# Update Col 4 (Tác Giả)
ws.cell(row=row, column=4).value = "Phạm Hoàng Minh"

# Update Col 5 (Thể Loại)  
ws.cell(row=row, column=5).value = "Sảng Văn"

# Update Col 8 (Trạng Thái) - now published
ws.cell(row=row, column=8).value = "publish"

# Update Col 12 (Tóm Tắt)
ws.cell(row=row, column=12).value = "Nguyễn Khánh Linh ngồi ở góc khuất nhất phòng kế toán Tập đoàn Bất động sản Thành Phát — cái bàn sát máy in. Ba năm bị gọi là \"con bé đánh máy\", cô âm thầm phát hiện lỗ hổng 47 tỷ đồng trong sổ sách. Khi cô quyết định lên tiếng, cả phòng kế toán quay lưng."

# Update Col 13 (Điểm Cần Sửa)
ws.cell(row=row, column=13).value = "REWRITE TOÀN BỘ: Chuyển Việt Nam, kiểm toán nội bộ, loại bỏ siêu nhiên + tên TQ → ĐÃ HOÀN THÀNH"

# Update Col 14 (Trạng Thái Sửa) 
ws.cell(row=row, column=14).value = "☑️ Đã sửa"

wb.save("danh_sach_truyen_doctieuthuyet.xlsx")
print("✅ Excel updated for story 1933:")
print(f"   Title: {ws.cell(row=row, column=3).value}")
print(f"   Author: {ws.cell(row=row, column=4).value}")
print(f"   Genre: {ws.cell(row=row, column=5).value}")
print(f"   Status: {ws.cell(row=row, column=8).value}")
print(f"   Fix Status: {ws.cell(row=row, column=14).value}")
