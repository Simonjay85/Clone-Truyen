import importlib.util
import json
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
spec = importlib.util.spec_from_file_location("novel_editor", ROOT / "scratch" / "novel_editor.py")
editor = importlib.util.module_from_spec(spec)
spec.loader.exec_module(editor)

STORY_ID = 5780
TITLE = "Bị Ép Nghỉ Vì Từ Chối Phong Bì, Tôi Lập Viện Riêng"
INTRO = (
    "Mười lăm năm làm Trưởng khoa Ngoại ở BV Nhân Dân, bác sĩ Trần Minh Đức nổi tiếng vì không nhận phong bì. "
    "Một ca mổ cứu người khiến ông trả lại chiếc phong bì dày trước mặt cả phòng trực, và cũng khiến ban giám đốc coi ông là cái gai. "
    "Bị cô lập, bị ép nghỉ vì 'gây mất đoàn kết', ông rời bệnh viện lớn với hai bàn tay sạch. "
    "Nhưng khi ông lập một viện nhỏ minh bạch, bệnh nhân bắt đầu bỏ nơi hào nhoáng để tìm đến nơi người bác sĩ còn nhìn họ như con người."
)

CHAPTERS = {
    5782: ("Chương 1: Phong Bì Bị Trả Lại", """
Ca cấp cứu nhập viện lúc hai giờ mười bảy phút sáng. Một phụ nữ bốn mươi sáu tuổi vỡ ruột thừa, sốt cao, bụng cứng như mặt trống, người nhà chạy sau băng ca vừa khóc vừa gọi bác sĩ. Trần Minh Đức đang trực ngoại tổng quát. Ông đã năm mươi hai tuổi, tóc muối tiêu, mắt sâu vì thiếu ngủ, nhưng khi bước vào phòng cấp cứu, giọng vẫn bình tĩnh như dao mổ vừa được khử trùng.

"Chuẩn bị phòng mổ. Công thức máu, điện giải, kháng sinh ngay. Người nhà ký cam kết, đừng để bệnh nhân chờ."

Con trai bệnh nhân kéo ông ra góc hành lang, dúi vào túi áo ông một phong bì dày. Cậu ta run đến mức lời nói vỡ vụn: "Bác sĩ cứu mẹ cháu. Nhà cháu có nhiêu đây. Nếu thiếu cháu vay thêm." Minh Đức lấy phong bì ra, đặt lại vào tay cậu. "Tiền này giữ để mua thuốc sau mổ. Mẹ cậu cần bác sĩ, không cần hối lộ."

Cậu thanh niên sững người. Hai điều dưỡng quay đi, như thể cảnh ấy quá quen nhưng vẫn quá nguy hiểm để nhìn thẳng. Ở BV Nhân Dân, người ta không gọi phong bì là hối lộ. Người ta gọi nó là "tấm lòng", "cà phê", "biết điều". Ai nhận thì sống yên, ai không nhận thì làm người khác khó chịu. Minh Đức đã làm khó chịu cả bệnh viện suốt mười lăm năm.

Ca mổ kéo dài hơn hai giờ. Ruột thừa đã hoại tử, ổ bụng nhiễm trùng nặng. Minh Đức đứng chính, tay chắc, lời ngắn. Khi bệnh nhân qua cơn nguy, ông ra ngoài báo cho gia đình. Cậu con trai òa khóc, quỳ xuống ngay giữa hành lang. Minh Đức đỡ cậu dậy. "Đừng quỳ. Về pha cho mẹ cốc nước ấm khi bà tỉnh."

Sáng hôm sau, câu chuyện phong bì bị trả lại lan khắp khoa. Người khen nhỏ, người chê lớn. Một bác sĩ trẻ nói nửa đùa nửa thật: "Thầy trả thì thầy cao đẹp, tụi em nhận lại thành xấu xa hết." Minh Đức nhìn cậu ta rất lâu. "Nếu em thấy xấu, có nghĩa em vẫn còn biết ranh giới."

Đến mười giờ, giám đốc bệnh viện Nguyễn Văn Hải gọi ông lên phòng. Phòng giám đốc rộng, có máy lạnh thơm mùi tinh dầu, tường treo bằng khen và ảnh bắt tay lãnh đạo. Hải trẻ hơn Minh Đức gần mười tuổi, áo blouse luôn trắng đến lạnh. Ông ta không mời ngồi ngay, chỉ chỉ vào ghế sau khi để Minh Đức đứng đủ lâu.

"Anh Đức, tôi nghe chuyện đêm qua."

"Bệnh nhân ổn."

"Tôi không nói bệnh nhân. Tôi nói cách anh làm mọi người khó xử." Hải gõ ngón tay lên bàn. "Anh trả phong bì trước mặt nhân viên, người nhà. Anh muốn truyền thông biết bệnh viện này có phong bì à?"

Minh Đức bình thản: "Nếu sợ người ta biết, tốt nhất là đừng để nó tồn tại."

Không khí trong phòng lạnh xuống. Hải tựa lưng, nụ cười mỏng như lưỡi dao cùn. "Anh giỏi chuyên môn, tôi công nhận. Nhưng bệnh viện là tập thể. Tập thể cần hòa khí. Anh cứ làm như mình là người sạch duy nhất, anh khiến anh em mất đoàn kết."

"Tôi không nhận tiền của người đang sợ mất mẹ. Nếu chuyện đó làm mất đoàn kết, đoàn kết ấy đáng xem lại."

Hải đóng tập hồ sơ trước mặt. "Anh nên nghỉ vài ngày. Suy nghĩ về thái độ của mình."

Minh Đức đứng dậy. Ông không cãi thêm. Người từng cầm dao trong ổ bụng nhiễm trùng hiểu có những vết mủ phải rạch đúng lúc, nhưng cũng có lúc cần để nó tự gom lại cho rõ. Khi ông bước ra, cô thư ký nhìn ông bằng ánh mắt vừa thương vừa sợ. Dọc hành lang, tiếng xe đẩy, tiếng máy monitor, tiếng người nhà gọi bác sĩ vẫn vang lên. Bệnh viện bên ngoài vẫn chạy, nhưng bên trong Minh Đức, một điều gì đó đã nứt.

Tối đó, người phụ nữ được cứu tỉnh lại. Cậu con trai tìm ông, đưa một túi cam thay cho phong bì. "Bác sĩ bảo giữ tiền mua thuốc, nhà cháu mua cam bằng tiền lẻ còn lại. Cái này bác sĩ nhận được không?" Minh Đức nhận hai quả, không nhận cả túi. Ông đặt một quả lên bàn trực, bóc một quả chia cho điều dưỡng. Vị cam chua, nhưng trong cổ họng ông lại ngọt. Đó là thứ biết ơn không làm người bệnh nghèo thêm.
"""),
    5783: ("Chương 2: Bị Cô Lập Trong Bệnh Viện", """
Sau cuộc gặp với giám đốc Hải, Minh Đức không bị kỷ luật ngay. Người ta tinh vi hơn thế. Lịch mổ của ông bị cắt dần, những ca khó chuyển sang người khác rồi âm thầm gọi ông xuống "hỗ trợ" khi biến chứng. Các cuộc họp chuyên môn đổi giờ mà không báo. Hồ sơ đề xuất mua máy nội soi mới bị trả về vì thiếu chữ ký phụ trách, trong khi chữ ký ấy thuộc về một phó khoa vốn chưa từng đọc hết bản thuyết minh.

Ở phòng ăn bác sĩ, ghế cạnh ông thường trống. Không ai nói ghét ông. Họ chỉ đột nhiên bận điện thoại, bận lấy thêm canh, bận ngồi bàn khác. Một điều dưỡng lâu năm lén đặt trước mặt ông ly cà phê đen rồi nói nhỏ: "Thầy đừng buồn. Người ta sợ." Minh Đức mỉm cười. "Tôi biết. Sợ cũng là một bệnh nghề nghiệp ở đây."

Bác sĩ trẻ từng học dưới tay ông bắt đầu tránh mắt. Trong số đó có Khải, học trò cũ ông quý nhất. Khải giỏi, tay mổ khéo, từng nói muốn trở thành bác sĩ không để người nghèo sợ bệnh viện. Nhưng giờ Khải phụ trách các ca dịch vụ, mặc áo blouse thêu tên bóng bẩy, và biết cách nhận phong bì kín đáo hơn đàn anh. Khi gặp Minh Đức ở cầu thang, Khải cúi đầu chào rồi đi rất nhanh.

Sự cô lập không đau bằng một nhát dao. Nó giống nước nhỏ vào đá, mỗi ngày một ít. Sáng ông vào khoa, vài câu chuyện ngưng giữa chừng. Trưa ông ký hồ sơ, người ta soi từng lỗi chính tả. Chiều ông góp ý ca mổ, phó khoa đáp: "Thời của anh khác rồi. Bây giờ bệnh viện phải tự chủ tài chính." Minh Đức hỏi: "Tự chủ nghĩa là bệnh nhân tự lo phong bì à?" Cả phòng họp im bặt.

Hải tận dụng sự im lặng ấy. Ông ta phát động chiến dịch "văn hóa phục vụ mới", treo poster bác sĩ cười, nhân viên cúi chào, quầy dịch vụ sạch như khách sạn. Nhưng ở phòng cấp cứu, bệnh nhân nghèo vẫn nằm ghép. Ở khoa ngoại, người nhà vẫn thì thầm hỏi điều dưỡng nên "gửi bác sĩ bao nhiêu". Minh Đức thấy sự hào nhoáng phủ lên một vết thương hôi, và càng thấy khó thở.

Một buổi chiều, ông phát hiện bệnh nhân ung thư đại tràng bị trì hoãn mổ vì chưa đóng đủ gói dịch vụ. Người đàn ông ấy làm thợ hồ, vợ bán rau, con gái ôm sổ viện phí ngồi khóc ngoài hành lang. Minh Đức ký chỉ định mổ cấp thiết, ghi rõ nguy cơ tắc ruột. Phòng tài chính gọi lên chất vấn. Ông đáp: "Nếu ruột vỡ trong lúc các anh chờ tiền, ai ký giấy chịu trách nhiệm?"

Ca mổ thành công. Nhưng ngay tối đó, Hải gửi email toàn viện: nghiêm cấm bác sĩ tự ý can thiệp quy trình tài chính, mọi trường hợp vượt tuyến dịch vụ phải được ban giám đốc phê duyệt. Không nêu tên, nhưng ai cũng biết nói về ai. Dưới email, phó giám đốc chuyên môn nhắn riêng cho ông: "Anh Đức, anh đang tự đưa mình vào thế khó."

Minh Đức ngồi trong phòng trực, đọc tin nhắn rồi tắt màn hình. Ngoài cửa kính, con gái người thợ hồ cúi đầu cảm ơn từng điều dưỡng. Cô bé mười sáu tuổi, tóc buộc vội, mắt đỏ nhưng sáng. Cô đem cho ông một hộp cơm chay. "Mẹ con nói nhà con không có gì, bác sĩ ăn tạm." Ông nhận hộp cơm. Cơm nguội, đậu hũ mặn, rau luộc nhạt, nhưng đó là bữa ăn khiến ông thấy mình còn là bác sĩ.

Đêm ấy, Khải vào phòng trực. Cậu đứng rất lâu mới nói: "Thầy đừng chống nữa. Thầy không thay được cả hệ thống đâu."

Minh Đức hỏi: "Vậy em chọn để hệ thống thay em?"

Khải cúi mặt. "Em còn gia đình, còn khoản vay mua nhà."

"Tôi không trách em sợ. Tôi chỉ mong em đừng gọi nỗi sợ là thực tế rồi bắt bệnh nhân trả giá."

Khải rời đi, vai nặng. Minh Đức nhìn theo, lòng đau. Một bệnh viện có thể làm hỏng người bằng nhiều cách. Không cần biến họ thành ác quỷ. Chỉ cần bắt họ quen với một cái sai nhỏ mỗi ngày, rồi trả lương đúng hạn, rồi gọi đó là trưởng thành.
"""),
    5784: ("Chương 3: Nộp Đơn Nghỉ - Và Lập Viện", """
Quyết định ép nghỉ đến trong một cuộc họp kín. Hải đặt trước mặt Minh Đức tập biên bản dày: gây mất đoàn kết nội bộ, không tuân thủ quy trình tài chính, phát ngôn ảnh hưởng hình ảnh bệnh viện, tạo áp lực tâm lý cho đồng nghiệp. Mỗi dòng được viết bằng thứ ngôn ngữ hành chính sạch sẽ đến mức không còn thấy máu người bệnh ở đâu.

"Anh ký đơn xin nghỉ vì lý do sức khỏe, bệnh viện sẽ tổ chức chia tay tử tế," Hải nói. "Nếu không, hội đồng kỷ luật sẽ làm đúng quy trình."

Minh Đức lật từng trang. Có chữ ký của những người từng cùng ông thức trắng đêm mổ cấp cứu. Có ý kiến của phó khoa từng được ông cứu lỗi chuyên môn. Có bản tường trình của Khải, viết rằng "thầy Đức thường xuyên gây áp lực đạo đức lên bác sĩ trẻ". Ông dừng ở dòng đó lâu nhất, không phải vì bất ngờ, mà vì đau theo một cách rất yên lặng.

"Tôi ký đơn nghỉ," ông nói.

Hải hơi ngả người, như vừa thắng một ván cờ. "Anh thông minh đấy."

"Nhưng lý do là tôi không còn phù hợp với một nơi coi phong bì là văn hóa."

Nụ cười của Hải tắt. "Anh nên cẩn thận lời nói."

Minh Đức lấy bút, viết đơn ngay trên bàn. Chữ ông đều, không run. Mười lăm năm ở BV Nhân Dân kết thúc trong bảy dòng. Khi ông ra khỏi phòng, trời đổ mưa. Dọc hành lang, vài người tránh mắt, vài người nhìn theo. Điều dưỡng trưởng Lan, người đã làm với ông từ ngày ông mới về khoa, bước tới đưa ông một túi nhỏ. Bên trong là ống nghe cũ của ông để quên ở phòng mổ.

"Thầy đi đâu?" Lan hỏi.

"Chưa biết."

"Nếu thầy mở phòng khám, em theo."

Ông nhìn chị. "Lương thấp lắm."

"Ở đây lương có cao đâu. Chỉ là mình trả bằng thứ khác."

Câu nói ấy đi theo ông suốt buổi chiều. Ông về căn nhà nhỏ, đặt đơn nghỉ lên bàn ăn. Vợ ông mất đã lâu, con gái đang học y năm cuối ở Sài Gòn. Căn nhà im lặng. Ông mở tủ, lấy những cuốn sổ ghi chép ca bệnh cũ, những đề án cải tiến quy trình từng bị bệnh viện bỏ xó, những bản phác thảo về một mô hình viện ngoại khoa nhỏ: giá minh bạch, không phong bì, quỹ hỗ trợ bệnh nhân nghèo, công khai chỉ định, bác sĩ nhận lương tử tế để không phải nhận tay dưới.

Ý tưởng ấy ông từng viết cho vui trong những đêm trực dài. Bây giờ nó trở thành lối ra duy nhất không làm ông xấu hổ. Ông gọi điện cho con gái, Minh An. Cô nghe xong im rất lâu, rồi hỏi: "Ba có sợ không?"

"Có."

"Vậy sao vẫn làm?"

"Vì nếu ba nghỉ rồi im, họ sẽ nói người sạch chỉ tồn tại được khi còn ghế. Ba muốn thử sống sạch cả khi mất ghế."

Ba ngày sau, căn phòng khách nhà ông biến thành văn phòng tạm. Lan đến với một xấp hồ sơ nhân sự. Một kỹ thuật viên gây mê nghỉ hưu nhận lời hỗ trợ. Bác sĩ Hòa, chuyên chẩn đoán hình ảnh, góp máy siêu âm cũ. Một doanh nhân từng được Minh Đức mổ miễn phí cho mẹ đề nghị cho vay vốn không lãi, nhưng ông chỉ nhận phần đủ thuê mặt bằng và sửa chữa, ghi giấy nợ rõ ràng.

Họ thuê một tòa nhà ba tầng cũ gần bến xe tỉnh. Tường bong, điện yếu, thang máy hay kẹt. Lan nhìn quanh, thở dài: "Bệnh nhân có tin không thầy?"

Minh Đức đặt tay lên khung cửa. "Nếu mình làm đúng, họ sẽ tin chậm. Chậm cũng được."

Tên viện được chọn rất đơn giản: Viện Ngoại khoa Minh Đức. Không phải vì ông muốn đặt tên mình lên bảng hiệu, mà vì Lan nói bệnh nhân đã nhớ cái tên ấy như một lời hứa. Bảng giá được treo ngay cửa. Dòng chữ lớn nhất không phải khẩu hiệu, mà là quy định: "Không nhận phong bì. Mọi khoản thu có hóa đơn. Ai khó khăn xin gặp phòng công tác xã hội."

Ngày treo bảng, một số người đi ngang cười. "Bác sĩ bị đuổi mở viện riêng kìa." Minh Đức nghe thấy, không giận. Một cái cây bị nhổ khỏi mảnh đất thối chưa chắc đã chết. Nếu rễ còn khỏe, nó có thể tìm đất khác. Và ông tin rễ của nghề y không nằm trong ghế trưởng khoa, mà nằm ở khoảnh khắc một người đau đớn tin giao thân thể mình cho bàn tay bác sĩ.
"""),
    5785: ("Chương 4: Bệnh Nhân Đổ Về", """
Ngày đầu khai trương, Viện Minh Đức chỉ có bốn bệnh nhân. Một cụ bà đau thoát vị bẹn, một chú tài xế bị sỏi mật, một cô công nhân bị u tuyến giáp và một người đến hỏi rồi đi vì sợ viện nhỏ. Lan ghi tên từng người cẩn thận, giải thích bảng giá từng khoản. Minh Đức tự khám, tự đọc phim, tự gọi điện hỏi bệnh sau khi người ta về. Đến tối, cả viện ngồi ăn cơm hộp trong phòng thủ thuật chưa kịp lắp điều hòa.

"Bốn người cũng là bắt đầu," Minh Đức nói.

Lan cười: "Em chỉ sợ tháng sau không đủ tiền điện."

Nỗi sợ ấy thật. Viện nhỏ không có ngân sách nhà nước, không có gói dịch vụ bóng bẩy, không có đội truyền thông chuyên nghiệp. Trong khi đó, BV Nhân Dân tung tin rằng Minh Đức bị buộc nghỉ vì vi phạm quy trình, viện riêng chưa đủ năng lực cấp cứu, bệnh nhân đến đó là mạo hiểm. Một số trang địa phương đăng bài lập lờ, ảnh chụp bảng hiệu mới của viện bị chỉnh tối màu như nơi đáng nghi.

Nhưng bệnh nhân có trí nhớ của riêng họ. Người phụ nữ vỡ ruột thừa kể với hàng xóm rằng bác sĩ Đức cứu bà mà không lấy phong bì. Gia đình người thợ hồ kể rằng ông mổ trước, tính tiền sau. Bác bảo vệ cũ của bệnh viện đưa vợ đến khám vì "ở đó người ta nhìn mình như người nhà quê, bác sĩ Đức thì không". Mỗi câu chuyện là một tờ rơi không in, chuyền từ chợ sang bến xe, từ quán nước sang xưởng may.

Tuần thứ ba, phòng chờ không còn trống. Bệnh nhân ngồi trên ghế nhựa, tay cầm số thứ tự, nhìn bảng giá công khai. Có người vẫn lén đưa phong bì vì quen. Lan trả lại, chỉ vào bảng. "Ở đây không nhận. Nếu cô muốn cảm ơn, sau mổ viết đánh giá thật hoặc góp vào quỹ hỗ trợ bệnh nhân nghèo có biên lai." Một bà cụ cười ngượng: "Tui đưa vì sợ không đưa thì bác sĩ lơ." Lan đáp: "Tụi con lập viện này để cô khỏi sợ như vậy."

Ca mổ lớn đầu tiên là một bé trai mười hai tuổi bị viêm ruột thừa biến chứng. Cha mẹ làm công nhân, từng đưa con đến BV Nhân Dân nhưng được khuyên chọn gói dịch vụ để "được mổ nhanh". Họ không đủ tiền, ôm con sang Minh Đức trong tuyệt vọng. Minh Đức khám xong, chỉ định mổ ngay. Khi người cha hỏi phải đóng trước bao nhiêu, ông nói: "Đóng phần viện phí cơ bản. Thiếu thì quỹ hỗ trợ. Đừng để con đau thêm vì người lớn đếm tiền."

Ca mổ thành công. Đêm đó, người cha đứng trước quầy thu ngân, hai tay cầm hóa đơn, khóc như trẻ con. "Lần đầu tiên tôi thấy bệnh viện đưa hóa đơn mà không thấy nhục." Câu ấy được một người nhà khác quay lại, đăng lên mạng. Video không đẹp, âm thanh rè, nhưng thật. Nó lan nhanh hơn mọi bài bôi nhọ.

Hải ở BV Nhân Dân bắt đầu chú ý. Số ca ngoại dịch vụ giảm. Một vài bệnh nhân yêu cầu chuyển hồ sơ sang Viện Minh Đức. Bác sĩ trẻ trong khoa lén xem fanpage của viện nhỏ, thấy bảng giá, thấy quỹ hỗ trợ công khai, thấy ảnh Minh Đức ngồi nói chuyện với bệnh nhân sau mổ. Có người mỉa mai là làm màu. Có người im lặng rất lâu.

Áp lực ở Viện Minh Đức cũng tăng. Bệnh nhân đông, nhân sự ít, Lan phải chia ca sát giờ. Minh Đức đặt ra một quy định: không được vì đông mà cộc với bệnh nhân; không được vì bệnh nhân nghèo mà giải thích qua loa; không được nhận bất cứ thứ gì đưa riêng. Một kỹ thuật viên than: "Thầy khó vậy giữ người sao nổi?" Ông đáp: "Nếu mình dễ với cái sai từ đầu, viện này sẽ thành bản sao nhỏ của nơi tôi vừa rời."

Tối cuối tháng, viện có lãi rất ít. Sau khi trả lương, tiền thuê, thuốc men, còn lại chẳng bao nhiêu. Nhưng cuốn sổ quỹ hỗ trợ có mười bảy khoản đóng góp tự nguyện, từ hai chục nghìn đến năm triệu. Minh Đức nhìn những con số ấy, lòng nhẹ hơn khoản lợi nhuận. Ông biết người ta không chỉ đến vì rẻ. Họ đến vì ở đây, nỗi sợ bị moi tiền được tháo ra ngay từ cửa.
"""),
    5786: ("Chương 5: Giám Đốc Hải Phản Công", """
Nguyễn Văn Hải không quen thua chậm. Ông ta có thể chấp nhận một bác sĩ cũ mở phòng khám nhỏ để sống qua ngày, nhưng không chấp nhận bệnh nhân rời BV Nhân Dân để đến đó như một lời bỏ phiếu. Trong cuộc họp giao ban, Hải đặt báo cáo doanh thu xuống bàn, giọng lạnh: "Một viện tư mới mở không thể tự nhiên kéo bệnh nhân như vậy. Phải có vấn đề về truyền thông bẩn."

Phòng truyền thông được lệnh tung bài. Những dòng tiêu đề xuất hiện: "Viện nhỏ tự xưng không phong bì có thật sự an toàn?", "Bệnh nhân nghèo có đang bị lợi dụng làm hình ảnh?", "Cựu trưởng khoa bị cho nghỉ vì bất tuân nay mở viện riêng". Không bài nào nói dối hoàn toàn, nhưng bài nào cũng bẻ chữ đủ để người đọc nghi ngờ. Ảnh Minh Đức được chọn ở góc mệt mỏi nhất, dưới ánh đèn bệnh viện xanh xao.

Chưa dừng lại, Hải gửi đơn kiến nghị Sở Y tế kiểm tra đột xuất Viện Minh Đức. Đoàn kiểm tra đến vào sáng thứ hai, đúng lúc viện có hai ca mổ. Lan bình tĩnh đưa hồ sơ pháp lý, giấy phép, danh mục thuốc, quy trình vô khuẩn. Minh Đức mời đoàn xem trực tiếp phòng mổ, kho thuốc, khu hồi sức. Ông không sợ kiểm tra. Ông chỉ sợ kiểm tra được dùng như cây gậy.

Một cán bộ trong đoàn hỏi sắc: "Tại sao viện treo khẩu hiệu không nhận phong bì? Anh có hàm ý các cơ sở khác nhận à?"

Minh Đức đáp: "Tôi chỉ nói quy định của viện tôi. Nếu nơi khác cũng không nhận, họ có thể treo giống tôi."

Lan cúi mặt giấu cười. Cán bộ kia đỏ tai, ghi chép mạnh tay hơn. Đoàn kiểm tra không tìm được sai phạm lớn, chỉ nhắc vài lỗi hành chính: nhãn tủ thuốc chưa đồng bộ, khu chờ thiếu bảng chỉ dẫn thoát hiểm phụ, sổ lưu mẫu cần ký đủ hai chữ ký. Minh Đức sửa ngay trong ngày, đăng công khai biên bản kiểm tra và danh sách việc đã khắc phục. Cú đánh của Hải biến thành bằng chứng rằng viện nhỏ đủ minh bạch để mở cửa cho kiểm tra.

Hải chuyển sang cách bẩn hơn. Một người đàn ông đưa mẹ đến Minh Đức khám sỏi mật, cố tình quay clip khi Lan từ chối phong bì. Ông ta gào lên: "Không nhận tiền riêng thì coi thường bệnh nhân hả? Mẹ tôi bệnh nặng mà mấy người làm màu!" Clip được cắt đúng đoạn ông ta khóc, bỏ đoạn Lan giải thích đóng viện phí ở quầy. Trong hai giờ, mạng xã hội đầy bình luận chửi Viện Minh Đức giả tạo.

Bảo vệ muốn kéo người đó ra, nhưng Minh Đức ngăn lại. Ông mời hai mẹ con vào phòng tư vấn, bật camera viện, yêu cầu có nhân chứng. Bà mẹ ngồi im, mặt tái vì đau và xấu hổ. Minh Đức khám cho bà, kết luận cần mổ nhưng chưa cấp cứu, giải thích ba phương án điều trị và chi phí. Người đàn ông vẫn cố khiêu khích: "Ông sợ clip nên mới tử tế chứ gì?"

Minh Đức nhìn thẳng vào camera. "Tôi tử tế vì mẹ anh đang đau. Còn chuyện anh quay gì, sau khi bà ổn, luật sư của viện sẽ làm việc."

Ca mổ của bà diễn ra hai ngày sau, thành công. Chính bà là người xin gặp Lan, đưa ra chiếc điện thoại cũ có tin nhắn từ một số lạ hứa trả tiền nếu con trai bà gây chuyện ở viện. Bà khóc: "Nó túng nợ, nó ngu. Nhưng bác sĩ vẫn mổ cho tôi." Minh Đức không công bố tên bà. Ông chỉ chuyển bằng chứng cho luật sư và cơ quan chức năng.

Tối đó, Khải gọi điện. Giọng cậu thấp, căng. "Thầy, em nghĩ thầy cần biết. Bệnh viện đang gom hồ sơ biến chứng cũ để đổ cho thầy."

Minh Đức nhắm mắt. "Em gọi cho tôi, Hải biết thì sao?"

"Em không biết. Nhưng em không muốn im nữa."

Cú điện thoại ấy không xóa được bản tường trình cũ của Khải, nhưng nó tạo một vết nứt trong bức tường sợ hãi. Minh Đức ghi lại thông tin, cảm ơn rồi tắt máy. Ông biết Hải sẽ còn đánh. Nhưng lần này ông không đứng một mình trong hành lang bệnh viện. Sau lưng ông có Lan, có bệnh nhân, có camera công khai, có sổ sách sạch. Người sạch nếu chỉ im lặng sẽ bị bôi bẩn. Người sạch biết lưu bằng chứng mới có thể đứng vững.
"""),
    5787: ("Chương 6: Bí Mật Của BV Nhân Dân", """
Hồ sơ Khải gửi đến lúc ba giờ sáng. Một file nén đặt tên vô nghĩa, kèm một câu: "Thầy xem rồi quyết định, nhưng đừng để lộ em." Minh Đức ngồi trong phòng làm việc của viện, ngoài cửa sổ trời còn tối. Ông mở từng thư mục. Bên trong là biên bản hội chẩn bị sửa, danh sách thuốc vật tư đội giá, hợp đồng gói dịch vụ với công ty sân sau và những báo cáo biến chứng được chuyển lỗi cho bác sĩ trực thay vì lỗi quy trình.

Điều khiến ông lạnh người nhất là thư mục "ngoại tổng quát". Nhiều ca bệnh ông từng đề nghị mổ sớm bị trì hoãn vì chưa đóng gói dịch vụ. Có ca sau đó biến chứng nặng, hồ sơ được chỉnh lại thời gian chỉ định. Chữ ký điện tử của ông xuất hiện trong một số phiếu mà ông chưa từng ký. Ai đó đã dùng tài khoản trưởng khoa cũ để hợp thức hóa quyết định sau khi ông bị cô lập.

Minh Đức gọi Lan. Chị đến trong vòng hai mươi phút, tóc còn ướt vì rửa mặt vội. Hai người ngồi đọc đến sáng. Lan càng đọc càng run. "Nếu công bố, thầy sẽ đụng cả một đường dây."

"Nếu không công bố, họ sẽ tiếp tục đụng vào bệnh nhân."

Nhưng công bố thế nào là vấn đề. Minh Đức không muốn biến sự thật y khoa thành màn đấu tố cảm tính. Ông liên hệ một luật sư, một nhà báo điều tra từng theo mảng y tế và một bác sĩ kiểm định độc lập. Mỗi người nhận một phần tài liệu, kèm yêu cầu xác minh chéo. Ông không cho phép ai dùng hồ sơ bệnh nhân khi chưa che thông tin cá nhân. Sự thật dù cần phơi bày cũng không được giẫm lên quyền riêng tư của người bệnh.

Trong lúc ấy, Hải tung đòn trước. BV Nhân Dân tổ chức họp báo về "nỗ lực nâng cao y đức", ngầm nói một số cá nhân cũ lợi dụng danh tiếng bệnh viện để kinh doanh riêng. Hải đứng trước logo lớn, giọng đĩnh đạc: "Chúng tôi không dung túng sai phạm, dù người đó từng giữ vị trí cao." Câu nói ấy được nhiều trang đăng lại. Bình luận chia đôi. Có người bênh Minh Đức, có người bắt đầu nghi ngờ.

Minh Đức xem họp báo trong phòng bệnh của một cụ ông sau mổ. Cụ hỏi: "Bác sĩ có tức không?"

"Có."

"Vậy sao bác sĩ còn cười?"

"Vì tức mà mất tay vững thì bệnh nhân thiệt."

Cụ bật cười, rồi nắm tay ông. "Tui không biết giấy tờ ai đúng. Tui chỉ biết ở đây mổ xong không ai hỏi phong bì. Bác sĩ giữ được cái đó là được."

Câu nói của cụ giữ Minh Đức khỏi phản ứng vội. Hai ngày sau, nhà báo điều tra gửi lại bản xác minh đầu tiên: công ty cung cấp vật tư giá cao thuộc về em vợ phó giám đốc; một số bệnh nhân có lời khai về việc bị gợi ý đóng gói dịch vụ; log hệ thống cho thấy tài khoản của Minh Đức được truy cập sau thời điểm ông đã bị thu quyền chuyên môn. Luật sư đề nghị gửi đơn đến Sở Y tế, Thanh tra tỉnh và đồng thời chuẩn bị công bố có kiểm soát.

Khải nhắn thêm một file âm thanh. Trong đó, giọng Hải nói với phòng tài chính: "Đừng để ca nghèo chiếm phòng mổ dịch vụ. Bác sĩ Đức thích làm thánh thì để ông ta ra ngoài làm." Minh Đức nghe đi nghe lại ba lần. Ông không thấy hả hê. Ông thấy buồn. Một bệnh viện được xây để cứu người, nhưng ở đâu đó trong các phòng máy lạnh, bệnh nhân đã bị biến thành hạng ghế.

Đêm trước khi gửi đơn, Minh Đức đứng ở hành lang Viện Minh Đức. Phòng bệnh im, chỉ còn tiếng máy oxy đều đều. Lan bước đến hỏi ông có chắc không. Ông nhìn những người đang ngủ sau cửa kính: bác tài xế, cô công nhân, bé trai mổ ruột thừa, bà mẹ từng bị thuê gây chuyện. "Tôi không chắc mình thắng," ông nói. "Nhưng tôi chắc nếu im, mình thua nghề."
"""),
    5788: ("Chương 7: Mẹ Và Viên Thuốc", """
Giữa lúc cuộc chiến hồ sơ căng nhất, mẹ Minh Đức nhập viện. Bà tám mươi tuổi, tăng huyết áp, suy thận nhẹ, vốn sống cùng em gái ông ở quê. Bà ngã trong nhà tắm, may không gãy xương nhưng tụt huyết áp và rối loạn điện giải. Em gái gọi điện, giọng hoảng. Minh Đức lập tức lái xe về đón mẹ lên Viện Minh Đức.

Bà cụ nằm trên giường bệnh nhỏ, tóc bạc mỏng, tay gầy nổi gân. Thấy con trai mặc blouse, bà mắng ngay: "Mẹ có già chút thôi, đừng làm như mẹ sắp chết." Minh Đức cười, nhưng mắt cay. Bao năm ông cứu mẹ người khác, đến lúc mẹ mình nằm trước mặt, mọi kiến thức y khoa bỗng không đủ che nỗi sợ của một đứa con.

Bà biết chuyện ông bị ép nghỉ qua hàng xóm và báo mạng. Khi chỉ còn hai mẹ con, bà hỏi: "Con có hối hận không?"

"Không."

"Có khổ không?"

"Có."

Bà gật đầu như đã biết. "Ngày xưa ba con làm thầy thuốc ở xã, người ta đem gà, đem gạo đến cảm ơn. Ổng nhận khi biết nhà người ta còn đủ ăn, từ chối khi biết người ta phải vay. Ổng nói thuốc cứu người không được làm người ta nghèo thêm. Con giống ba con ở chỗ cứng đầu."

Minh Đức ngồi cạnh giường, bóc từng viên thuốc bỏ vào hộp chia liều. Mẹ ông nhìn hộp thuốc, bỗng nói: "Nhưng con nhớ, sạch không có nghĩa là phải lạnh. Có người đưa phong bì vì họ xấu, nhưng cũng có người đưa vì họ sợ. Con trả lại thì phải trả sao cho người ta không nhục."

Câu ấy ghim vào ông. Ông nhớ nhiều lần mình trả phong bì quá thẳng, khiến người nhà đỏ mặt trước đám đông. Ông đúng, nhưng có lẽ đôi khi cái đúng của ông sắc quá, làm xước cả người đang hoảng loạn. Buổi chiều, ông gọi Lan vào sửa quy trình: thay vì chỉ treo bảng cấm, viện sẽ có tờ hướng dẫn giải thích vì sao không nhận phong bì, bệnh nhân có thể cảm ơn bằng thư, đánh giá, góp quỹ công khai hoặc đơn giản là quay lại tái khám đúng hẹn.

Mẹ ông nghe xong cười: "Vậy mới là bác sĩ. Không chỉ mổ vết thương trên bụng, còn phải coi vết thương trong lòng người ta."

Hai ngày sau, một cụ bà bán vé số được đưa vào viện vì thoát vị nghẹt. Cụ cứ nắm chặt túi vải, bên trong có một phong bì nhỏ ba trăm nghìn. Con trai cụ mất, cụ sống một mình, nghe hàng xóm nói phải đưa tiền bác sĩ mới được lo. Lan định trả ngay, nhưng Minh Đức ngăn lại. Ông ngồi xuống ngang tầm mắt cụ.

"Má giữ tiền này mua cháo sau mổ. Ở viện con, bác sĩ không nhận riêng. Nếu má vẫn muốn cảm ơn, má viết cho con một câu: mổ xong nhớ đi tái khám."

Cụ bà ngơ ngác rồi bật cười móm mém. "Tui không biết chữ."

"Vậy má hứa miệng."

"Ừ, tui hứa."

Ca mổ diễn ra thuận lợi. Khi cụ xuất viện, bà dúi vào tay Lan một trái ổi chín, nói ngoài chợ người ta cho. Lan nhận, ghi vào sổ quà tặng chung của viện: một trái ổi, trị giá không định giá, dùng chia cho phòng trực. Minh Đức nhìn dòng ghi ấy, lòng nhẹ. Minh bạch không phải biến con người thành máy. Minh bạch là để lòng biết ơn còn nguyên mà không biến thành món nợ ngầm.

Đêm đó, mẹ ông đòi ăn cháo trắng. Bà ăn được nửa chén rồi nói: "Con đánh nhau với người ta, nhớ đừng để mình thành người chỉ biết đánh. Bệnh nhân đến với con vì họ đau, không phải vì họ muốn làm chứng cho con thắng."

Minh Đức nắm tay mẹ. Ngoài kia, hồ sơ chống lại BV Nhân Dân đã sẵn sàng. Nhưng trong căn phòng bệnh nhỏ, ông hiểu cuộc chiến lớn nhất không phải hạ Hải. Cuộc chiến lớn nhất là giữ cho viện của mình không bị hận thù nặn thành một nơi cứng lạnh. Một bệnh viện tử tế phải vừa sạch, vừa ấm.
"""),
    5789: ("Chương 8: BV Nhân Dân Thay Đổi", """
Đơn tố cáo được gửi đi vào sáng thứ hai. Đến chiều, bài điều tra đầu tiên lên trang. Không giật tít quá đà, không gọi ai là ác quỷ, chỉ trình bày chuỗi hợp đồng vật tư, log truy cập hồ sơ, lời khai bệnh nhân và đoạn ghi âm của Hải. Chính sự điềm tĩnh ấy làm nó nặng. Trong vòng một ngày, Sở Y tế công bố thành lập đoàn thanh tra. BV Nhân Dân tạm đình chỉ một số cá nhân liên quan, trong đó có trưởng phòng tài chính và phó giám đốc chuyên môn.

Hải vẫn cố chống. Ông ta trả lời báo rằng tài liệu bị cắt ghép, rằng Minh Đức trả thù cá nhân. Nhưng khi nhà báo công bố bản đối chiếu log hệ thống và hợp đồng công ty sân sau, giọng Hải mất dần sức. Những bệnh nhân từng bị ép gói dịch vụ bắt đầu lên tiếng. Một người mẹ kể con mình phải nằm chờ vì thiếu tiền. Một bác tài xế đưa hóa đơn vật tư cao gấp ba giá thị trường. Cơn giận của cộng đồng không ồn ào ngay lập tức; nó dâng lên như nước triều, chậm nhưng không lùi.

Trong BV Nhân Dân, bác sĩ trẻ hoang mang. Khải bị gọi làm việc vì bị nghi rò rỉ tài liệu. Cậu không khai Minh Đức. Cậu chỉ nói mình cung cấp bằng chứng cho cơ quan có thẩm quyền vì không muốn tiếp tục ký vào hồ sơ sai. Khi rời phòng kiểm điểm, cậu thấy nhiều đồng nghiệp nhìn mình khác. Không hẳn khâm phục, không hẳn ghét. Có lẽ họ đang tự hỏi mình đã im lặng bao lâu.

Minh Đức không ăn mừng. Viện của ông vẫn khám, vẫn mổ, vẫn họp chuyên môn mỗi chiều. Ông cấm nhân viên đăng bài chế giễu BV Nhân Dân. "Nơi đó vẫn có bệnh nhân đang nằm. Vẫn có bác sĩ tốt đang làm việc. Mình đánh sai phạm, không đánh người bệnh." Lan đồng ý, nhưng nói thẳng: "Thầy hiền quá, họ từng muốn chôn thầy." Minh Đức đáp: "Nếu mình dùng cùng cách của họ, mình chỉ đổi vai trong một trò bẩn."

Thanh tra kéo dài ba tuần. Kết luận sơ bộ buộc BV Nhân Dân thay đổi quy trình tài chính, công khai đường dây nóng nhận phản ánh phong bì, kiểm toán hợp đồng vật tư và khôi phục một số chỉ định chuyên môn bị can thiệp sai. Hải bị tạm đình chỉ chức vụ để phục vụ điều tra. Ngày tin ấy phát ra, trước cổng bệnh viện có phóng viên, nhưng bên trong khoa ngoại, điều khiến nhiều người chú ý hơn là tấm bảng mới treo ở quầy: "Không gợi ý, không nhận chi phí ngoài hóa đơn."

Bảng ấy còn mới, chữ in đẹp, nhưng chưa ai biết có sống được không. Minh Đức nhận được ảnh bảng từ Khải. Tin nhắn kèm theo chỉ có ba chữ: "Muộn nhưng cần." Ông nhìn rất lâu rồi nhắn lại: "Giữ nó bằng việc làm, không bằng ảnh."

Một tuần sau, Khải đến Viện Minh Đức. Cậu mặc áo sơ mi, không blouse, đứng ở cửa như học trò phạm lỗi. Lan thấy cậu thì lạnh mặt. Minh Đức mời vào phòng. Khải đặt lên bàn bản tường trình ngày xưa, bản có câu "thầy gây áp lực đạo đức", và nói: "Em xin lỗi. Lúc đó em sợ mất vị trí."

Minh Đức không nhận lời xin lỗi ngay. Ông hỏi: "Bây giờ em còn sợ không?"

"Còn."

"Vậy em định làm gì với nỗi sợ?"

Khải im lặng, rồi nói muốn ở lại BV Nhân Dân để sửa từ bên trong. Cậu không xin sang Viện Minh Đức, không xin thầy che chở. Minh Đức gật đầu. "Nếu em ở lại, em sẽ khó hơn tôi. Vì tôi đi rồi được ghét công khai. Em ở đó, mỗi ngày phải chọn giữa tiện và đúng."

Trước khi về, Khải cúi đầu rất thấp. Minh Đức nhìn học trò, thấy trong lòng mình bớt nặng. Một bệnh viện thay đổi không phải vì một giám đốc ngã xuống, mà vì những người còn lại bắt đầu thấy xấu hổ khi làm sai. Xấu hổ, nếu được dùng đúng, cũng là một loại thuốc.
"""),
    5790: ("Chương 9: Bệnh Viện Minh Đức Ra Đời", """
Sáu tháng sau, Viện Minh Đức không còn là tòa nhà cũ phải chống dột bằng xô nhựa. Nhờ khoản vay được trả đúng hạn và các hợp đồng bảo hiểm minh bạch, viện mở thêm khu hồi sức, phòng chẩn đoán hình ảnh và một quầy công tác xã hội đúng nghĩa. Nhưng Minh Đức từ chối gọi nơi đó là bệnh viện lớn. Ông nói lớn nhỏ không đo bằng số tầng, mà bằng việc bệnh nhân nghèo có dám bước vào hay không.

Ngày khánh thành khu mới, nhiều người đề nghị tổ chức long trọng. Minh Đức chỉ mời nhân viên, một số bệnh nhân cũ và đại diện địa phương. Băng rôn treo trước cửa rất đơn giản: "Minh bạch để người bệnh bớt sợ." Không có múa lân, không có phát biểu dài. Người cắt băng cùng ông là cụ bà bán vé số từng hứa tái khám và bé trai mổ ruột thừa ngày đầu viện đông khách. Cụ bà cầm kéo run run, cười đến mất cả mắt.

Bảng giá mới vẫn treo ở cửa, nhưng bên cạnh có thêm bảng quỹ hỗ trợ cập nhật mỗi tháng. Thu bao nhiêu, chi cho ai, còn lại bao nhiêu, tất cả đều công khai. Có người bảo làm vậy dễ bị soi. Lan đáp: "Mình lập viện này để được soi." Câu nói ấy lan trong nhân viên như một tiêu chuẩn. Ai nhập kho thuốc phải ký hai người. Ai tư vấn gói mổ phải ghi rõ lựa chọn rẻ hơn nếu phù hợp. Ai nhận quà của bệnh nhân phải đưa vào sổ chung.

Không phải mọi thứ đều đẹp. Có bác sĩ mới nghỉ sau hai tháng vì thấy quy định quá gò bó. Có bệnh nhân bực vì không được ưu tiên dù quen người trong viện. Có nhà cung cấp gợi ý chiết khấu ngoài hợp đồng, bị Minh Đức từ chối thì quay sang nói viện khó làm ăn. Mỗi lần như vậy, ông đều họp ngắn, nhắc lại lý do ban đầu. "Chúng ta không chống tiền. Chúng ta chống tiền đi đường tối."

BV Nhân Dân sau biến cố bắt đầu cải tổ. Hải bị miễn nhiệm, một giám đốc mới được bổ nhiệm. Khải gửi cho Minh Đức bản quy trình chống phong bì đang thử nghiệm và xin góp ý. Minh Đức sửa rất kỹ, không giữ thù. Lan nhìn ông đọc tài liệu của nơi từng ép mình nghỉ, lắc đầu: "Thầy đúng là hết thuốc." Ông cười: "Nếu họ sửa thật, bệnh nhân được lợi. Mình đâu lập viện để độc quyền tử tế."

Một buổi chiều, đoàn sinh viên y đến tham quan Viện Minh Đức. Minh An, con gái ông, giờ là bác sĩ nội trú, dẫn nhóm đi qua phòng khám, phòng mổ, quầy tài chính. Một sinh viên hỏi Minh Đức: "Thầy nghĩ bác sĩ nghèo thì có giữ được y đức không?" Câu hỏi thật đến mức cả phòng im.

Ông trả lời chậm: "Bác sĩ cũng cần sống tử tế bằng lương. Một hệ thống bắt bác sĩ nghèo rồi mắng họ tham là hệ thống giả đạo đức. Nhưng nghèo không phải giấy phép để bán nỗi sợ của bệnh nhân. Vì vậy muốn sạch lâu dài, phải xây cơ chế: lương rõ, giá rõ, quỹ rõ, kiểm tra rõ. Đạo đức cá nhân quan trọng, nhưng không đủ nếu đặt nó trong một cái bẫy."

Sinh viên ghi chép rất nhanh. Minh Đức nhìn những gương mặt trẻ và nhớ Khải của ngày xưa, nhớ chính mình lúc mới vào nghề. Ông không muốn các em lớn lên bằng sự vỡ mộng. Ông muốn các em biết nghề y sẽ làm mình mệt, nghèo hơn vài lựa chọn khác, đau lòng hơn rất nhiều nghề khác, nhưng nếu giữ được bàn tay sạch và trái tim ấm, mỗi ca bệnh qua khỏi sẽ trả lại một thứ không đồng tiền nào mua được.

Tối hôm đó, viện nhận một ca cấp cứu từ tai nạn xe máy. Máu chảy nhiều, người nhà hoảng loạn, một người theo thói quen dúi phong bì vào tay bác sĩ trực. Bác sĩ trẻ trả lại, chỉ vào bảng hướng dẫn, giọng rất nhẹ: "Anh giữ tiền mua thuốc. Chúng tôi lo chuyên môn." Minh Đức đứng ở cuối hành lang, nghe câu ấy, không bước vào. Ông biết một nơi thật sự ra đời không phải khi bảng hiệu được treo, mà khi người trong đó tự nói đúng câu cần nói dù không có ông đứng cạnh.
"""),
    5791: ("Chương 10: Ca Trực Đêm", """
Một năm sau ngày rời BV Nhân Dân, Minh Đức vẫn giữ thói quen trực đêm mỗi tuần một lần. Nhân viên bảo ông không cần, viện đã có đội trực đủ. Ông chỉ cười: "Người lập luật phải thỉnh thoảng đứng ở nơi luật chạm vào người bệnh." Đêm trực giúp ông nghe những điều ban ngày bị lịch họp che mất: tiếng người nhà thở dài trước quầy viện phí, tiếng điều dưỡng dỗ bệnh nhân già, tiếng bác sĩ trẻ gọi hội chẩn vì không muốn quyết định một mình.

Đêm ấy mưa lớn. Gần một giờ sáng, xe cấp cứu đưa vào một người đàn ông bị thủng dạ dày, sốc nhẹ. Người nhà là cô con gái khoảng hai mươi tuổi, áo mưa rách, tay ôm túi giấy tờ ướt. Cô run rẩy hỏi: "Bác sĩ ơi, nhà con chỉ còn sáu triệu. Ba con có được mổ không?" Câu hỏi ấy đưa Minh Đức về đúng đêm phong bì đầu tiên, về hành lang BV Nhân Dân, về cậu thanh niên sợ mất mẹ.

Ông đặt tay lên vai cô. "Ba con được mổ vì ông ấy cần mổ. Tiền tính sau, có quỹ hỗ trợ và phương án trả dần. Bây giờ con ký giấy, rồi ngồi xuống uống nước."

Ca mổ kéo dài đến gần bốn giờ. Bác sĩ trực chính là một người trẻ, Minh Đức chỉ đứng hỗ trợ. Khi đường khâu cuối cùng hoàn tất, ông thấy tay mình hơi mỏi. Tuổi tác không nói dối. Nhưng ông cũng thấy trong phòng mổ này không còn mùi sợ hãi của phong bì. Có áp lực chuyên môn, có mồ hôi, có sai số phải kiểm soát, nhưng không có câu hỏi bệnh nhân đã "biết điều" chưa.

Ra khỏi phòng mổ, cô con gái đứng bật dậy. Minh Đức báo ca mổ tạm ổn. Cô khóc, lục túi lấy ra một phong bì đã ướt mép. Có lẽ ai đó ngoài kia vẫn dặn cô phải đưa. Cả hành lang im. Bác sĩ trẻ nhìn Minh Đức, chờ xem ông làm gì. Ông nhận phong bì, mở ra trước mặt cô. Bên trong là sáu triệu đồng, toàn tiền lẻ.

Ông không trả lại ngay. Ông đưa cô đến quầy thu ngân, cùng nhân viên lập phiếu tạm ứng đúng sáu triệu, in hóa đơn, ghi phần còn lại chuyển quỹ hỗ trợ xét duyệt. Sau đó ông nói: "Con không đưa riêng cho bác sĩ. Con đóng viện phí cho ba. Khác nhau nhiều lắm."

Cô gái ôm hóa đơn vào ngực, khóc nấc. Minh Đức không thấy mình chiến thắng. Ông chỉ thấy một mắt xích nhỏ được đặt đúng chỗ. Cùng là tiền từ tay người nhà bệnh nhân, nhưng đi qua ánh sáng thì nó không còn là nỗi nhục.

Gần sáng, ông ngồi ở phòng trực, ăn nửa hộp cơm nguội. Lan đem vào hai quả cam, giống hệt đêm năm nào. "Người nhà bệnh nhân thủng dạ dày gửi quầy chung," chị nói. "Có biên nhận quà tặng rồi." Minh Đức bật cười, bóc cam chia cho bác sĩ trẻ. Vị cam vẫn chua, vẫn làm ông nhớ vì sao mình bắt đầu.

Minh An ghé viện lúc sáu giờ, thấy cha đang ghi hồ sơ. Cô hỏi ông có mệt không. Ông nói có. Cô hỏi có đáng không. Ông nhìn qua cửa kính, nơi cô con gái bệnh nhân ngủ gục trên ghế chờ, tay vẫn nắm hóa đơn. "Đáng," ông đáp. "Vì hôm nay có một người biết rằng bệnh viện không phải cái chợ đêm của nỗi sợ."

Buổi sáng, mặt trời lên sau cơn mưa. Bảng "Không nhận phong bì" trước cửa viện ướt nước nhưng vẫn rõ chữ. Minh Đức đứng nhìn nó một lúc rồi bước vào vòng khám mới. Ông không còn là trưởng khoa của bệnh viện lớn nhất tỉnh. Ông cũng không cần danh xưng ấy nữa. Ông là một bác sĩ trong một viện nhỏ đang học cách sạch và ấm mỗi ngày.

Ngoài sân, một cụ bà xuất viện chắp tay cảm ơn. Minh Đức đỡ tay cụ xuống, nói câu ông đã nói hàng trăm lần nhưng chưa bao giờ thấy cũ: "Cụ về uống thuốc đúng giờ, tái khám đúng hẹn. Đó là cách cảm ơn tốt nhất." Cụ cười, con cháu cười. Tiếng cười ấy đi qua hành lang sáng, chạm vào những bức tường còn mới sơn, rồi ở lại trong lòng những người trực đêm.

Minh Đức quay về phòng mổ khi chuông cấp cứu lại reo. Nghề y không cho ai kết thúc bằng một bài diễn văn đẹp. Nó gọi người ta trở lại bằng tiếng bánh xe băng ca, bằng hơi thở yếu, bằng bàn tay người nhà níu áo blouse. Và ông, sau tất cả những lần bị cô lập, bị bôi nhọ, bị ép nghỉ, vẫn bước nhanh về phía tiếng gọi ấy. Vì ông hiểu phẩm giá của bác sĩ không nằm ở chiếc phong bì bị từ chối một lần, mà ở việc ngày nào cũng từ chối biến nỗi đau của người khác thành cơ hội cho mình.
"""),
}


def main():
    editor.upload_helper()
    try:
        editor.update_story_meta(STORY_ID, title=TITLE, intro=INTRO)
        updates = []
        for chapter_id, (title, content) in CHAPTERS.items():
            editor.update_chapter(chapter_id, title, content.strip())
            updates.append({"chapter_id": chapter_id, "title": title})
    finally:
        editor.remove_helper()

    workbook = ROOT / "danh_sach_truyen_doctieuthuyet.xlsx"
    wb = openpyxl.load_workbook(workbook)
    ws = wb.active
    for row in range(5, ws.max_row + 1):
        if ws.cell(row, 2).value == STORY_ID:
            ws.cell(row, 3).value = TITLE
            ws.cell(row, 12).value = INTRO
            ws.cell(row, 13).value = (
                "Đã đọc live và viết lại toàn bộ 10 chương từ bản quá ngắn; tăng xung đột phong bì, "
                "bị cô lập, lập viện riêng, phản công bằng hồ sơ và kết ca trực đêm."
            )
            ws.cell(row, 14).value = "☑️ Đã sửa"
            break
    wb.save(workbook)

    out = ROOT / "scratch" / "story_5780_fix_result.json"
    out.write_text(json.dumps(updates, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(updates, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
