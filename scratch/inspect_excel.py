import openpyxl

wb = openpyxl.load_workbook("danh_sach_truyen_doctieuthuyet.xlsx")
ws = wb.active

# Check first 5 rows to understand structure
for row in range(1, 6):
    vals = []
    for col in range(1, min(ws.max_column + 1, 15)):
        v = ws.cell(row=row, column=col).value
        if v is not None:
            vals.append(f"Col{col}={v}")
    print(f"Row {row}: {' | '.join(vals)}")

# Check row 192 (story 1933)
print(f"\nRow 192 (story 1933):")
for col in range(1, min(ws.max_column + 1, 15)):
    v = ws.cell(row=192, column=col).value
    if v is not None:
        print(f"  Col{col}: {str(v)[:100]}")

print(f"\nTotal rows: {ws.max_row}, Total columns: {ws.max_column}")
