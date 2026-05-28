#!/usr/bin/env python3
import importlib.util
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
EDITOR_PATH = ROOT / "scratch" / "novel_editor.py"
BACKUP = ROOT / "scratch" / "story_5803_before_fix.json"
XLSX = ROOT / "danh_sach_truyen_doctieuthuyet.xlsx"
OUT = ROOT / "scratch" / "story_5803_editorial_fix_result.json"

spec = importlib.util.spec_from_file_location("novel_editor", EDITOR_PATH)
editor = importlib.util.module_from_spec(spec)
spec.loader.exec_module(editor)

STORY_ID = 5803
TITLE = "Bị Bán Qua Biên Giới, Tôi Thành Luật Sư Bảo Vệ Trẻ Em"

INTRO = """
<p><strong>"Đi với chị xuống thành phố ba ngày thôi, có áo mới, có tiền gửi về cho mẹ."</strong> Năm mười hai tuổi, Vàng Thị Mai tin câu nói ấy và bước lên chiếc xe máy rời bản Lao Chải trong một buổi chiều sương phủ.</p>
<p>Ba ngày sau, cô bé người H'Mông bị khóa trong một căn nhà xa lạ bên kia biên giới. Mười tháng bị giam, bỏ đói, đe dọa và tước hết giấy tờ đã dạy Mai một điều đau đớn: trẻ em nghèo không chỉ cần lòng thương, các em cần người biết luật đứng ra bảo vệ.</p>
<p>Mai trốn thoát sau ba ngày băng rừng về đồn biên phòng, trở về nhà trong ánh mắt thương hại và định kiến. Nhưng cô không chết trong cái bẫy ấy. Cô đi học lại từ con số không, thi vào Đại học Luật, rồi trở thành luật sư chuyên cứu phụ nữ và trẻ em bị buôn bán.</p>
<p>Mười lăm năm sau, khi đường dây từng bán cô quay lại dụ một bé gái khác, Vàng Thị Mai không còn là nạn nhân im lặng. Lần này, cô mang hồ sơ, lời khai, tọa độ điện thoại và cả bản án ra trước ánh sáng.</p>
""".strip()


def p(text):
    return f"<p>{text}</p>"


EXPANSIONS = {
    5805: [
        "Buổi chiều hôm đó, sương xuống sớm hơn mọi ngày. Bản Lao Chải nằm co trong thung lũng, mái nhà gỗ ướt lạnh, khói bếp bay thấp như không muốn rời khỏi núi. Vàng Thị Mai ngồi trước hiên, vá lại quai dép bằng sợi dây gai mẹ tước từ bao tải cũ. Cô bé mười hai tuổi, gầy, mắt đen, tóc tết hai bên, cả đời mới xuống chợ huyện ba lần.",
        "Người phụ nữ lạ đến bản lúc gần năm giờ. Chị ta mặc áo khoác đỏ, đi giày trắng, nói tiếng H'Mông không trôi lắm nhưng đủ khiến người lớn bớt đề phòng. Chị ta bảo mình quen một chủ quán ăn dưới thành phố Lào Cai, cần mấy bé gái nhanh nhẹn phụ rửa bát ba ngày hội chợ, mỗi bé được trả tám trăm nghìn. Với nhà Mai, tám trăm nghìn là tiền muối, dầu, vở, thuốc ho cho em trai cả tháng.",
        "Mẹ Mai lưỡng lự. Bố cô đang đi phụ hồ xa, nhà chỉ còn mẹ, bà ngoại và ba đứa nhỏ. Người phụ nữ mở túi, lấy ra một gói kẹo, một chiếc kẹp tóc nhựa và hai tờ tiền mới. \"Em nó đi với chị, tối thứ ba chị đưa về. Có số điện thoại đây, có gì cứ gọi.\" Số điện thoại được viết trên mảnh giấy xanh, nét bút tròn trịa đến mức nhìn như thật.",
        "Bà ngoại Mai ngồi bên bếp, mắt nheo lại. Bà không thích người lạ cười quá nhiều. Bà kéo Mai lại, buộc vào cổ tay cô một sợi chỉ đỏ nhỏ. \"Nếu thấy đường lạ quá thì quay về,\" bà nói. \"Con hổ con có thể chưa biết săn, nhưng phải biết nhớ lối về hang.\" Mai cười, không hiểu vì sao bà nghiêm mặt đến vậy.",
        "Chiếc xe máy chở Mai rời bản khi trời chuyển màu chì. Ban đầu đường vẫn quen: con dốc đá, cây sa mộc cong, đám trẻ đuổi nhau bên bờ ruộng bậc thang. Rồi xe rẽ sang một nhánh đường cô chưa từng đi. Người phụ nữ bảo do đường chính sạt, phải vòng. Mai ôm túi áo cũ, trong túi có hai bộ quần áo, gói xôi mẹ nắm và chiếc kẹp tóc nhựa mà cô chưa dám cài.",
        "Tối xuống, họ không vào thành phố. Họ dừng ở một căn nhà tôn cạnh bìa rừng. Có hai người đàn ông hút thuốc trước cửa. Một người hỏi bằng thứ tiếng Mai không hiểu. Người phụ nữ đáp nhanh, rồi đẩy Mai vào trong. Lần đầu tiên, Mai thấy bụng mình lạnh đi. Cô hỏi: \"Chị ơi, bao giờ đến quán ăn?\" Không ai trả lời.",
        "Đêm ấy, Mai ngủ ngồi cạnh một cô gái lớn hơn chừng mười sáu tuổi. Cô gái ấy nhìn Mai rất lâu rồi thì thầm bằng tiếng H'Mông: \"Em bị lừa rồi.\" Mai không tin. Cô nắm chặt sợi chỉ đỏ trên cổ tay, tự nhủ sáng mai chị áo đỏ sẽ đưa mình xuống quán ăn. Nhưng sáng hôm sau, người phụ nữ biến mất. Hai người đàn ông đưa các cô lên một chiếc xe khác, cửa kính dán đen, mùi thuốc lá và xăng cũ làm Mai buồn nôn.",
        "Đến ngày thứ ba, Mai không còn biết mình đang ở đâu. Biển hiệu ngoài đường đổi chữ. Giọng nói quanh cô đổi âm. Cánh cửa căn nhà cuối cùng đóng sập lại từ bên ngoài, tiếng khóa kim loại vang lên rất khô. Một người đàn ông ném xuống nền túi quần áo của Mai, rồi chỉ vào góc phòng. Trong đó có ba cô gái khác, mắt sưng, môi khô, đều mặc đồ cũ và đều không dám khóc to.",
        "Mai ngồi xuống sàn, lưng áp vào tường lạnh. Cô nhìn cổ tay. Sợi chỉ đỏ của bà vẫn còn, nhưng đường về bản đã biến mất sau quá nhiều khúc rẽ. Lúc ấy cô mới hiểu: mình không đi làm thuê. Mình bị bán. Ý nghĩ ấy quá lớn với một đứa trẻ mười hai tuổi. Nó không biến thành tiếng hét. Nó nằm trong ngực cô như một hòn đá, nặng đến mức cô không thở nổi.",
    ],
    5806: [
        "Căn nhà nơi Mai bị nhốt có hai lớp cửa. Cửa gỗ bên trong, cửa sắt bên ngoài, cửa sổ hàn thanh ngang. Ban ngày, ánh sáng lọt qua khe mái, vẽ lên nền xi măng những vệt nhỏ như sợi chỉ. Bốn cô gái trong phòng học cách đo thời gian bằng những vệt sáng ấy, vì không ai cho họ đồng hồ.",
        "Người trông giữ căn nhà là một phụ nữ trung niên mặt lạnh. Bà ta mang cơm hai lần mỗi ngày: cơm nguội, rau luộc nhạt, thỉnh thoảng có vài miếng đậu phụ. Nước được đựng trong can nhựa, có mùi lạ. Nếu ai hỏi khi nào được về, bà ta gõ mạnh chiếc muôi vào song sắt. \"Về? Tiền đã trả rồi. Muốn về thì bảo nhà mày đem tiền sang chuộc.\"",
        "Mai không biết nhà mình phải chuộc bao nhiêu. Cô chỉ biết mẹ không có tiền. Mỗi tối, cô nằm co trên tấm chiếu rách, nhớ tiếng gió luồn qua vách gỗ ở Lao Chải, nhớ tiếng bà ngoại kể chuyện con hổ H'Mông mắc bẫy nhưng không cúi đầu. Hồi ở nhà, cô từng nghe chuyện ấy đến chán. Trong căn phòng khóa ngoài, câu chuyện trở thành thứ duy nhất giữ cô khỏi tan ra.",
        "Ba cô gái cùng phòng đều có câu chuyện giống nhau. Súa bị rủ đi bán hàng ở chợ biên giới. Lử bị người quen của dì hứa cho làm giúp việc. Chị Mỷ lớn nhất, mười bảy tuổi, bị lừa bằng lời hứa lấy chồng giàu rồi gửi tiền về sửa nhà. Không ai ngu. Không ai tham. Họ chỉ nghèo, ít thông tin và tin nhầm người nói cùng tiếng mẹ đẻ.",
        "Mai là nhỏ nhất nên được các chị nhường chỗ nằm sát tường, nơi ít gió lạnh hơn. Đổi lại, cô học cách quan sát. Người phụ nữ mặt lạnh giấu chìa khóa ở túi áo bên trái. Gã đàn ông canh đêm thường uống rượu sau mười giờ. Cửa sổ phòng bếp có một thanh sắt bị gỉ ở chân. Mỗi chi tiết nhỏ được Mai giữ trong đầu như giữ hạt ngô giống qua mùa đói.",
        "Có lần, một người đàn ông lạ đến xem mặt các cô. Mai bị kéo ra sân. Cô không hiểu hết lời họ nói, nhưng hiểu ánh mắt như nhìn món hàng. Cô cắn chặt môi đến bật máu. Chị Mỷ sau đó ôm cô rất lâu, nói: \"Em phải sống. Sống rồi mới có ngày kể lại.\" Câu ấy ở lại với Mai suốt nhiều năm. Không phải lời an ủi, mà là mệnh lệnh.",
        "Mười tháng không trôi thẳng. Nó vỡ thành từng mảnh: mảnh là ngày Súa sốt cao nhưng không được đưa đi khám; mảnh là đêm Lử khóc không ra tiếng vì nhớ em trai; mảnh là buổi Mai nghe tiếng trẻ con ngoài ngõ cười và nhận ra mình cũng từng cười như thế. Cô học cách không tiêu hết nước mắt trong một đêm, vì sáng hôm sau vẫn phải mở mắt.",
        "Chiếc thìa nhôm trong bát cơm trở thành bí mật của cô. Mỗi lần ăn xong, Mai lén mài cạnh thìa vào nền xi măng ở góc khuất. Ban đầu chỉ để bớt sợ. Sau đó, cạnh thìa sắc lên thật. Chị Mỷ phát hiện, không mắng. Chị chỉ lấy thân người che cho cô mỗi lần người trông giữ mở cửa.",
        "Đêm mưa cuối tháng mười, tiếng sấm làm cả căn nhà rung nhẹ. Gã canh cửa uống say hơn mọi ngày. Người phụ nữ mặt lạnh đi vắng từ chiều. Mai ngồi cạnh cửa sổ bếp, tay cầm chiếc thìa mài mòn, tim đập đến đau. Chị Mỷ đặt tay lên vai cô. \"Nếu chạy được, chạy thẳng về phía núi,\" chị thì thầm. \"Gặp ai mặc quân phục Việt Nam thì lao vào, đừng sợ.\"",
        "Mai nhìn ba cô gái trong phòng. Cô muốn kéo tất cả đi. Nhưng thanh cửa hẹp, gã canh ở sân, một đứa trẻ mười hai tuổi không thể cứu cả thế giới trong một đêm. Chị Mỷ như đọc được mắt cô, lắc đầu. \"Em nhỏ nhất. Em lọt được. Em về được thì kể cho người ta biết chỗ này.\"",
        "Lúc ấy Mai hiểu một điều tàn nhẫn: đôi khi sống sót cũng là một món nợ. Cô nắm sợi chỉ đỏ trên cổ tay, cắm chiếc thìa vào khe gỗ đã mục, bắt đầu nạy.",
    ],
    5807: [
        "Thanh gỗ bật ra không kêu lớn, nhưng trong tai Mai, âm thanh ấy như sấm. Cô nín thở, chui qua ô cửa hẹp, da vai bị cạnh gỗ cào rách. Ngoài sân, mưa rơi thành màn trắng. Gã canh cửa gục bên bàn, chai rượu lăn dưới chân. Mai không dám nhìn lâu. Cô chạy.",
        "Đường làng bên kia biên giới tối và lạ. Mai không biết phía nào là Việt Nam, chỉ nhớ lời chị Mỷ: chạy về phía núi. Cô men theo bờ ruộng, chân trần lún trong bùn lạnh. Mỗi tiếng chó sủa làm cô muốn nằm rạp xuống. Mỗi ánh đèn xe từ xa làm cô trốn vào bụi cỏ, ôm miệng để tiếng thở không bật ra.",
        "Đêm đầu tiên, Mai ngủ dưới gốc cây, đúng hơn là ngất đi vì kiệt sức. Sáng dậy, cổ họng khô rát, bụng đau vì đói. Cô uống nước đọng trên lá, nhai một mẩu xôi khô còn sót trong túi áo từ ngày rời bản, đã mốc và cứng như đá. Cô nhổ ra, rồi lại nhặt lên cắn tiếp. Sống quan trọng hơn sạch.",
        "Đến trưa, cô gặp một con đường đất có vết bánh xe. Mai biết đi theo đường dễ bị bắt lại, nhưng đi trong rừng thì mất hướng. Cô chọn cách đi song song, cách đường chừng hai mươi bước, vừa đi vừa nhìn mặt trời. Bà ngoại từng dạy nếu lạc trên nương, hãy nhớ hướng nắng và tiếng suối. Bây giờ những bài học tưởng nhỏ ấy thành bản đồ.",
        "Chiều ngày thứ nhất, có tiếng xe máy phía sau. Mai lao xuống rãnh, lăn qua bụi gai. Gai cào vào cẳng chân, máu rịn ra. Hai người đàn ông chạy xe qua, nói thứ tiếng cô không hiểu. Một người quay đầu nhìn về phía bụi. Mai cắn vào tay mình để khỏi bật khóc. Xe đi xa rồi, cô vẫn nằm thêm rất lâu, đến khi kiến bò lên cổ mới dám nhúc nhích.",
        "Đêm thứ hai lạnh hơn. Mưa tạnh, sương phủ dày. Mai nghe tiếng nước chảy và lần theo, vì suối thường dẫn về bản. Cô rửa vết thương ở chân, nước lạnh buốt như kim. Trong bóng tối, cô thấy khuôn mặt mình in lờ mờ trên mặt nước: gầy, bẩn, mắt trũng. Cô suýt không nhận ra cô bé từng ngồi vá dép trước hiên nhà.",
        "Cơn sốt đến vào nửa đêm. Mai nằm giữa đám lá ướt, người lúc nóng lúc lạnh. Cô mơ thấy mẹ gọi ăn cơm, mơ thấy bà ngoại kể chuyện con hổ bị bẫy. Trong mơ, con hổ không gầm. Nó chỉ nhìn cô, mắt vàng, kiên nhẫn. Mai tỉnh dậy vì chính tiếng mình gọi bà.",
        "Sáng ngày thứ ba, cô thấy một cột mốc đá. Chữ trên đó không hoàn toàn quen, nhưng bên kia triền dốc có lá cờ đỏ rất nhỏ bay trên mái nhà thấp. Mai không chắc đó là đồn biên phòng Việt Nam hay một căn nhà khác. Cô chỉ biết nếu không đi tiếp, cô sẽ chết trong rừng.",
        "Cô bước xuống dốc bằng đôi chân không còn cảm giác. Đến gần cổng, một chiến sĩ biên phòng nhìn thấy cô trước. Anh chạy ra, tay giơ lên ra hiệu dừng vì sợ cô hoảng. Mai nhìn màu áo xanh, nhìn phù hiệu, nhìn lá cờ. Mọi thứ trong ngực cô vỡ ra.",
        "Cô khuỵu xuống trước cổng đồn Bát Xát, hai tay vẫn nắm sợi chỉ đỏ đã sờn. Người chiến sĩ hỏi cô bằng tiếng Kinh, cô không trả lời được. Mãi đến khi một cán bộ biết tiếng H'Mông chạy ra, Mai mới nói được câu đầu tiên sau ba ngày trốn chạy: \"Em muốn về nhà.\"",
        "Câu nói ấy không lớn. Nó khản đặc, đứt quãng, gần như bị gió cuốn mất. Nhưng với Mai, nó là cánh cửa đầu tiên mở ra sau mười tháng khóa ngoài.",
    ],
    5808: [
        "Ngày Mai trở về Lao Chải, cả bản kéo ra đầu dốc. Mẹ cô chạy trước, ngã dúi một lần rồi lại đứng dậy. Bà ôm Mai chặt đến mức vết thương trên vai cô đau nhói, nhưng Mai không kêu. Cô áp mặt vào áo mẹ, ngửi mùi khói bếp quen thuộc và hiểu mình thật sự chưa chết.",
        "Bố Mai từ công trình dưới xuôi về trong đêm. Ông ngồi bên giường nhìn con gái ngủ, bàn tay to thô đặt trên mép chăn nhưng không dám chạm mạnh. Ông cứ lặp đi lặp lại: \"Bố xin lỗi.\" Mai nghe thấy, nhưng cô không biết trả lời thế nào. Một đứa trẻ mười hai tuổi không thể gánh tội lỗi của người lớn, cũng không biết trả lại nó cho ai.",
        "Những ngày đầu, cán bộ xã, biên phòng và công an huyện lên lấy lời khai. Mỗi câu hỏi đều cần thiết, nhưng với Mai, mỗi câu trả lời giống như mở lại một cánh cửa khóa. Cô kể tên người phụ nữ áo đỏ, căn nhà hai lớp cửa, chị Mỷ, Súa, Lử, đường trốn qua suối. Có lúc cô run đến không cầm được cốc nước. Mẹ cô muốn dừng, nhưng Mai lắc đầu. Chị Mỷ đã bảo cô về được thì kể.",
        "Bản làng ban đầu thương cô. Người đem trứng, người đem gạo, người đem áo ấm. Nhưng sau thương hại là những lời thì thầm. <em>Con bé bị bán rồi.</em> <em>Sau này ai dám lấy?</em> <em>Không biết bên kia nó bị làm sao.</em> Mai nghe thấy khi đi qua bờ rào, khi lấy nước ở đầu bản, khi ngồi trong lớp học tạm. Những lời ấy không đánh vào da, nhưng làm cô thấy mình bẩn dù không làm gì sai.",
        "Cô im lặng ba tháng. Không phải im vì ngoan, mà vì trong người như có một hòn đá chặn ngang cổ họng. Đêm nào Mai cũng giật mình khi nghe tiếng xe máy. Cửa nhà đóng rồi cô vẫn kiểm tra chốt. Có bữa mẹ nấu cơm, mùi khói bếp làm cô nhớ căn nhà bị nhốt, cô nôn hết ra sân.",
        "Cô giáo Thảo, giáo viên cấp hai dưới huyện, được cử lên hỗ trợ Mai học lại. Cô không ép Mai kể chuyện. Cô chỉ mang sách, vở, bút chì màu và một tờ bản đồ Việt Nam. Buổi đầu, Mai không nói câu nào. Buổi thứ hai, cô giáo chỉ vào tỉnh Lào Cai trên bản đồ, rồi chỉ xuống Hà Nội, Huế, Cần Thơ. \"Đất nước mình dài lắm,\" cô nói. \"Con đã đi qua chỗ tối, nhưng đời con không chỉ có chỗ tối đó.\"",
        "Ngày thứ chín mươi, Mai mở miệng xin đi học. Câu nói làm mẹ cô khóc. Bà sợ con gái xuống huyện sẽ bị trêu, bị nhìn, bị nhắc lại chuyện cũ. Bố cô cũng sợ. Nhưng bà ngoại chống gậy ra hiên, nhìn Mai rất lâu rồi nói: \"Hổ ra khỏi bẫy mà cứ nằm trong hang thì người đặt bẫy thắng.\"",
        "Hôm Mai trở lại lớp, cả phòng học im lặng. Có bạn nhìn cô tò mò, có bạn tránh mắt. Mai ngồi bàn cuối, đặt vở lên bàn, bàn tay run dưới gầm ghế. Giờ ra chơi, một nhóm con gái thì thầm rồi cười. Mai nghe thấy tên mình. Cô cúi xuống, móng tay bấm vào mép vở.",
        "Cô giáo Thảo bước vào, không mắng ai. Cô viết lên bảng hai chữ: <strong>nạn nhân</strong>. Rồi cô quay xuống lớp: \"Người bị hại không có lỗi vì kẻ xấu chọn làm hại mình. Nếu các em không hiểu điều đó, các em đang đứng về phía kẻ xấu.\"",
        "Lớp học im phăng phắc. Mai không ngẩng lên, nhưng nước mắt rơi xuống trang vở. Lần đầu tiên sau ngày trở về, có một người lớn nói rõ ràng rằng lỗi không nằm ở cô. Câu ấy mở một khe nhỏ trong hòn đá chặn cổ họng.",
        "Tối hôm đó, Mai viết dòng đầu tiên trong cuốn vở mới: <em>Mình sẽ học để biết cách nói cho đúng, để không ai bảo trẻ bị hại là xấu nữa.</em>",
    ],
    5809: [
        "Con đường từ Lao Chải xuống trường huyện dài mười cây số, mùa nắng bụi đỏ bám đến đầu gối, mùa mưa bùn dính nặng như kéo chân lại. Mai dậy lúc bốn giờ sáng, ăn nắm cơm nguội với muối vừng, đeo cặp vải mẹ khâu, đi khi bản còn tối. Bà ngoại thường chống gậy đứng ở cửa, chờ đến khi bóng cô khuất sau rặng sa mộc mới vào nhà.",
        "Ngày đầu tiên đi học đều lại, Mai đau chân đến phát khóc. Vết sẹo do gai rừng ở bắp chân vẫn căng mỗi khi leo dốc. Nhưng đau chân dễ chịu hơn đau vì ngồi yên. Mỗi bước xuống huyện giống như cô tự kéo mình ra khỏi căn phòng khóa ngoài, lần này bằng sách vở thay vì chiếc thìa mài sắc.",
        "Bạn bè không phải ai cũng ác. Có người lén để vào ngăn bàn cô một cái bánh ngô. Có người cho mượn compa. Nhưng cũng có những câu nói đâm thẳng: \"Nó từng bị bán đấy.\" \"Chắc bên kia giàu lắm nên mới về.\" \"Sau này đừng ngồi gần nó.\" Mai nghe hết. Cô không đánh nhau. Cô ghi vào một góc vở: <em>Lời nói cũng có thể là dây trói.</em>",
        "Cô học chậm hơn các bạn vì mất gần một năm. Toán làm cô sợ, tiếng Kinh trong sách giáo khoa nhiều từ cô không hiểu. Mỗi tối, cô giáo Thảo cho cô mượn từ điển cũ, dạy cách gạch chân từ khó. Mai chép mỗi từ mười lần: quyền, nghĩa vụ, bảo vệ, chứng cứ, tố cáo. Khi chưa hiểu luật là gì, cô đã thích những chữ nghe như có thể dựng thành hàng rào.",
        "Năm lớp tám, có một buổi tuyên truyền phòng chống buôn người ở trường. Cán bộ đọc tài liệu rất nhanh, học sinh ngáp, vài bạn cười khi nghe đến chuyện bị lừa lấy chồng xa. Mai ngồi dưới, tay lạnh đi. Cô biết những tờ rơi kia không đủ. Kẻ xấu không đến với mặt dữ và dây thừng. Chúng đến bằng tiếng mẹ đẻ, kẹo ngọt, tiền mới và lời hứa gửi về cho mẹ.",
        "Cuối buổi, cô xin phát biểu. Cả hội trường quay lại nhìn. Mai đứng dậy, váy H'Mông cũ chạm đầu gối, giọng ban đầu rất nhỏ. Cô kể rằng người lừa cô nói tiếng H'Mông, biết tên mẹ cô, biết nhà cô nghèo. Cô kể rằng số điện thoại để lại là số giả. Cô kể rằng nếu một đứa trẻ biến mất quá một ngày, đừng chờ thêm vì xấu hổ.",
        "Không ai cười nữa. Cán bộ tuyên truyền đặt tập giấy xuống. Cô giáo Thảo đứng cuối hội trường, mắt đỏ. Sau buổi đó, ba bạn gái đến hỏi Mai nếu có người rủ đi làm ở chợ biên giới thì phải làm sao. Mai không biết trả lời đủ, chỉ nói: \"Đừng đi một mình. Báo cô giáo. Báo biên phòng. Gọi người lớn trước khi lên xe.\"",
        "Chính cảm giác không biết đủ ấy làm Mai càng học dữ hơn. Cô muốn có câu trả lời chắc chắn, muốn biết khi nào công an được can thiệp, khi nào trẻ em có quyền từ chối, gia đình phải trình báo ra sao, lời khai cần ghi thế nào để không bị xem nhẹ.",
        "Năm lớp chín, Mai đạt giải nhì học sinh giỏi môn Giáo dục công dân cấp huyện. Phần thi tình huống hỏi về quyền trẻ em bị xâm hại và trách nhiệm tố giác. Mai viết dài nhất phòng thi. Cô không dùng từ hoa mỹ. Cô chỉ viết như thể đang gửi giấy về căn phòng hai lớp cửa năm xưa.",
        "Khi nhận giấy khen, hiệu trưởng đọc tên cô rất rõ: Vàng Thị Mai, bản Lao Chải. Dưới sân, vài bạn từng trêu cô cúi mặt. Mai không thấy hả hê. Cô chỉ thấy một cánh cửa mới mở ra: nếu mình học đủ, một ngày nào đó tên mình sẽ không còn gắn với câu \"con bé bị bán\", mà với câu \"người biết cách cứu người khác\".",
    ],
    5810: [
        "Ngày Mai nhận giấy báo trúng tuyển Đại học Luật Hà Nội, bưu tá phải hỏi đường ba lần mới tìm đến nhà. Bố cô cầm phong bì, không dám xé. Mẹ cô lau tay vào tạp dề rồi cũng không dám chạm. Bà ngoại là người mở thư. Bà nhìn những dòng chữ tiếng Kinh rất lâu dù không đọc hết, rồi đưa cho Mai. \"Nó gọi con đi xa lần nữa,\" bà nói. \"Lần này con tự chọn đường.\"",
        "Cả bản đến xem giấy báo. Có người mừng thật, có người lo thật, cũng có người nói con gái đi xa học nhiều rồi khó lấy chồng. Mai nghe, mỉm cười rất nhẹ. Cô từng bị đưa đi xa vì người khác quyết định thay mình. Lần này, cô cầm giấy báo bằng chính tay mình, biết ga tàu, biết địa chỉ ký túc xá, biết số điện thoại phòng công tác sinh viên.",
        "Hà Nội làm Mai choáng. Đường rộng, xe máy như nước lũ, nhà cao tầng che mất trời. Đêm đầu trong ký túc xá, cô không ngủ. Tiếng cửa phòng đóng mở ngoài hành lang làm cô nhớ tiếng khóa căn nhà bên kia biên giới. Cô đặt ghế chắn trước cửa, rồi thấy xấu hổ với chính mình. Bạn cùng phòng tên Hạnh không hỏi nhiều, chỉ lặng lẽ kéo giường của mình gần cửa hơn. \"Tớ ngủ ngoài này cho,\" Hạnh nói.",
        "Những tháng đầu, Mai học như leo dốc đá. Ngôn ngữ luật khô và khó: cấu thành tội phạm, chủ thể, khách thể, chứng cứ, tố tụng. Bạn bè thành phố quen thuyết trình, quen đọc tài liệu tiếng Anh. Mai phải tra từng từ, ghi từng khái niệm bằng cả tiếng Kinh lẫn tiếng H'Mông để tự hiểu.",
        "Tiền là nỗi lo thường trực. Học bổng chỉ đủ một phần. Mai rửa bát ở quán cơm gần trường từ sáu giờ tối đến mười giờ đêm, tay nứt vì nước rửa chén. Có hôm về phòng, cô mở sách mà mắt díp lại, chữ trên trang nhòe đi. Hạnh giật sách khỏi tay: \"Ngủ đi, mai đọc.\" Mai lắc đầu. \"Mai còn ca làm.\"",
        "Năm hai, lớp học môn Luật Hình sự có chuyên đề về mua bán người. Giảng viên phân tích điều luật, khung hình phạt, yếu tố xuyên biên giới. Cả phòng ghi chép. Mai ngồi bất động. Những chữ trong giáo trình bỗng có mùi cơm nguội, mùi nền xi măng ẩm và tiếng khóa cửa. Cô xin phép ra ngoài, đứng trong nhà vệ sinh rất lâu, tay bám vào bồn rửa.",
        "Giảng viên tìm thấy cô sau giờ học. Thầy không hỏi chuyện riêng, chỉ đưa cô một tập tài liệu về hỗ trợ nạn nhân. \"Nếu em muốn nghiên cứu mảng này, thầy có thể giới thiệu em đến một trung tâm pháp lý.\" Mai nhận tập giấy. Lần đầu tiên, ký ức của cô không chỉ là vết thương; nó có thể trở thành năng lực chuyên môn.",
        "Từ đó, Mai tham gia nhóm trợ giúp pháp lý cho phụ nữ di cư. Cô dịch tiếng H'Mông trong vài buổi tư vấn, nghe những câu chuyện giống mình đến rợn người: bị rủ đi làm, bị hứa hôn nhân, bị giữ giấy tờ, bị đe dọa gia đình sẽ xấu hổ nếu trình báo. Mỗi lần dịch xong, cô về phòng viết lại quy trình bằng ngôn ngữ dễ hiểu hơn.",
        "Bạn bè hỏi vì sao cô chọn lĩnh vực nặng nề như vậy. Mai không trả lời ngay. Đến lễ tốt nghiệp, khi được mời phát biểu đại diện sinh viên dân tộc thiểu số, cô mới nói: \"Có những đứa trẻ không thiếu nghị lực. Các em thiếu người lớn biết tin các em, biết ghi lời khai đúng, biết gọi đúng cơ quan, biết nói với bản làng rằng người bị hại không có lỗi. Tôi học luật để trở thành một người lớn như thế.\"",
        "Dưới hàng ghế cuối, bà ngoại Mai mặc áo chàm, ngồi cạnh mẹ cô. Bà không hiểu hết bài phát biểu, nhưng khi cả hội trường vỗ tay, bà vỗ chậm theo. Mai nhìn xuống, thấy sợi chỉ đỏ mới trên cổ tay bà. Cô biết mình đã đi rất xa khỏi căn nhà khóa ngoài, nhưng đường về hang của con hổ vẫn nằm trong máu.",
    ],
    5811: [
        "Sau khi tốt nghiệp loại giỏi, Mai nhận được lời mời ở lại Hà Nội làm cho một văn phòng luật sư lớn. Mức lương khởi điểm cao hơn thu nhập cả năm của bố mẹ cô cộng lại. Hợp đồng đặt trước mặt cô, giấy trắng, chữ đẹp, điều khoản rõ ràng. Nếu ký, Mai sẽ có phòng trọ tốt hơn, không phải rửa bát, không phải gửi từng đồng học bổng còn lại về nhà.",
        "Nhưng cùng tuần đó, cô nhận cuộc gọi từ cô giáo Thảo. Ở một bản gần Bắc Hà, một bé gái mười ba tuổi mất tích sau khi đi chợ phiên với người quen mới. Gia đình ban đầu không dám báo vì sợ bị cười, sợ con gái \"mang tiếng\". Đến khi báo thì đã muộn hai ngày.",
        "Mai cầm hợp đồng Hà Nội suốt một đêm. Sáng hôm sau, cô từ chối. Trưởng văn phòng hỏi cô có chắc không. Mai nói chắc. Ông bảo: \"Về tỉnh làm trợ giúp pháp lý cực lắm, em sẽ bị cuốn vào những vụ không ai muốn đụng.\" Mai đáp: \"Chính vì không ai muốn đụng nên em phải về.\"",
        "Văn phòng đầu tiên của Mai ở thị trấn Sa Pa chỉ rộng mười sáu mét vuông. Một bàn gỗ cũ, hai ghế nhựa, tủ hồ sơ mua thanh lý và tấm biển viết tay: <strong>Tư vấn pháp lý miễn phí cho phụ nữ, trẻ em bị mua bán, bạo lực, ép hôn.</strong> Ngày treo biển, vài người đi ngang đọc rồi cười nhỏ. Một ông chủ nhà bên cạnh hỏi cô có chắc muốn viết chữ \"bị mua bán\" to như vậy không, sợ xấu khu phố.",
        "Mai giữ nguyên tấm biển. Cô từng thấy xấu hổ bị đặt nhầm lên vai người bị hại. Bây giờ cô muốn những chữ ấy đứng ngoài đường, rõ ràng, để ai cần có thể tìm thấy.",
        "Khách đầu tiên là một người mẹ Dao đỏ. Chị ôm tấm ảnh con gái, tay run đến mức ảnh cong lại. Con chị bị lừa đi làm quán ăn ở Móng Cái, ba tuần không liên lạc. Công an xã đã ghi nhận nhưng bảo gia đình chờ thêm vì có thể con bé tự đi làm. Mai hỏi từng mốc thời gian, từng cuộc gọi, từng người rủ đi, từng biển số xe nghe được. Chị mẹ vừa trả lời vừa khóc vì lần đầu có người hỏi kỹ như vậy.",
        "Mai lập đơn trình báo bổ sung, kèm đề nghị áp dụng biện pháp khẩn cấp xác minh qua nhà mạng và camera bến xe. Cô gọi cho biên phòng, gọi cho hội phụ nữ, gọi cho một anh khóa trên làm ở cơ quan điều tra. Ba ngày sau, cô bé được tìm thấy ở một nhà nghỉ gần biên giới, chưa bị đưa qua cửa khẩu.",
        "Khi cô bé về, em không ôm mẹ ngay. Em đứng nép sau cán bộ, mắt nhìn xuống đất. Mai bước đến, không chạm vào em. Cô chỉ nói bằng tiếng Dao do người mẹ dạy vội: \"Em về rồi. Em không có lỗi.\"",
        "Câu nói ấy làm cô bé bật khóc. Người mẹ quỳ xuống cảm ơn, Mai đỡ dậy ngay. \"Đừng quỳ trước em,\" cô nói. \"Chúng ta còn phải đi lấy lời khai, còn phải giữ chứng cứ, còn phải làm cho kẻ rủ con chị không rủ được ai nữa.\"",
        "Vụ đầu tiên kết thúc bằng một bản án chưa thật nặng như Mai mong muốn, nhưng đủ để cả vùng biết văn phòng nhỏ ở thị trấn không chỉ treo biển cho đẹp. Những người từng cười bắt đầu dẫn người thân đến. Có người đến ban đêm vì sợ hàng xóm thấy. Có người chỉ đứng ngoài cửa nửa tiếng rồi đi. Mai không ép. Cô hiểu bước qua ngưỡng cửa ấy cần nhiều can đảm hơn người ngoài tưởng.",
        "Mỗi tối, Mai ghi lại các lỗ hổng trong từng vụ: gia đình trình báo muộn vì xấu hổ, cán bộ cơ sở thiếu biểu mẫu, trẻ em không biết số gọi khẩn cấp, người dân không phân biệt môi giới việc làm thật và bẫy buôn người. Dần dần, cô nhận ra cứu từng vụ là cần, nhưng chưa đủ. Phải dựng một mạng lưới trước khi chiếc xe lạ rời bản.",
    ],
    5812: [
        "Mạng lưới An Toàn bắt đầu không phải bằng dự án tài trợ lớn, mà bằng một cuộc họp trong nhà văn hóa bản. Mai trải lên bàn ba tờ giấy: một tờ ghi các chiêu lừa phổ biến, một tờ ghi số điện thoại cần gọi, một tờ là bản đồ những điểm xe khách hay đón người đi biên giới. Người đến họp phần lớn là phụ nữ: mẹ, chị, cô giáo, cán bộ hội phụ nữ, vài người từng bị lừa hụt.",
        "Ban đầu, không khí rất ngại. Nói về buôn người ở bản mình chẳng khác nào thừa nhận bản mình có vết thương. Một bác trưởng họ bảo: \"Chuyện xấu thì để công an làm, nói nhiều trẻ con lại biết đường hư.\" Mai đứng dậy, giọng bình tĩnh: \"Trẻ con không hư vì biết nguy hiểm. Trẻ con gặp nguy hiểm khi người lớn bắt các em không được biết.\"",
        "Câu ấy làm căn phòng im lại. Mai không dùng slide phức tạp. Cô diễn lại một tình huống: một người phụ nữ ăn mặc đẹp rủ bé gái xuống thành phố bán hàng ba ngày, đưa tiền trước, để số điện thoại giả. Cô hỏi cả phòng: dấu hiệu nguy hiểm nằm ở đâu? Lúc đầu chỉ vài người trả lời. Sau đó, các bà mẹ bắt đầu nói: không có giấy giới thiệu, không cho gọi video với chủ việc, hứa tiền quá cao, giục đi ngay trong ngày.",
        "Mạng lưới được chia thành từng cụm. Mỗi cụm có ba người giữ số điện thoại khẩn cấp, một người biết đọc giấy tờ, một người nói tốt tiếng Kinh, một người quen đường ra bến xe. Nếu có trẻ chuẩn bị đi làm xa, gia đình phải báo cụm kiểm tra: địa chỉ nơi đến, tên chủ, hợp đồng, số xe, ảnh người đưa đi. Không ai được mắng gia đình nghèo muốn kiếm tiền. Mục tiêu là kiểm chứng, không làm nhục.",
        "Mai cùng biên phòng tổ chức các buổi nói chuyện ở chợ phiên. Cô không đứng trên sân khấu đọc luật khô khan. Cô đưa ra những câu hỏi rất đời: Nếu số điện thoại người tuyển dụng tắt máy thì gọi ai? Nếu con gái gửi tin nhắn \"con ổn\" nhưng không nghe máy thì có chắc là ổn không? Nếu hàng xóm sợ mất mặt không báo, hậu quả là gì?",
        "Có người phản đối. Một nhóm môi giới lao động chửi Mai phá miếng ăn của họ. Một người đàn ông từng rủ mấy cô gái đi làm dưới xuôi đứng giữa chợ nói cô bịa chuyện để nổi tiếng. Mai không cãi tay đôi. Cô mời anh ta đưa giấy phép, hợp đồng mẫu và địa chỉ nơi làm để mạng lưới xác minh miễn phí. Anh ta bỏ đi trước mặt cả chợ.",
        "Vụ ngăn chặn đầu tiên đến sau ba tuần. Một bé gái mười bốn tuổi ở bản Tả Van được rủ đi \"phục vụ nhà hàng\" ở Quảng Tây, tiền lương mười lăm triệu một tháng. Gia đình đã chuẩn bị cho em đi vì nhà nợ tiền phân bón. Tình nguyện viên gọi Mai lúc mười giờ đêm. Mai cùng cán bộ xã đến ngay, kiểm tra số điện thoại người tuyển dụng và phát hiện ảnh căn cước gửi qua Zalo là giả.",
        "Sáng hôm sau, người đến đón bị giữ lại ở bến xe. Trong điện thoại của hắn có danh sách sáu bé gái khác, kèm ghi chú tuổi, bản, hoàn cảnh gia đình. Khi cán bộ đọc danh sách ấy, mẹ của bé gái Tả Van ngồi bệt xuống nền, hai tay ôm mặt. Bà không khóc thành tiếng. Có lẽ vì chỉ cách một đêm nữa, con bà đã biến mất.",
        "Tin vụ đó lan nhanh hơn mọi buổi tuyên truyền. Người dân bắt đầu lưu số của mạng lưới. Các cô giáo dán bảng cảnh báo trong lớp. Mấy bác xe ôm ở bến chợ tự nguyện báo nếu thấy trẻ đi cùng người lạ. Một chủ quán phở cho Mai mượn góc tường treo poster vì \"con gái tôi cũng mười ba tuổi\".",
        "Ba năm, Mạng lưới An Toàn ngăn chặn hai mươi ba vụ, hỗ trợ cứu bốn mươi phụ nữ và trẻ em, nhưng con số không làm Mai ngủ ngon hơn. Mỗi vụ thành công đều nhắc cô có bao nhiêu vụ khác không kịp gọi. Vì vậy cô tiếp tục đi bản, tiếp tục gõ cửa, tiếp tục nhắc đi nhắc lại một câu đến khản giọng: nghèo không có nghĩa là phải giao con mình cho lời hứa không giấy tờ.",
    ],
    5813: [
        "Bà Vàng Thị Páo đã tám mươi tuổi nhưng mắt vẫn sáng. Mỗi chiều, bà ngồi trước hiên thêu thổ cẩm, kim đi lên đi xuống đều như nhịp thở của núi. Mai mỗi lần mệt lại về ngồi cạnh bà. Không cần kể hết vụ án, bà vẫn biết cháu mình vừa đi qua chuyện nặng. Bà chỉ rót chén nước ngô, đẩy sang.",
        "\"Bà ơi, sao ngày xưa bà cứ kể chuyện con hổ?\" Mai hỏi trong một buổi chiều mưa. \"Hồi nhỏ con tưởng bà kể để dọa trẻ con không đi xa.\"",
        "Bà Páo cười, nếp nhăn quanh mắt xếp lại. \"Không phải dọa. Người H'Mông mình ở núi lâu, biết hổ không phải con vật chỉ biết cắn. Hổ biết chờ. Biết nhớ đường. Biết bị thương thì liếm vết thương trong hang, đến khi đủ sức mới đi tiếp.\"",
        "Mai nhìn xuống bàn tay mình. Tay luật sư không còn chai như tay mẹ may vá, nhưng có vết hằn do cầm bút quá lâu, có vết sẹo mờ từ đêm chui qua cửa sổ. \"Có lúc con thấy mình không giống hổ,\" cô nói. \"Có lúc con vẫn sợ tiếng xe máy. Vẫn mơ thấy cửa khóa. Vẫn thấy mình mười hai tuổi.\"",
        "Bà đặt khung thêu xuống. \"Hổ cũng biết sợ. Không sợ thì nó chết sớm. Nhưng sợ mà vẫn nhớ mình phải về đâu, đó mới là hổ.\"",
        "Câu nói ấy theo Mai vào rất nhiều phiên làm việc. Có lần, một bé gái được cứu về không chịu nói, chỉ ngồi xé mép áo. Cán bộ sốt ruột vì cần lời khai. Mai ngồi xuống sàn cùng em, không hỏi gì suốt hai mươi phút. Cô chỉ lấy từ túi ra sợi chỉ đỏ bà ngoại mới buộc cho, đặt lên bàn. \"Chị cũng từng cần một thứ để nhớ đường về,\" cô nói.",
        "Cô bé nhìn sợi chỉ, rồi bật khóc. Lời khai hôm đó không đầy đủ ngay, nhưng đủ mở đầu. Mai hiểu hơn ai hết: người bị hại không phải cái máy kể lại tội ác theo trình tự cho hồ sơ. Họ cần được nhắc rằng mình còn quyền im, quyền thở, quyền nói từng chút một.",
        "Bà Páo không biết chữ nhiều, nhưng bà là người đầu tiên dạy Mai ngôn ngữ của phẩm giá. Khi hàng xóm từng xì xào về Mai, bà chống gậy ra ngõ, nói thẳng: \"Đứa bị bán không xấu. Kẻ bán người mới xấu. Ai nói ngược thì mang cái miệng bẩn về rửa.\" Từ đó, ít người dám nói trước mặt gia đình cô nữa.",
        "Năm Mai ba mươi tuổi, cô đưa bà xuống dự một buổi tuyên dương phụ nữ vùng cao chống buôn người. Người dẫn chương trình giới thiệu Mai là luật sư, người sáng lập Mạng lưới An Toàn. Cả hội trường vỗ tay. Bà Páo ngồi hàng đầu, áo chàm mới, tay cầm sợi chỉ đỏ, mắt long lanh.",
        "Sau buổi lễ, bà hỏi: \"Con thành hổ thật chưa?\" Mai cười, định trả lời như mọi khi rằng con chưa biết. Nhưng nhìn những cô gái trẻ vây quanh hỏi cô cách học luật, nhìn các bà mẹ xin thêm tờ rơi để mang về bản, cô bỗng hiểu câu trả lời không nằm ở cảm giác của mình.",
        "\"Con đang học làm hổ,\" Mai nói. \"Nhưng con không đi một mình nữa.\"",
        "Bà Páo gật đầu. \"Hổ đi một mình thì mạnh. Cả đàn người biết đường về thì còn mạnh hơn.\"",
    ],
    5814: [
        "Năm giờ sáng, Sa Pa chìm trong sương. Mai đứng trên con dốc nhìn xuống Lao Chải, nơi ruộng bậc thang còn phủ một lớp trắng mỏng, mái nhà gỗ nhô lên như những hòn đá ấm. Mười lăm năm trước, cô rời bản trên chiếc xe máy của một người phụ nữ áo đỏ. Hôm nay, cô cũng chuẩn bị rời bản, nhưng trong túi là thẻ luật sư, điện thoại đầy số khẩn cấp và một tập hồ sơ mới.",
        "Cuộc gọi đến lúc bốn giờ bốn mươi. Một cô giáo ở Bắc Hà nói có bé gái tên Dợ, mười ba tuổi, vừa nhắn cho bạn rằng sẽ đi làm xa với một người chị mới quen. Gia đình tưởng em sang Lào Cai phụ bán hàng, nhưng số xe khách trong tin nhắn không đi Lào Cai. Nó đi về hướng biên giới.",
        "Mai không hỏi những câu thừa. Cô yêu cầu cô giáo giữ liên lạc với bạn của Dợ, gửi ảnh người rủ đi, biển số xe nếu có, vị trí cuối cùng, tên bản, tên bố mẹ. Tay cô ghi nhanh vào sổ. Mỗi dòng chữ là một sợi dây kéo một đứa trẻ khỏi mép vực.",
        "Mẹ Mai thức dậy khi cô chuẩn bị áo mưa. Bà không còn hoảng như nhiều năm trước, nhưng ánh mắt vẫn không giấu được lo. \"Con đi một mình à?\" Mai lắc đầu. \"Con gọi biên phòng rồi. Chị Sính ở mạng lưới Bắc Hà đang ra bến xe. Công an xã cũng biết.\"",
        "Bà ngoại Páo ngồi bên bếp, tóc bạc xõa sau vai. Bà buộc vào cổ tay Mai một sợi chỉ đỏ mới. \"Đường có sương,\" bà nói. Mai cúi xuống để bà buộc cho chắc. Sợi chỉ cũ năm mười hai tuổi đã được cô cất trong hộp gỗ. Sợi chỉ mới không phải để nhắc cô sợ. Nó nhắc cô nhớ mình đã sống sót để đi những con đường này.",
        "Xe máy lao xuống dốc, cắt qua màn sương. Hai bên đường, hoa dại ướt nước, vài đứa trẻ đeo cặp đi học sớm nép vào lề khi nghe tiếng xe. Mai chạy chậm qua trường cũ. Cửa lớp chưa mở, nhưng bảng hiệu vẫn đó. Cô nhớ cô bé từng đi bộ mười cây số, nhớ những lời trêu, nhớ cô giáo Thảo viết hai chữ nạn nhân lên bảng.",
        "Đến bến xe Bắc Hà, chị Sính đã giữ được Dợ ở quán nước. Cô bé ngồi co ro, bên cạnh là một ba lô nhỏ. Người phụ nữ rủ em đi đã biến mất khi thấy cán bộ xã xuất hiện. Dợ không khóc. Em chỉ nhìn Mai bằng ánh mắt vừa bướng vừa sợ, giống hệt ánh mắt Mai từng thấy trong gương nước suối năm xưa.",
        "\"Em ghét ở nhà,\" Dợ nói. \"Nhà em nợ tiền. Người ta bảo em đi làm ba tháng là trả hết.\"",
        "Mai ngồi xuống ngang tầm mắt em. \"Chị tin là em muốn giúp nhà. Nhưng người thật sự tuyển em sẽ cho địa chỉ, hợp đồng, số điện thoại gọi được, và không bắt em đi ngay lúc trời chưa sáng.\"",
        "Cô không mắng Dợ ngu. Không hỏi em có biết nguy hiểm không. Mai mở điện thoại, cho em xem từng dấu hiệu lừa đảo, từng vụ đã được cứu, từng số cần gọi. Khi bố mẹ Dợ chạy đến, người cha định tát con vì sợ quá hóa giận. Mai đứng chắn lại. \"Đừng biến nỗi sợ của bác thành cái tát. Hôm nay cháu còn ở đây là may rồi.\"",
        "Buổi trưa, người phụ nữ rủ Dợ bị giữ ở một điểm đón khác nhờ camera bến xe và tin báo của tài xế. Trong điện thoại của chị ta có ảnh năm bé gái, lịch hẹn và tin nhắn chuyển tiền đặt cọc. Mai nhìn những dòng chữ ấy, cổ họng khô lại. Mười lăm năm trôi qua, chiếc bẫy chỉ đổi vỏ. Kẹo ngọt thành điện thoại thông minh, lời hứa thành tin nhắn Zalo, nhưng mục tiêu vẫn là những đứa trẻ nghèo muốn cứu gia đình.",
        "Chiều muộn, sau khi lấy lời khai ban đầu, Mai đứng ngoài sân trụ sở xã. Dợ ngồi bên mẹ, tay cầm tờ giấy có số của Mạng lưới An Toàn. Cô bé chưa hiểu hết hôm nay mình vừa thoát điều gì. Có lẽ nhiều năm sau em mới hiểu. Cũng có thể em sẽ giận vì người lớn chặn giấc mơ kiếm tiền của mình. Mai chấp nhận. Cứu một đứa trẻ không phải lúc nào cũng được cảm ơn ngay.",
        "Trên đường về, sương đã tan. Nắng chiều đổ xuống ruộng bậc thang, làm từng thửa lúa sáng lên như vảy bạc. Mai dừng xe ở lưng dốc, nhìn về phía đường biên mờ xa. Cô nghĩ đến chị Mỷ, Súa, Lử, những người cô chưa tìm lại được. Nghĩ đến cô bé mười hai tuổi từng ngã trước cổng đồn biên phòng và chỉ nói được một câu muốn về nhà.",
        "Điện thoại lại rung. Một tin nhắn từ cô giáo Thảo: <em>Mai, tuần sau trường muốn em về nói chuyện với học sinh lớp bảy.</em> Mai mỉm cười, nhắn lại: <em>Em về.</em>",
        "Cô nổ máy, chạy về phía bản. Sau lưng, mặt trời hạ chậm sau dãy núi. Trước mặt, con đường vẫn quanh co, vẫn có sương, vẫn có những khúc khuất không nhìn hết. Nhưng lần này, con hổ H'Mông không chỉ biết đường về. Cô biết cả cách quay lại, tìm những đứa trẻ khác trước khi cánh cửa khóa sập.",
    ],
}


def wc(text):
    return len(re.findall(r"\w+", re.sub(r"<[^>]+>", " ", text), flags=re.UNICODE))


def main():
    data = json.loads(BACKUP.read_text(encoding="utf-8"))
    updates = []
    editor.upload_helper()
    try:
        meta = editor.update_story_meta(STORY_ID, title=TITLE, intro=INTRO)
        if not meta.get("success"):
            raise RuntimeError(f"Story meta update failed: {meta}")
        updates.append({"story_id": STORY_ID, "meta": "updated"})

        for chapter in data["chapters"]:
            chapter_id = chapter["id"]
            paragraphs = EXPANSIONS[chapter_id]
            content = "\n".join(p(x) for x in paragraphs)
            res = editor.update_chapter(chapter_id, chapter["title"], content)
            if not res.get("success"):
                raise RuntimeError(f"Chapter update failed {chapter_id}: {res}")
            updates.append({"chapter_id": chapter_id, "title": chapter["title"], "local_words": wc(content)})
    finally:
        editor.remove_helper()

    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    headers = [c.value for c in ws[4]]
    col = {v: i + 1 for i, v in enumerate(headers)}
    for row in range(5, ws.max_row + 1):
        if ws.cell(row, col["ID Truyện"]).value == STORY_ID:
            ws.cell(row, col["Tên Truyện"]).value = TITLE
            ws.cell(row, col["Điểm Cần Sửa"]).value = "Đã đọc live và viết lại/mở rộng 10 chương lên chuẩn 1000-1500 từ; tăng mở truyện, lực cản, hành trình pháp lý và kết cứu người."
            ws.cell(row, col["Trạng Thái Sửa"]).value = "☑️ Đã sửa"
            break
    wb.save(XLSX)

    OUT.write_text(json.dumps({"success": True, "updates": updates}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"success": True, "updates": updates}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
