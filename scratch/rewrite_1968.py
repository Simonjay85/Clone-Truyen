import json
import urllib.request
import openpyxl

story_data = {
    "secret_token": "ZEN_TRUYEN_2026_BYPASS",
    "story_id": 1968,
    "title": "Bị Vu Oan Gây Cháy Nhà Máy, Tôi Dùng Camera Giấu Kín Lật Mặt Giám Đốc Bảo Hiểm",
    "intro": "<p>Nguyễn Trần Minh, kỹ sư phòng cháy chữa cháy tại Khu công nghiệp Tân Tạo, bị sa thải và vu oan gây ra vụ cháy nhà máy dệt Hưng Phát — thiệt hại ước tính một trăm hai mươi tỷ đồng.</p>\n<p>Tất cả bằng chứng đều chỉ vào anh: hệ thống báo cháy tự động bị tắt đúng ca trực của anh, báo cáo kiểm tra PCCC cuối cùng mang chữ ký của anh, và camera an ninh ghi lại anh rời nhà máy mười lăm phút trước khi lửa bùng lên.</p>\n<p>Nhưng Trần Minh biết sự thật: anh không tắt hệ thống báo cháy. Ai đó đã thay đổi log hệ thống. Và chiếc camera an ninh bị cắt đúng bốn phút — bốn phút đủ để ai đó châm lửa.</p>\n<p>Khi cả công ty quay lưng, khi cơ quan điều tra liệt anh vào danh sách nghi phạm, Trần Minh chỉ có một thứ: chiếc camera GoPro cá nhân mà anh đeo trên mũ bảo hộ mỗi ca trực — thứ mà không ai biết anh có.</p>",
    "author": "Lê Trọng Nghĩa",
    "seo": {
        "focus_keyword": "vu oan cháy nhà máy, camera giấu kín, bảo hiểm gian lận",
        "seo_title": "Bị Vu Oan Gây Cháy Nhà Máy — Truyện Sảng Văn Kịch Tính | Đọc Tiểu Thuyết",
        "seo_description": "Kỹ sư PCCC bị vu oan gây cháy nhà máy 120 tỷ. Chỉ có chiếc GoPro giấu kín trên mũ bảo hộ mới chứng minh được anh vô tội — và lật mặt kẻ thực sự châm lửa."
    },
    "chapters": []
}

ch1 = {
    "title": "Chương 1: Đám Cháy Lúc Ba Giờ Sáng",
    "content": """<p>Nguyễn Trần Minh nghe tiếng còi báo cháy lúc ba giờ mười bảy phút sáng, đúng lúc anh đang kiểm tra van chữa cháy khu B3 — cách nhà máy dệt Hưng Phát bốn trăm mét.</p>

<p>Tiếng còi rú lên rồi tắt ngấm sau hai giây. Không phải lỗi hệ thống — hệ thống báo cháy Notifier NFS2-3030 của nhà máy Hưng Phát được thiết kế để rú liên tục cho đến khi có người nhấn nút reset thủ công. Hai giây nghĩa là ai đó đã tắt nó.</p>

<p>Trần Minh chạy. Anh ba mươi hai tuổi, cựu lính cứu hỏa Phòng Cảnh sát PCCC Công an TP.HCM, giờ là kỹ sư an toàn cháy nổ tại Ban Quản lý Khu công nghiệp Tân Tạo. Bốn năm trong nghề, anh chưa bao giờ thấy hệ thống báo cháy bị tắt chỉ sau hai giây.</p>

<p>Khi anh chạy đến cổng nhà máy Hưng Phát, lửa đã bùng lên ở khu vực kho nguyên liệu — hàng tấn vải polyester, bông thô, và sợi tổng hợp đang cháy dữ dội. Ngọn lửa cao mười mét, nhiệt bức xạ nóng rát mặt từ khoảng cách năm mươi mét.</p>

<p>Anh nhấn nút báo cháy khẩn cấp trên bộ đàm, gọi đội cứu hỏa Khu công nghiệp. Trong lúc chờ xe chữa cháy, anh mở van nước cứu hỏa gần nhất và bắt đầu phun nước vào khu vực cháy.</p>

<p>Nhưng nước không ra.</p>

<p>Trần Minh nhìn đồng hồ áp suất: kim chỉ số không. Bơm chữa cháy không hoạt động.</p>

<p>Anh chạy sang van thứ hai. Số không.</p>

<p>Van thứ ba. Số không.</p>

<p>Ba bơm chữa cháy khu B — tất cả đều chết. Cùng lúc.</p>

<p>Trần Minh đứng giữa sân nhà máy, nhìn lửa nuốt chửng kho nguyên liệu, và anh biết: đây không phải tai nạn. Ai đó đã tắt hệ thống báo cháy, tắt bơm nước, và châm lửa. Ba hành động có chủ đích, được thực hiện trong cùng một khung giờ.</p>

<p>Xe cứu hỏa đến sau mười một phút. Quá muộn — kho nguyên liệu đã cháy hoàn toàn, lửa đang lan sang xưởng sản xuất chính.</p>

<p>Đến sáu giờ sáng, đám cháy được dập tắt. Nhà máy dệt Hưng Phát — một trong những nhà máy lớn nhất Khu công nghiệp Tân Tạo, thuộc sở hữu của Công ty Cổ phần Dệt may Hưng Phát do ông Lê Đức Thắng làm Tổng Giám đốc — thiệt hại ước tính một trăm hai mươi tỷ đồng.</p>

<p>Và Nguyễn Trần Minh là người cuối cùng rời khỏi nhà máy trước khi lửa bùng.</p>

<p>Ít nhất, đó là những gì camera an ninh ghi lại.</p>

<hr>

<p>Bảy giờ sáng. Trần Minh ngồi trong phòng họp Ban Quản lý Khu công nghiệp, đối diện với Phó Ban quản lý Hoàng Gia Bảo và hai điều tra viên từ Cơ quan Cảnh sát Điều tra Công an quận Bình Tân.</p>

<p>"Anh Minh, camera an ninh cổng B2 ghi lại anh rời nhà máy Hưng Phát lúc 3 giờ 02 phút. Mười lăm phút sau, lửa bùng lên. Anh giải thích được không?"</p>

<p>"Tôi rời cổng B2 lúc 3 giờ 02 vì tôi đang đi kiểm tra van chữa cháy khu B3 — đó là lịch trình ca trực đêm của tôi."</p>

<p>"Nhưng log hệ thống cho thấy tài khoản của anh là tài khoản cuối cùng đăng nhập vào bảng điều khiển PCCC Notifier trước khi hệ thống bị tắt."</p>

<p>Trần Minh cảm thấy sống lưng lạnh toát.</p>

<p>"Tôi không tắt hệ thống. Log đó bị sửa."</p>

<p>"Anh có bằng chứng?"</p>

<p>Trần Minh im lặng. Anh không có bằng chứng — chưa.</p>

<p>Nhưng anh có một thứ mà không ai biết: chiếc camera GoPro Hero 11 gắn trên mũ bảo hộ. Anh bắt đầu đeo nó từ ba tháng trước, khi anh phát hiện rằng ai đó đang can thiệp vào hệ thống PCCC của khu công nghiệp — những lần reset bất thường, những thay đổi cấu hình không có log.</p>

<p>Chiếc GoPro quay liên tục mỗi ca trực, lưu trữ trên thẻ nhớ 256GB. Đêm qua, nó đã ghi lại toàn bộ hành trình của anh — bao gồm cả thời điểm anh ở khu B3, cách nhà máy Hưng Phát bốn trăm mét, khi lửa bùng lên.</p>

<p>Nhưng anh không thể đưa nó ra ngay. Vì nếu kẻ chủ mưu biết anh có video, chúng sẽ tìm cách hủy bằng chứng.</p>

<p>"Tôi cần thời gian để cung cấp bằng chứng," Trần Minh nói.</p>

<p>"Anh có bảy ngày," điều tra viên đáp. "Trong thời gian này, anh bị tạm đình chỉ công việc."</p>

<p>Trần Minh gật đầu, đứng dậy, bước ra.</p>

<p>Bảy ngày. Một chiếc GoPro. Và một trăm hai mươi tỷ đồng đang nằm trong hồ sơ bảo hiểm của ai đó.</p>"""
}

ch2 = {
    "title": "Chương 2: Người Hưởng Lợi Từ Đám Cháy",
    "content": """<p>Ngày thứ nhất sau đình chỉ.</p>

<p>Trần Minh về căn phòng trọ ở quận Bình Tân, đóng cửa, kéo rèm, và mở chiếc GoPro.</p>

<p>Video dài bốn tiếng mười bảy phút — toàn bộ ca trực đêm qua. Anh tua đến 3:02 — thời điểm camera an ninh ghi lại anh rời cổng B2. Trên video GoPro, anh đang đi bộ dọc hành lang khu B3, kiểm tra van chữa cháy số 7. Góc quay rõ ràng: đồng hồ trên tường hành lang hiện 3:04, phía sau anh là kho nguyên liệu khu B3 — không phải nhà máy Hưng Phát.</p>

<p>Anh tua tiếp đến 3:17 — lúc tiếng còi báo cháy rú lên. Video ghi lại phản ứng của anh: giật mình, nhìn về phía nhà máy Hưng Phát, rồi chạy. GoPro rung lắc theo từng bước chân.</p>

<p>Bằng chứng ngoại phạm hoàn hảo.</p>

<p>Nhưng Trần Minh không chỉ cần chứng minh mình vô tội. Anh cần tìm ra ai đã châm lửa — và tại sao.</p>

<p>Anh mở laptop, bắt đầu nghiên cứu.</p>

<hr>

<p>Nhà máy dệt Hưng Phát thuộc Công ty Cổ phần Dệt may Hưng Phát, thành lập năm 2015, vốn điều lệ năm mươi tỷ đồng. Tổng Giám đốc: Lê Đức Thắng, năm mươi bốn tuổi, từng là Phó Giám đốc Sở Công Thương TP.HCM trước khi nghỉ hưu sớm để kinh doanh.</p>

<p>Trần Minh nhớ: ba tháng trước, anh nghe đồng nghiệp trong Ban Quản lý khu công nghiệp nói rằng nhà máy Hưng Phát đang gặp khó khăn tài chính. Đơn hàng xuất khẩu giảm mạnh do thị trường Mỹ áp thuế chống bán phá giá, doanh thu quý III giảm bốn mươi phần trăm so với cùng kỳ.</p>

<p>Anh tra cứu trên Google: "Nhà máy dệt Hưng Phát bảo hiểm."</p>

<p>Kết quả: Hưng Phát đã mua bảo hiểm cháy nổ tại Công ty Bảo hiểm Phú An — gói Platinum, mức bồi thường tối đa hai trăm tỷ đồng. Hợp đồng bảo hiểm được ký cách đây sáu tháng — tức là ngay sau khi doanh thu bắt đầu giảm.</p>

<p>Mua bảo hiểm trước khi cháy. Mức bồi thường cao hơn giá trị thiệt hại. Đây là dấu hiệu kinh điển của gian lận bảo hiểm.</p>

<p>Nhưng ai là người thực hiện? Lê Đức Thắng — chủ nhà máy — là nghi phạm rõ ràng nhất. Nhưng ông ta không thể tự châm lửa — ông ta cần người bên trong khu công nghiệp.</p>

<p>Trần Minh nghĩ đến một người: Hoàng Gia Bảo, Phó Ban quản lý Khu công nghiệp Tân Tạo. Người đã ngồi đối diện anh sáng nay trong phòng họp, mặt lạnh như đá, không hỏi một câu nào để bênh vực anh — dù anh là nhân viên dưới quyền bốn năm.</p>

<p>Trần Minh mở hồ sơ nhân sự điện tử — anh vẫn có quyền truy cập vì tài khoản chưa bị khóa. Anh tra cứu thông tin Hoàng Gia Bảo.</p>

<p>Bốn mươi bảy tuổi. Tốt nghiệp Đại học Bách khoa TP.HCM. Vào Ban Quản lý khu công nghiệp năm 2018. Trước đó — và đây là chi tiết khiến Trần Minh dừng lại — Hoàng Gia Bảo từng là Trưởng phòng Kỹ thuật tại Công ty Bảo hiểm Phú An.</p>

<p>Phú An. Cùng công ty bảo hiểm đã bán gói Platinum cho Hưng Phát.</p>

<p>Trần Minh đóng laptop. Anh cần thêm bằng chứng — nhưng bức tranh đang dần hiện lên.</p>

<p>Ông chủ nhà máy Hưng Phát mua bảo hiểm cháy nổ mức cao bất thường. Phó Ban quản lý khu công nghiệp — cựu nhân viên công ty bảo hiểm — có khả năng tiếp cận hệ thống PCCC. Và kỹ sư PCCC Nguyễn Trần Minh được chọn làm vật tế thần.</p>

<p>Anh nhìn chiếc GoPro trên bàn. Bốn tiếng mười bảy phút video — nặng 28 gigabyte.</p>

<p>Hai mươi tám gigabyte sự thật. Nhưng chỉ một mình anh biết nó tồn tại.</p>"""
}

ch3 = {
    "title": "Chương 3: Log Hệ Thống Bị Sửa",
    "content": """<p>Ngày thứ hai.</p>

<p>Trần Minh không ngồi yên. Anh liên lạc với Đặng Văn Phúc — đồng nghiệp cũ ở Phòng Cảnh sát PCCC, giờ là Thiếu tá phụ trách điều tra cháy nổ.</p>

<p>"Phúc, tao cần mày giúp một việc."</p>

<p>"Mày đang bị nghi phạm trong vụ Hưng Phát, Minh. Tao không thể can thiệp."</p>

<p>"Tao không nhờ mày can thiệp. Tao nhờ mày kiểm tra một thứ: log hệ thống Notifier NFS2-3030 của nhà máy Hưng Phát. Hệ thống đó lưu log trên hai nơi — server nội bộ và cloud backup của hãng Notifier. Nếu ai đó sửa log trên server nội bộ, bản cloud sẽ khác."</p>

<p>Im lặng trên điện thoại. Rồi Phúc nói: "Mày chắc không?"</p>

<p>"Hệ thống Notifier NFS2-3030 đời 2022 trở lên có tính năng cloud backup tự động qua giao thức MQTT. Bản backup không thể sửa từ phía người dùng — chỉ hãng Notifier mới có quyền. Nếu log nội bộ nói tài khoản của tao tắt hệ thống, nhưng bản cloud nói khác — thì ai đó đã sửa log."</p>

<p>"Để tao kiểm tra. Nhưng mày phải giữ im chuyện này."</p>

<p>"Được."</p>

<hr>

<p>Trong lúc chờ Phúc, Trần Minh tiếp tục điều tra Hoàng Gia Bảo.</p>

<p>Anh biết rằng với tư cách Phó Ban quản lý, Bảo có thẻ truy cập tất cả cơ sở hạ tầng trong khu công nghiệp — bao gồm phòng điều khiển PCCC trung tâm, nơi đặt bảng điều khiển Notifier. Bảo cũng có mật khẩu quản trị hệ thống — vì anh ta từng yêu cầu Trần Minh cấp quyền "để phục vụ công tác quản lý" cách đây một năm.</p>

<p>Trần Minh kiểm tra lại email cũ. Đúng: ngày 15 tháng 3 năm ngoái, Hoàng Gia Bảo gửi email yêu cầu cấp tài khoản admin trên hệ thống PCCC. Trần Minh đã cấp theo quy trình — vì Bảo là cấp trên trực tiếp.</p>

<p>Tức là Bảo có quyền đăng nhập, thay đổi cấu hình, và tắt hệ thống báo cháy. Và quan trọng hơn — Bảo có quyền sửa log.</p>

<p>Trần Minh forward email đó cho Phúc, kèm ghi chú: "Kiểm tra xem tài khoản admin này có log đăng nhập vào đêm cháy không."</p>

<hr>

<p>Chiều hôm đó, Trần Minh nhận cuộc gọi từ một số lạ.</p>

<p>"Anh Minh? Tôi là Nguyễn Thị Hạnh, nhân viên kế toán nhà máy Hưng Phát. Tôi cần gặp anh."</p>

<p>"Tại sao?"</p>

<p>"Vì tôi biết ai đã châm lửa. Và tôi sợ."</p>

<p>Trần Minh hẹn gặp Hạnh tại quán cà phê trên đường Kinh Dương Vương, quận Bình Tân — đủ xa khu công nghiệp để không bị phát hiện.</p>

<p>Hạnh đến muộn mười phút. Cô ba mươi tuổi, gầy, mắt có quầng thâm, tay cầm ly nước run nhẹ.</p>

<p>"Anh Minh, tôi làm kế toán cho Hưng Phát ba năm. Hai tuần trước vụ cháy, giám đốc Lê Đức Thắng bảo tôi chuẩn bị hồ sơ yêu cầu bồi thường bảo hiểm. Ông ta đưa tôi danh sách tài sản — nhưng giá trị ghi trong danh sách cao hơn giá trị thực ít nhất ba lần."</p>

<p>"Chuẩn bị hồ sơ bồi thường trước khi cháy?"</p>

<p>"Dạ. Ông ta nói 'chuẩn bị sẵn để phòng trường hợp'. Lúc đó tôi không nghĩ gì. Nhưng khi nhà máy cháy thật, tôi hiểu."</p>

<p>Trần Minh nhìn Hạnh. Đôi mắt cô đầy sợ hãi — nhưng cũng đầy phẫn nộ.</p>

<p>"Chị có giữ bản danh sách đó không?"</p>

<p>"Có. Tôi photo một bản giấu ở nhà. Và tôi có email ông Thắng gửi cho tôi với file đính kèm — ngày gửi là hai tuần trước vụ cháy."</p>

<p>Trần Minh cảm thấy mạch máu đập nhanh hơn. Đây là bằng chứng mấu chốt: email có timestamp, chứng minh Lê Đức Thắng đã chuẩn bị hồ sơ bồi thường bảo hiểm trước khi cháy.</p>

<p>"Chị Hạnh, tôi cần chị forward email đó cho tôi. Và tôi cần chị sẵn sàng làm nhân chứng."</p>

<p>Hạnh nhìn xuống ly cà phê, hai tay nắm chặt.</p>

<p>"Tôi sợ, anh Minh. Ông Thắng có quan hệ rộng. Nếu ông ta biết tôi tiết lộ..."</p>

<p>"Tôi hiểu. Nhưng nếu chị không lên tiếng, tôi sẽ vào tù vì tội phóng hỏa — một tội mà tôi không làm. Và ông Thắng sẽ nhận hai trăm tỷ tiền bảo hiểm cho một vụ cháy mà chính ông ta gây ra."</p>

<p>Hạnh im lặng rất lâu.</p>

<p>Rồi cô gật đầu.</p>"""
}

ch4 = {
    "title": "Chương 4: Bản Cloud Không Biết Nói Dối",
    "content": """<p>Ngày thứ ba.</p>

<p>Đặng Văn Phúc gọi điện cho Trần Minh lúc sáu giờ sáng.</p>

<p>"Minh, mày đúng."</p>

<p>"Đúng cái gì?"</p>

<p>"Log cloud backup của Notifier khác hoàn toàn log nội bộ. Bản cloud ghi rõ: hệ thống báo cháy bị tắt lúc 3 giờ 14 phút bởi tài khoản admin — không phải tài khoản của mày. Tài khoản của mày không có log đăng nhập nào từ 2 giờ đến 4 giờ sáng."</p>

<p>Trần Minh nhắm mắt, hít một hơi sâu. Ba ngày anh nín thở — bây giờ anh mới thở được.</p>

<p>"Tài khoản admin đó là của ai?"</p>

<p>"Tên đăng nhập: HGB_admin. Tao đã kiểm tra — tài khoản này được tạo bởi email hoanggiabao@tanTao_kcc.gov.vn."</p>

<p>Hoàng Gia Bảo. Phó Ban quản lý Khu công nghiệp Tân Tạo.</p>

<p>"Phúc, còn gì nữa không?"</p>

<p>"Có. Bản cloud còn ghi log hệ thống bơm chữa cháy. Ba bơm khu B bị tắt thủ công lúc 2 giờ 58 phút — mười chín phút trước khi cháy. Cũng bởi tài khoản HGB_admin."</p>

<p>Tắt hệ thống báo cháy. Tắt bơm nước. Rồi châm lửa. Ba bước, một tay đạo diễn.</p>

<p>"Phúc, tao cần mày làm một việc nữa: yêu cầu hãng Notifier cung cấp bản cloud backup chính thức, có xác nhận, có dấu mộc. Đó là bằng chứng pháp lý."</p>

<p>"Tao hiểu. Nhưng mày cần biết: nếu Hoàng Gia Bảo là người tắt hệ thống, thì hắn không hành động một mình. Hắn cần người bên ngoài châm lửa vật lý — và hắn cần người trả tiền."</p>

<p>"Lê Đức Thắng."</p>

<p>"Chưa đủ cơ sở. Mày cần bằng chứng liên kết hai người."</p>

<hr>

<p>Trần Minh mở email mà Nguyễn Thị Hạnh đã forward cho anh đêm qua. Email từ Lê Đức Thắng gửi cho Hạnh, ngày 5 tháng 5 — hai tuần trước vụ cháy. File đính kèm: "Danh_sach_tai_san_bao_hiem.xlsx".</p>

<p>Anh mở file. Danh sách ghi rõ: máy dệt nhập khẩu Đức trị giá tám mươi tỷ (giá thực: ba mươi lăm tỷ), kho nguyên liệu trị giá bốn mươi tỷ (giá thực: mười hai tỷ), hệ thống điện và tự động hóa ba mươi tỷ (giá thực: mười bốn tỷ).</p>

<p>Tổng giá trị khai báo: một trăm năm mươi tỷ. Giá trị thực: khoảng sáu mươi mốt tỷ. Chênh lệch: gần chín mươi tỷ đồng.</p>

<p>Lê Đức Thắng đã khai khống giá trị tài sản gần gấp ba lần để nhận bồi thường bảo hiểm cao hơn. Và ông ta đã chuẩn bị hồ sơ này hai tuần trước khi cháy.</p>

<p>Trần Minh lưu email và file Excel vào hai USB khác nhau, một để ở nhà, một gửi cho Phúc.</p>

<p>Bây giờ anh có: video GoPro chứng minh ngoại phạm, bản cloud log chứng minh Hoàng Gia Bảo tắt hệ thống PCCC, và email chứng minh Lê Đức Thắng chuẩn bị gian lận bảo hiểm trước khi cháy.</p>

<p>Thiếu một mảnh ghép cuối: bằng chứng liên kết Lê Đức Thắng và Hoàng Gia Bảo.</p>

<p>Trần Minh ngồi lại, nhìn lên trần nhà. Anh nghĩ đến một nơi mà mối quan hệ giữa hai người có thể lộ ra: sao kê ngân hàng.</p>

<p>Nhưng anh không phải cơ quan điều tra — anh không có quyền yêu cầu sao kê ngân hàng.</p>

<p>Trừ khi... ai đó trong ngân hàng sẵn sàng giúp.</p>

<p>Trần Minh lấy điện thoại, cuộn danh bạ đến chữ T: Trần Quang — anh trai, Phó phòng Giao dịch Vietcombank chi nhánh Bình Tân.</p>

<p>"Anh Quang, em cần nói chuyện."</p>"""
}

ch5 = {
    "title": "Chương 5: Đường Tiền",
    "content": """<p>Trần Quang ngồi đối diện em trai trong quán cơm bình dân trên đường Tên Lửa, Bình Tân. Anh bốn mươi tuổi, mặc sơ mi trắng, mặt nghiêm, hai tay đan trước ngực.</p>

<p>"Minh, em biết anh không thể cung cấp sao kê của khách hàng. Đó là vi phạm Luật Các Tổ chức Tín dụng."</p>

<p>"Em biết. Em không nhờ anh cung cấp sao kê. Em nhờ anh xác nhận một thông tin: Hoàng Gia Bảo có tài khoản tại Vietcombank Bình Tân không?"</p>

<p>Trần Quang im lặng một lúc.</p>

<p>"Có."</p>

<p>"Trong sáu tháng qua, tài khoản đó có giao dịch lớn bất thường không?"</p>

<p>"Minh..."</p>

<p>"Anh ơi, em đang bị vu oan tội phóng hỏa. Nếu không tìm ra bằng chứng trong bốn ngày, em sẽ bị khởi tố. Em không cần số cụ thể — chỉ cần biết có hay không."</p>

<p>Trần Quang nhìn em trai. Đôi mắt Trần Minh — đôi mắt mà anh đã nhìn thấy từ khi nó còn là thằng nhóc mười tuổi chạy theo anh ra cánh đồng ở Củ Chi — bây giờ đầy mệt mỏi nhưng kiên quyết.</p>

<p>"Có," Trần Quang nói, giọng rất nhỏ. "Ba tháng trước, tài khoản của Hoàng Gia Bảo nhận một khoản chuyển đến hai tỷ đồng từ một công ty tên 'Công ty TNHH Tư vấn Đại Thắng'. Anh nhớ vì cái tên nghe kỳ — và vì số tiền lớn bất thường so với mức lương công chức."</p>

<p>Công ty TNHH Tư vấn Đại Thắng. Đại Thắng. Lê Đức Thắng.</p>

<p>Trần Minh tra cứu trên điện thoại: Công ty TNHH Tư vấn Đại Thắng, thành lập tháng 1 năm nay, trụ sở quận 7, giám đốc: Lê Thị Thanh Hằng — vợ Lê Đức Thắng.</p>

<p>Hai tỷ đồng chuyển từ công ty của vợ Lê Đức Thắng sang tài khoản cá nhân Hoàng Gia Bảo, ba tháng trước vụ cháy.</p>

<p>Đường tiền. Mảnh ghép cuối cùng.</p>

<hr>

<p>Tối hôm đó, Trần Minh ngồi trước laptop, tổng hợp toàn bộ bằng chứng:</p>

<p>1. Video GoPro — chứng minh anh ở khu B3, cách hiện trường bốn trăm mét, khi cháy xảy ra.</p>

<p>2. Bản cloud backup Notifier — chứng minh tài khoản HGB_admin (Hoàng Gia Bảo) tắt hệ thống báo cháy và bơm nước trước khi cháy.</p>

<p>3. Email Lê Đức Thắng — chuẩn bị hồ sơ bồi thường bảo hiểm với giá trị khai khống hai tuần trước vụ cháy.</p>

<p>4. Đường tiền — hai tỷ đồng từ công ty vợ Lê Đức Thắng sang Hoàng Gia Bảo.</p>

<p>Bốn mảnh ghép. Một bức tranh hoàn chỉnh: Lê Đức Thắng thuê Hoàng Gia Bảo đốt nhà máy để nhận bồi thường bảo hiểm, và dùng Trần Minh làm vật tế thần.</p>

<p>Trần Minh viết một bản tóm tắt dài mười hai trang, đính kèm tất cả bằng chứng, in ra ba bản:</p>

<p>Bản một: gửi Thiếu tá Đặng Văn Phúc — Phòng Cảnh sát PCCC.</p>
<p>Bản hai: gửi Cơ quan Cảnh sát Điều tra Công an quận Bình Tân — đơn vị đang xử lý vụ án.</p>
<p>Bản ba: giữ tại nhà, trong két sắt nhỏ.</p>

<p>Anh nhìn đồng hồ: mười một giờ đêm. Còn bốn ngày.</p>

<p>Nhưng anh sẽ không cần bốn ngày. Ngày mai, sự thật sẽ lên tiếng.</p>"""
}

ch6 = {
    "title": "Chương 6: Phiên Đối Chất",
    "content": """<p>Ngày thứ tư. Trụ sở Công an quận Bình Tân.</p>

<p>Phòng làm việc của Đội Điều tra tội phạm kinh tế. Bốn người ngồi quanh bàn: Trần Minh, Thiếu tá Phúc, Thượng tá Nguyễn Thanh Tùng — Đội trưởng Đội Điều tra, và một chuyên viên giám định kỹ thuật từ Phòng Kỹ thuật Hình sự.</p>

<p>"Anh Minh, hôm nay chúng tôi mời anh đến để xem xét các bằng chứng mới mà anh cung cấp," Thượng tá Tùng nói, giọng trung tính. "Chúng tôi cũng đã mời ông Hoàng Gia Bảo đến — ông ta đang ở phòng bên."</p>

<p>Trần Minh gật đầu.</p>

<p>Thượng tá Tùng mở laptop, chiếu video GoPro lên màn hình lớn. Cả phòng xem đoạn video từ 3:00 đến 3:20 — rõ ràng Trần Minh đang ở khu B3, kiểm tra van chữa cháy, khi tiếng còi báo cháy rú lên.</p>

<p>"Video này đã được chuyên viên kỹ thuật xác minh: không có dấu hiệu chỉnh sửa, metadata file cho thấy ngày quay trùng khớp với đêm xảy ra cháy."</p>

<p>Tiếp theo, Thượng tá Tùng trình bày bản cloud backup từ Notifier — bản chính thức có xác nhận của đại diện hãng tại Việt Nam.</p>

<p>"Bản cloud cho thấy tài khoản HGB_admin đăng nhập vào hệ thống PCCC lúc 2:51, tắt hệ thống báo cháy lúc 3:14, và tắt ba bơm chữa cháy khu B lúc 2:58. Bản log nội bộ — bản mà phòng kỹ thuật nhà máy cung cấp cho chúng tôi trước đó — đã bị sửa, thay tên đăng nhập thành tài khoản của anh Minh."</p>

<p>"Ai có quyền sửa log nội bộ?"</p>

<p>"Chỉ người có quyền admin. Tài khoản HGB_admin được cấp cho ông Hoàng Gia Bảo — Phó Ban quản lý Khu công nghiệp."</p>

<hr>

<p>Hoàng Gia Bảo được đưa vào phòng.</p>

<p>Anh ta bước vào, mặt bình thường — cho đến khi nhìn thấy Trần Minh. Mắt anh ta giật nhẹ, nhưng nhanh chóng lấy lại bình tĩnh.</p>

<p>"Ông Bảo, ông giải thích về việc tài khoản admin do ông quản lý đã được sử dụng để tắt hệ thống PCCC đêm xảy ra cháy?" Thượng tá Tùng hỏi.</p>

<p>"Tôi không sử dụng tài khoản đó. Có thể ai đó đã lấy mật khẩu của tôi."</p>

<p>"Mật khẩu của ông được bảo mật như thế nào?"</p>

<p>"Tôi lưu trên máy tính cá nhân, có khóa vân tay."</p>

<p>"Ông có chia sẻ mật khẩu cho ai không?"</p>

<p>"Không."</p>

<p>"Vậy chỉ ông mới có quyền truy cập tài khoản HGB_admin?"</p>

<p>Im lặng kéo dài.</p>

<p>Thượng tá Tùng mở tiếp hồ sơ:</p>

<p>"Ông Bảo, chúng tôi cũng phát hiện rằng ba tháng trước vụ cháy, tài khoản cá nhân của ông tại Vietcombank Bình Tân nhận một khoản chuyển hai tỷ đồng từ Công ty TNHH Tư vấn Đại Thắng — công ty do vợ ông Lê Đức Thắng, Tổng Giám đốc Hưng Phát, làm giám đốc. Ông giải thích khoản tiền này là gì?"</p>

<p>Mặt Hoàng Gia Bảo trắng bệch. Tay anh ta bắt đầu run.</p>

<p>"Tôi... đó là tiền tư vấn. Tôi tư vấn cho công ty Đại Thắng về... về thủ tục hành chính."</p>

<p>"Tư vấn gì mà hai tỷ, ông Bảo? Trong khi ông là công chức nhà nước, lương hàng tháng mười tám triệu đồng?"</p>

<p>Hoàng Gia Bảo không trả lời. Anh ta nhìn xuống bàn, hai tay nắm chặt đến mức các đốt ngón tay trắng bệch.</p>

<p>Trần Minh ngồi yên, không nói gì. Anh không cần nói — bằng chứng đã nói thay anh.</p>"""
}

ch7 = {
    "title": "Chương 7: Hoàng Gia Bảo Khai Hết",
    "content": """<p>Hoàng Gia Bảo khai tất cả vào ngày thứ năm.</p>

<p>Theo lời khai, kế hoạch đốt nhà máy được Lê Đức Thắng đề xuất cách đây bốn tháng, trong một bữa nhậu tại nhà hàng Riverside Palace, quận 4. Thắng nói thẳng: nhà máy đang thua lỗ, đơn hàng cạn, nợ ngân hàng đến hạn — cách duy nhất để cứu công ty là đốt nhà máy, nhận tiền bảo hiểm.</p>

<p>Bảo được giao nhiệm vụ: tắt hệ thống PCCC vào đêm được chọn, tắt bơm nước, và sửa log hệ thống để đổ tội cho Trần Minh — vì Trần Minh là người trực ca đêm, có tài khoản trên hệ thống, và không có ai chứng kiến anh ở đâu vào lúc cháy.</p>

<p>Người châm lửa thực tế là một thanh niên tên Trần Văn Mạnh — cháu vợ Hoàng Gia Bảo, hai mươi lăm tuổi, thất nghiệp. Mạnh được trả năm trăm triệu đồng để lẻn vào kho nguyên liệu qua cửa phụ (khóa đã được Bảo mở trước), đổ xăng, và châm lửa.</p>

<p>Tiền trả cho Bảo: hai tỷ đồng trước (qua công ty Đại Thắng), và hai tỷ nữa sau khi nhận bồi thường bảo hiểm.</p>

<p>Tiền trả cho Mạnh: năm trăm triệu, chia ba đợt.</p>

<hr>

<p>Khi Thượng tá Tùng đọc lại bản khai cho Bảo ký, Trần Minh ngồi ở phòng bên, uống một ly nước lọc.</p>

<p>Phúc ngồi cạnh, vỗ vai anh.</p>

<p>"Xong rồi, Minh. Mày được minh oan."</p>

<p>Trần Minh không nói gì. Anh nhìn ra cửa sổ — bầu trời Bình Tân chiều nay trong vắt, nắng vàng rót xuống mái tôn nhà dân.</p>

<p>"Phúc, tao nghĩ mãi một điều."</p>

<p>"Gì?"</p>

<p>"Nếu tao không đeo GoPro, tao đã vào tù."</p>

<p>"Đúng."</p>

<p>"Và nếu chị Hạnh — kế toán Hưng Phát — không dám lên tiếng, tao cũng vào tù."</p>

<p>"Cũng đúng."</p>

<p>"Mấy ngày qua, không ai trong Ban Quản lý bênh tao. Hai mươi đồng nghiệp, không ai gọi hỏi thăm. Họ sợ Hoàng Gia Bảo."</p>

<p>Phúc im lặng. Anh hiểu cảm giác đó — cảm giác bị cả tổ chức bỏ rơi, khi mọi người chọn im lặng thay vì đứng về phía sự thật.</p>

<p>"Nhưng mày không bỏ cuộc," Phúc nói. "Đó là điều quan trọng."</p>

<p>Trần Minh gật đầu.</p>

<p>"Tao sẽ không quay lại Ban Quản lý nữa."</p>

<p>"Vậy mày sẽ làm gì?"</p>

<p>Trần Minh nhìn chiếc GoPro trên bàn — chiếc camera nhỏ bé đã cứu đời anh.</p>

<p>"Tao sẽ mở một công ty tư vấn PCCC. Chuyên kiểm toán hệ thống phòng cháy cho khu công nghiệp. Vì nếu đêm đó hệ thống hoạt động đúng, nhà máy không cháy thành tro."</p>

<p>Phúc cười.</p>

<p>"Ít nhất thì mày vẫn chọn đúng nghề."</p>"""
}

ch8 = {
    "title": "Chương 8: Kỹ Sư Phòng Cháy Không Bao Giờ Tắt Lửa",
    "content": """<p>Sáu tháng sau.</p>

<p>Lê Đức Thắng bị khởi tố về hai tội: "Hủy hoại tài sản" theo Điều 178 Bộ luật Hình sự và "Gian lận trong kinh doanh bảo hiểm" theo Điều 213. Công ty Bảo hiểm Phú An từ chối chi trả bồi thường và khởi kiện ngược Hưng Phát đòi lại phí bảo hiểm đã đóng.</p>

<p>Hoàng Gia Bảo bị khởi tố về tội "Hủy hoại tài sản" với vai trò đồng phạm. Anh ta bị cách chức, khai trừ Đảng, và bị tạm giam.</p>

<p>Trần Văn Mạnh — người trực tiếp châm lửa — bị bắt tại Bình Dương sau hai tuần truy nã.</p>

<p>Nguyễn Thị Hạnh — kế toán Hưng Phát, nhân chứng quan trọng — được Cơ quan Điều tra bảo vệ theo chương trình bảo vệ nhân chứng.</p>

<hr>

<p>Trần Minh thành lập Công ty TNHH Kiểm toán PCCC Trần Minh — tên đầy đủ mà anh tự chọn, vì anh muốn mỗi khách hàng biết rằng người đứng sau công ty là một kỹ sư đã từng bị vu oan, và đã tự mình chứng minh sự thật.</p>

<p>Văn phòng đặt tại quận 7, TP.HCM — một căn hộ chung cư chuyển đổi công năng, nhỏ nhưng gọn gàng. Ba nhân viên: Trần Minh, một kỹ sư mới ra trường, và chị Hạnh — đúng, Nguyễn Thị Hạnh, cựu kế toán Hưng Phát, bây giờ là kế toán kiêm trợ lý hành chính của công ty.</p>

<p>"Anh Minh, hợp đồng kiểm toán PCCC với Khu công nghiệp Hiệp Phước đã ký xong," Hạnh nói, đặt tập hồ sơ lên bàn. "Họ muốn kiểm tra toàn bộ hệ thống Notifier và bơm chữa cháy của mười hai nhà máy."</p>

<p>"Mười hai nhà máy. Mất bao lâu?"</p>

<p>"Hai tuần nếu làm liên tục."</p>

<p>"Tốt. Bắt đầu từ thứ Hai."</p>

<p>Hạnh gật đầu, quay đi. Nhưng cô dừng lại ở cửa.</p>

<p>"Anh Minh, em muốn hỏi một chuyện."</p>

<p>"Gì?"</p>

<p>"Hôm đó ở quán cà phê, khi em nói em sợ — anh có sợ không?"</p>

<p>Trần Minh nhìn Hạnh. Cô đã thay đổi nhiều trong sáu tháng — không còn gầy gò, quầng thâm dưới mắt đã bớt, giọng nói tự tin hơn.</p>

<p>"Có. Tao sợ. Sợ nhất là khi ngồi trong phòng điều tra, nhìn mặt Hoàng Gia Bảo, và biết rằng nếu không có chiếc GoPro, hắn sẽ thắng."</p>

<p>"Vậy tại sao anh vẫn đeo GoPro mỗi ca trực? Từ trước khi vụ cháy xảy ra?"</p>

<p>Trần Minh cười nhẹ.</p>

<p>"Vì tao là kỹ sư phòng cháy, Hạnh. Nghề của tao là chuẩn bị cho những thứ chưa xảy ra. Tao lắp bình chữa cháy ở những nơi chưa bao giờ cháy. Tao kiểm tra van nước ở những đường ống chưa bao giờ dùng. Và tao đeo camera ở những ca trực chưa bao giờ có sự cố."</p>

<p>"Rồi sự cố xảy ra."</p>

<p>"Rồi sự cố xảy ra. Và tao đã chuẩn bị."</p>

<p>Hạnh mỉm cười, quay đi.</p>

<hr>

<p>Trần Minh ngồi lại một mình trong văn phòng nhỏ. Trên bàn, cạnh laptop, là chiếc GoPro Hero 11 — vẫn chiếc đó, vẫn cái vỏ đen xước mép, vẫn dây đeo đã sờn.</p>

<p>Anh cầm nó lên, xoay trong tay.</p>

<p>Hai mươi tám gigabyte. Bốn tiếng mười bảy phút. Một đời người.</p>

<p>Anh đặt GoPro xuống, mở laptop, bắt đầu viết báo cáo kiểm toán PCCC cho mười hai nhà máy ở Hiệp Phước.</p>

<p>Bên ngoài cửa sổ, Sài Gòn chiều nay nắng vàng. Không có khói. Không có lửa.</p>

<p>Chỉ có ánh sáng — thứ ánh sáng bình thường nhất trên đời, nhưng chỉ những người từng đứng giữa bóng tối mới thật sự biết nó quý giá đến nhường nào.</p>"""
}

story_data["chapters"] = [ch1, ch2, ch3, ch4, ch5, ch6, ch7, ch8]

# Save JSON
with open("scratch/story_1968_rewrite.json", "w", encoding="utf-8") as f:
    json.dump(story_data, f, ensure_ascii=False, indent=2)

print("✅ Truyện 1968 viết xong!")
for i, ch in enumerate(story_data['chapters']):
    print(f"  Ch{i+1}: {ch['title']} — {len(ch['content'])} chars")

# Upload to WordPress
print("\n📤 Uploading to WordPress...")
upload_url = "https://doctieuthuyet.com/overwrite_story_v13.php"
payload = json.dumps(story_data).encode('utf-8')
req = urllib.request.Request(upload_url, data=payload, headers={'Content-Type': 'application/json'})
try:
    resp = urllib.request.urlopen(req, timeout=120)
    result = json.loads(resp.read().decode('utf-8'))
    print(f"  ✅ Success! ID={result['story_id']}, Chapters={result['chapters_count']}")
except Exception as e:
    print(f"  ❌ Error: {e}")

# Update Excel
print("\n📊 Updating Excel...")
wb = openpyxl.load_workbook("danh_sach_truyen_doctieuthuyet.xlsx")
ws = wb.active
for r in range(5, ws.max_row+1):
    if ws.cell(row=r, column=2).value and str(ws.cell(row=r, column=2).value).strip() == "1968":
        ws.cell(row=r, column=3).value = story_data["title"]
        ws.cell(row=r, column=4).value = story_data["author"]
        ws.cell(row=r, column=5).value = "Sảng Văn"
        ws.cell(row=r, column=12).value = "Kỹ sư PCCC bị vu oan gây cháy nhà máy 120 tỷ. Dùng camera GoPro giấu kín trên mũ bảo hộ, anh chứng minh mình vô tội và lật mặt giám đốc bảo hiểm câu kết đốt nhà máy lấy tiền bồi thường."
        ws.cell(row=r, column=13).value = "REWRITE TOÀN BỘ: Cốt truyện fantasy lộn xộn, AI template, meta-breaking → ĐÃ HOÀN THÀNH"
        ws.cell(row=r, column=14).value = "☑️ Đã sửa"
        print(f"  ✅ Excel updated row {r}")
        break
wb.save("danh_sach_truyen_doctieuthuyet.xlsx")
print("Done!")
