import openpyxl
import os

excel_path = "danh_sach_truyen_doctieuthuyet.xlsx"

print(f"Opening {excel_path}...")
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Find row for story ID 1933
found = False
for row in range(2, ws.max_row + 1):
    cell_id = ws.cell(row=row, column=2)  # Column B = ID
    if cell_id.value and str(cell_id.value).strip() == "1933":
        print(f"Found story 1933 at row {row}")
        
        # Show current values
        print(f"  Current Title (C): {ws.cell(row=row, column=3).value}")
        print(f"  Current Status: {ws.cell(row=row, column=4).value if ws.cell(row=row, column=4).value else 'N/A'}")
        
        # Update title
        ws.cell(row=row, column=3).value = "Cô Gái Ngồi Cạnh Máy In, Phòng Kế Toán Gọi Tôi Là Đứa Đánh Máy"
        
        # Find "Đã sửa" column - check headers
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and "sửa" in str(header).lower():
                ws.cell(row=row, column=col).value = "✅"
                print(f"  Marked '✅' in column {col} ({header})")
                break
            if header and "ghi chú" in str(header).lower():
                ws.cell(row=row, column=col).value = "REWRITE TOÀN BỘ — chuyển Việt Nam, kiểm toán nội bộ, loại bỏ siêu nhiên"
                print(f"  Updated notes in column {col}")
        
        # Check for loi column
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and "lỗi" in str(header).lower():
                ws.cell(row=row, column=col).value = "Tên TQ, siêu nhiên, banned phrases, cốt truyện vô lý → ĐÃ SỬA"
                print(f"  Updated errors in column {col}")
                break
        
        found = True
        break

if not found:
    print("⚠️ Story 1933 not found in Excel. Searching by title...")
    for row in range(2, ws.max_row + 1):
        title = ws.cell(row=row, column=3).value
        if title and "máy in" in str(title).lower():
            print(f"  Found at row {row}: {title}")

# Print header row to understand columns
print("\nColumn headers:")
for col in range(1, min(ws.max_column + 1, 20)):
    header = ws.cell(row=1, column=col).value
    if header:
        print(f"  Col {col}: {header}")

wb.save(excel_path)
print(f"\n✅ Excel saved!")
