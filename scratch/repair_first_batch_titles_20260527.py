#!/usr/bin/env python3
import importlib.util
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "danh_sach_truyen_doctieuthuyet.xlsx"
EDITOR_PATH = ROOT / "scratch" / "novel_editor.py"

spec = importlib.util.spec_from_file_location("novel_editor", EDITOR_PATH)
editor = importlib.util.module_from_spec(spec)
spec.loader.exec_module(editor)

FIXES = {
    5923: "Bị Đuổi Khỏi Xưởng Trà, Tôi Thâu Tóm Vùng Nguyên Liệu",
    5899: "Bị Đuổi Khỏi Xưởng Tàu, Tôi Lật Sập Gói Thầu Nghìn Tỷ",
    5877: "Bị Đuổi Khỏi Đảo Lý Sơn, Tôi Vạch Trần Tỏi Giả",
    5838: "Bị Khinh Là Thợ Xây, Tôi Thành Kiến Trúc Sư Tự Học",
    5825: "Bị Sếp Đạo Chiến Dịch, Tôi Mở Agency Cướp Lại Khách Lớn",
    5803: "Bị Bán Qua Biên Giới, Tôi Thành Luật Sư Bảo Vệ Trẻ Em",
    5792: "Bị Đuổi Khỏi Salon, Tôi Mở Tiệm Vỉa Hè Khiến Thành Phố Xếp Hàng",
    5780: "Bị Ép Nghỉ Vì Từ Chối Phong Bì, Tôi Lập Viện Riêng",
    5770: "Bị Cha Nuôi Lừa Quỹ Đầu Tư, Tôi Đánh Bại Ông Trên Sàn",
    5757: "Bị Đẩy Khỏi Cuộc Chiến Thừa Kế, Em Gái Út Thâu Tóm Đế Chế",
    5746: "Bị Vu Oan Biển Thủ Mười Tỷ, Tôi Khiến Giám Đốc Phải Quỳ",
    5734: "Con Dâu Bị Khinh Rẻ, Nắm Trọn Đế Chế Khách Sạn",
    5709: "Bác Sĩ Giỏi Về Quê Mở Phòng Khám Miễn Phí",
    5699: "Tiệm Bánh Của Bà Ngoại Chữa Lành Tuổi Thơ",
    5687: "Thầy Giáo Mang Tủ Sách Lên Núi",
    5676: "Quán Cà Phê Nhỏ Giữa Hẻm Sâu Sài Gòn",
    5664: "Bố Mẹ Phá Sản, Tôi Thành Ông Trùm Bất Động Sản Miền Tây",
    5654: "Bị Tạt Axit, Tôi Thành Founder Mỹ Phẩm Cho Người Có Sẹo",
    5641: "Mất Ba Ngón Tay, Tôi Thành Nghệ Sĩ Piano Ở Vienna",
    5630: "Bị Bắt Nạt Vì Mồ Côi, Tôi Thành Nhà Tâm Lý Cứu Trẻ Em",
    5414: "Bị Sa Thải Vì Từ Chối Quảng Cáo Giả, Tôi Lập Agency Mới",
    5270: "Bị Vu Oan Mất 200 Tỷ, Tôi Dùng Sao Kê Blockchain Lật Kèo",
}


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    headers = [c.value for c in ws[4]]
    id_col = headers.index("ID Truyện") + 1
    title_col = headers.index("Tên Truyện") + 1

    editor.upload_helper()
    try:
        for row in range(5, ws.max_row + 1):
            story_id = ws.cell(row, id_col).value
            if story_id in FIXES:
                title = FIXES[story_id]
                res = editor.update_story_meta(story_id, title=title)
                if not res.get("success"):
                    raise RuntimeError(f"Failed {story_id}: {res}")
                ws.cell(row, title_col).value = title
                print(f"fixed {story_id}: {title}")
    finally:
        editor.remove_helper()

    wb.save(XLSX)


if __name__ == "__main__":
    main()
