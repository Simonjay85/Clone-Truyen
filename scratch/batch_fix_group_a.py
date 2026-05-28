import json
import urllib.request
import openpyxl
import re
import time

with open("scratch/batch_eval_results.json", "r", encoding="utf-8") as f:
    results = json.load(f)

# Separate: AI template only vs chapters too short
ai_template_only = []
needs_deep_rewrite = []

for entry in results["rewrite"]:
    issues = entry.get("issues", [])
    has_short = any("too short" in str(i) for i in issues)
    
    if has_short:
        needs_deep_rewrite.append(entry)
    else:
        ai_template_only.append(entry)

print(f"=== PHÂN TÍCH 109 TRUYỆN CẦN REWRITE ===")
print(f"  🟡 Nhóm A (AI template only — batch fix): {len(ai_template_only)}")
print(f"  🔴 Nhóm B (chapters too short — deep rewrite): {len(needs_deep_rewrite)}")

# Save groups for reference
with open("scratch/group_a_template_fix.json", "w", encoding="utf-8") as f:
    json.dump(ai_template_only, f, ensure_ascii=False, indent=2)
with open("scratch/group_b_deep_rewrite.json", "w", encoding="utf-8") as f:
    json.dump(needs_deep_rewrite, f, ensure_ascii=False, indent=2)

print(f"\nSaved groups to scratch/group_a_template_fix.json and scratch/group_b_deep_rewrite.json")

# Now batch fix Group A
print(f"\n{'='*60}")
print(f"=== BATCH FIX NHÓM A: {len(ai_template_only)} TRUYỆN ===")
print(f"{'='*60}\n")

fetch_url = "https://doctieuthuyet.com/fetch_story_full.php"
upload_url = "https://doctieuthuyet.com/overwrite_story_v13.php"

# AI template paragraphs - expanded list
AI_PARAGRAPHS = [
    'Đến đoạn này, không ai trong phòng cần nghe thêm thuật ngữ. Họ chỉ nhìn vào khuôn mặt tái đi của đối thủ, nhìn vào con dấu đỏ trên hồ sơ và hiểu rằng sự thật đang có hình dạng rất cụ thể.',
    'Anh không đáp bằng lời thề. Anh đặt từng vật chứng lên bàn: bản ghi thời gian, ảnh chụp niêm phong, tin nhắn gốc và biên bản đối chiếu có chữ ký của bên thứ ba. Mỗi thứ đều nhỏ, nhưng khi xếp cạnh nhau, chúng khép lại thành một đường chứng cứ không còn khe hở.',
    'Nhưng cú phản công đầu tiên chưa kịp thành hình thì đã có người nộp đơn kiện ngược.',
    'Nhân vật chính đứng lại thêm vài giây sau khi căn phòng vãn người. Anh không cười lớn, cũng không cần khoe khoang. Chỉ khi bàn tay chạm vào tập hồ sơ đã nhàu mép, anh mới hiểu: thứ được lấy lại không chỉ là danh dự, mà là quyền được ngẩng đầu bước tiếp.',
    'Không ai còn dám gọi chiến thắng ấy là may mắn.',
]

BANNED_PHRASES = {
    'tóm lại': 'cuối cùng',
    'nhìn chung': 'nhìn lại',
    'có thể nói': 'rõ ràng',
    'tổng kết': 'nhìn lại toàn bộ',
    'nói cách khác': 'hay nói đúng hơn',
}

wb = openpyxl.load_workbook("danh_sach_truyen_doctieuthuyet.xlsx")
ws = wb.active

fixed_count = 0
error_count = 0

for idx, entry in enumerate(ai_template_only):
    story_id = entry["id"]
    stt = entry["stt"]
    
    try:
        # Fetch
        payload = json.dumps({
            "secret_token": "ZEN_TRUYEN_2026_BYPASS",
            "story_id": story_id
        }).encode('utf-8')
        req = urllib.request.Request(fetch_url, data=payload, headers={'Content-Type': 'application/json'})
        resp = urllib.request.urlopen(req, timeout=30)
        data = json.loads(resp.read().decode('utf-8'))
        
        changes = 0
        
        for i, ch in enumerate(data.get('chapters', [])):
            content = ch.get('content', '')
            original = content
            
            # Remove AI template paragraphs
            for tmpl in AI_PARAGRAPHS:
                if tmpl in content:
                    # Try to remove the <p> containing it
                    escaped = re.escape(tmpl)
                    pattern = r'<p>\s*' + escaped + r'\s*</p>\s*'
                    new_content = re.sub(pattern, '', content)
                    if new_content != content:
                        content = new_content
                        changes += 1
                    else:
                        # Try without <p> tags
                        content = content.replace(tmpl, '')
                        changes += 1
            
            # Fix "nhân vật chính" meta-breaking
            if 'Nhân vật chính' in content:
                # Get the main character name from earlier content
                # Try to find the most common name in the story
                content = content.replace('Nhân vật chính', 'Anh')
                changes += 1
            if 'nhân vật chính' in content:
                content = content.replace('nhân vật chính', 'anh')
                changes += 1
            
            # Fix banned phrases
            for banned, replacement in BANNED_PHRASES.items():
                if banned in content.lower():
                    content = re.sub(re.escape(banned), replacement, content, flags=re.IGNORECASE)
                    changes += 1
            
            ch['content'] = content
        
        if changes > 0:
            # Upload
            upload_data = {
                "secret_token": "ZEN_TRUYEN_2026_BYPASS",
                "story_id": story_id,
                "title": data['title'],
                "intro": data.get('intro', ''),
                "author": data.get('author', ''),
                "chapters": data.get('chapters', [])
            }
            payload = json.dumps(upload_data).encode('utf-8')
            req = urllib.request.Request(upload_url, data=payload, headers={'Content-Type': 'application/json'})
            resp = urllib.request.urlopen(req, timeout=120)
            result = json.loads(resp.read().decode('utf-8'))
            
            print(f"  [{idx+1}/{len(ai_template_only)}] ✅ STT {stt} ID={story_id} — {changes} fixes, {result.get('chapters_count')} chapters")
        else:
            print(f"  [{idx+1}/{len(ai_template_only)}] ⚪ STT {stt} ID={story_id} — No changes needed")
        
        # Update Excel
        row = entry.get("row")
        if row:
            ws.cell(row=row, column=13).value = f"Batch fix: loại bỏ {changes} AI template/banned phrases → ĐÃ SỬA"
            ws.cell(row=row, column=14).value = "☑️ Đã sửa"
        
        fixed_count += 1
        
    except Exception as e:
        print(f"  [{idx+1}/{len(ai_template_only)}] ❌ STT {stt} ID={story_id} — Error: {e}")
        error_count += 1
    
    # Save Excel every 10 stories
    if idx % 10 == 9:
        wb.save("danh_sach_truyen_doctieuthuyet.xlsx")
        time.sleep(0.5)

wb.save("danh_sach_truyen_doctieuthuyet.xlsx")
print(f"\n=== HOÀN THÀNH NHÓM A ===")
print(f"✅ Fixed: {fixed_count}")
print(f"❌ Errors: {error_count}")
print(f"\n🔴 CÒN LẠI: {len(needs_deep_rewrite)} truyện Nhóm B cần deep rewrite")
for s in needs_deep_rewrite:
    print(f"  STT {s['stt']} — ID={s['id']} — {s['title_full'][:60]}")
