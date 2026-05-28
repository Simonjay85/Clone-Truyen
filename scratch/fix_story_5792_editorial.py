#!/usr/bin/env python3
import importlib.util
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
EDITOR_PATH = ROOT / "scratch" / "novel_editor.py"
XLSX = ROOT / "danh_sach_truyen_doctieuthuyet.xlsx"
OUT = ROOT / "scratch" / "story_5792_editorial_fix_result.json"

spec = importlib.util.spec_from_file_location("novel_editor", EDITOR_PATH)
editor = importlib.util.module_from_spec(spec)
spec.loader.exec_module(editor)

STORY_ID = 5792
TITLE = "Bị Đuổi Khỏi Salon, Tôi Mở Tiệm Vỉa Hè Khiến Thành Phố Xếp Hàng"

INTRO = """
<p><strong>"Mày cắt năm mươi nghìn thì salon hai trăm nghìn của tao bán gì? Cầm kéo rẻ tiền thì ra vỉa hè mà cắt!"</strong> Chủ salon Đẹp Plus ném tạp dề của Võ Quốc Bảo xuống sàn trước mặt khách, chỉ vì anh lén giảm giá cho một bác xe ôm không đủ tiền cắt tóc.</p>
<p>Bảy năm đứng sau ghế xoay, Bảo hiểu một điều rất đơn giản: người nghèo cũng muốn soi gương mà thấy mình gọn gàng, tử tế. Nhưng trong salon sang, lòng tự trọng của khách bị tính thêm vào hóa đơn, còn người thợ dám cắt rẻ bị gọi là phá giá.</p>
<p>Bị sa thải, bị tố mất vệ sinh, bị salon cũ thuê người quay clip bôi nhọ, Bảo kéo chiếc ghế nhựa ra vỉa hè, treo tấm bảng "Cắt tóc 30K". Từ ba khách đầu tiên, anh tạo ra hàng dài người chờ, rồi mở tiệm tóc giá rẻ khiến cả thành phố phải nhìn lại giá trị của một đường kéo sạch.</p>
<p>Đây không chỉ là câu chuyện một thợ cắt tóc lật kèo chủ cũ. Đây là hành trình chứng minh cái đẹp không thuộc riêng về người có tiền.</p>
""".strip()


def html(paras):
    return "\n".join(f"<p>{x}</p>" for x in paras)


CHAPTERS = {
    5794: ("Chương 1: Bị Đuổi Khỏi Salon Sang", html([
        "Salon Đẹp Plus ở mặt tiền Nguyễn Thị Thập lúc chín giờ sáng sáng rực như một showroom xe hơi. Kính lau bóng, ghế da Ý, mùi tinh dầu bạc hà trộn với mùi keo xịt tóc cao cấp. Trên bảng giá, dòng cắt tạo kiểu nam bắt đầu từ hai trăm nghìn, chưa tính gội đầu, wax và gói chăm sóc da đầu.",
        "Võ Quốc Bảo đứng ở ghế số ba, tay cầm kéo Nhật đã mòn ở cán. Anh vừa hoàn thành kiểu tóc undercut cho một nhân viên văn phòng trẻ. Đường fade sạch, gáy gọn, mái giữ được độ rơi tự nhiên. Khách soi gương, gật đầu hài lòng. Nhưng mắt Bảo lại liếc về phía cửa, nơi bác Tư xe ôm công nghệ đang lúng túng cầm nón bảo hiểm.",
        "Bác Tư từng là khách quen của Bảo từ hồi anh còn học nghề ở tiệm nhỏ chợ Bà Chiểu. Dạo gần đây bác ít đến, tóc bạc mọc lởm chởm quanh tai. Bác hỏi nhỏ lễ tân liệu có gói cắt nào rẻ hơn không, vì tuần này xe hỏng, tiền chạy app hụt. Lễ tân nhìn bảng giá, mỉm cười đúng quy trình: \"Dạ salon em không có dịch vụ dưới hai trăm ạ.\"",
        "Bảo bước tới. \"Bác ngồi ghế em. Con cắt nhanh cho bác.\"",
        "Lễ tân tái mặt. \"Anh Bảo, khách chưa thanh toán gói.\"",
        "\"Anh trừ vào tiền tip của anh.\"",
        "Bác Tư xua tay. \"Thôi con, bác đi chỗ khác.\"",
        "Bảo đặt tay lên lưng ghế. \"Bác chạy ngoài nắng cả ngày, tóc dài đội nón đau đầu. Ngồi đi. Mười phút.\"",
        "Mười phút ấy đủ làm cả salon im lặng. Bảo không làm qua loa. Anh phủ khăn sạch, xịt khử trùng tông đơ trước mặt bác, cắt từng đường gọn như đang phục vụ khách VIP. Khi bác Tư soi gương, đôi mắt già sáng lên. \"Nhìn đỡ nhếch nhác hẳn,\" bác cười, lục ví lấy tờ năm mươi nghìn.",
        "Chủ salon Trần Văn Hiếu xuất hiện ngay sau lưng. Vest xám, đồng hồ vàng, nụ cười tắt rất nhanh khi thấy tờ tiền trên tay bác Tư. Ông ta không mắng trong phòng riêng. Ông ta mắng ngay giữa sảnh, trước năm khách đang đợi.",
        "\"Bảo, mày cắt năm mươi nghìn thì salon hai trăm nghìn của tao bán gì? Mày phá giá thị trường à?\"",
        "Bảo tháo khăn cho bác Tư, giọng vẫn bình tĩnh. \"Em không phá giá. Em cắt cho người không trả nổi giá cao.\"",
        "\"Đó là vấn đề của ông ta, không phải của salon tao.\" Hiếu chỉ tay ra cửa. \"Salon này bán trải nghiệm cao cấp. Không phải trạm cứu trợ lòng thương. Mày thích làm người tốt thì ra vỉa hè mà cắt.\"",
        "Bác Tư đứng dậy, mặt đỏ bừng vì xấu hổ. Bảo đặt nhẹ tay lên vai bác. Chính khoảnh khắc ấy khiến anh quyết định không xin lỗi.",
        "Hiếu ném chiếc tạp dề đen của Bảo xuống sàn. \"Thu dọn đồ. Từ hôm nay nghỉ. Và nhớ trả lại thẻ nhân viên.\"",
        "Bảo cúi xuống nhặt tạp dề, không vì tiếc việc mà vì trên đó còn dính tóc của những người anh từng làm đẹp. Bảy năm ở Đẹp Plus, anh là thợ có tỉ lệ khách quay lại cao nhất, người sửa những ca cắt hỏng của thợ mới, người khách khó vẫn gọi đích danh. Vậy mà anh rời salon với một bộ tông đơ, hai cây kéo, chiếc lược gãy răng và năm mươi nghìn bác Tư dúi vào tay.",
        "Ngoài vỉa hè, nắng Sài Gòn hắt lên mặt kính salon. Bác Tư đi theo anh, giọng nghẹn: \"Tại bác mà con mất việc.\"",
        "Bảo nhìn tấm kính phản chiếu gương mặt mình: hai mươi bảy tuổi, áo sơ mi đen nhàu, tay vẫn còn mùi cồn sát khuẩn. Anh lắc đầu. \"Không, bác. Tại con nhớ nghề này sinh ra để làm con người ta thấy mình tử tế hơn. Không phải để đo ví tiền trước khi cầm kéo.\"",
        "Bác Tư không biết đáp gì. Bảo đeo túi đồ lên vai, bước khỏi bóng mát của salon. Anh chưa có kế hoạch. Chỉ có đôi tay, bộ kéo cũ và một câu Hiếu vừa ném vào mặt: ra vỉa hè mà cắt.",
    ])),
    5795: ("Chương 2: Chiếc Ghế Nhựa Trên Vỉa Hè", html([
        "Phòng trọ của Bảo nằm trong con hẻm nhỏ ở Bình Thạnh, nơi xe máy phải né nhau bằng cách nghiêng tay lái và mỗi chiều mùi cá kho, nước xả vải, khói nhang trộn vào nhau như một thứ hương riêng của người lao động. Anh thuê căn phòng tám mét vuông trên gác lửng, đủ kê nệm, một kệ đồ nghề và chiếc quạt kêu lạch cạch.",
        "Tối bị đuổi, Bảo không ngủ. Anh lau từng cây kéo, tháo lưỡi tông đơ vệ sinh, xếp lược theo kích cỡ. Những động tác quen thuộc giữ tay anh khỏi run. Trong ví còn bốn triệu hai, tiền nhà sắp đến hạn. Nếu tìm salon khác, anh có thể có việc trong một tuần. Nhưng tin anh bị Đẹp Plus sa thải vì phá giá đã lan trong nhóm chủ salon quận 7. Không ai muốn nhận một thợ bị gắn nhãn gây rối.",
        "Sáng hôm sau, Bảo xuống tiệm tạp hóa mua một tấm bìa carton và bút lông đỏ. Anh viết: <strong>CẮT TÓC 30K - SẠCH, GỌN, KHÔNG CHÊ KHÁCH.</strong> Viết xong, anh nhìn dòng chữ cuối hơi lâu. Không chê khách. Đó là thứ anh từng thấy quá nhiều ở salon sang: chê khách tóc xơ, da đầu dầu, mặc áo bảo vệ, đi dép tổ ong, hỏi giá trước khi ngồi.",
        "Chủ nhà trọ thấy Bảo kéo chiếc ghế nhựa ra trước cửa thì nhíu mày. \"Cắt ở đây công an phường có nói không?\"",
        "\"Con chỉ cắt gọn, không lấn đường. Con dọn sạch tóc.\"",
        "\"Đừng làm bẩn hẻm là được.\"",
        "Khách đầu tiên không phải bác Tư mà là cậu giao hàng tên Hậu, đội nón bảo hiểm đến hằn trán. Hậu đứng nhìn bảng, hỏi ba lần có thật ba mươi nghìn không. Bảo chỉ vào bình xịt cồn, khăn sạch và hộp lưỡi lam dùng một lần. \"Thật. Nhưng nếu muốn uốn nhuộm thì không có. Ở đây chỉ cắt gọn, đẹp vừa mặt.\"",
        "Hậu ngồi xuống ghế nhựa xanh. Ghế hơi thấp, Bảo phải khom lưng nhiều hơn trong salon. Xe máy chạy sát sau lưng, tiếng rao bánh mì vọng từ đầu hẻm. Không có gương lớn, chỉ có chiếc gương treo tạm trên cửa sắt nhà trọ. Nhưng khi Bảo cầm kéo, mọi thứ ồn ào lùi ra xa.",
        "Anh hỏi Hậu chạy ngoài nắng mấy tiếng, có đội nón cả ngày không, có cần tóc không chạm cổ áo đồng phục không. Hậu ngạc nhiên vì một thợ cắt ba mươi nghìn vẫn hỏi kỹ như salon. Bảo cười: \"Giá rẻ không có nghĩa là cắt đại.\"",
        "Mười lăm phút sau, Hậu soi gương, đưa tay vuốt gáy. \"Anh ơi, nhìn em đỡ như mới ngủ dậy hả?\"",
        "\"Đỡ nhiều.\"",
        "Hậu trả ba mươi nghìn, rồi chụp hình bảng gửi vào nhóm shipper. Một giờ sau, hai người nữa đến. Một chú bảo vệ ca đêm muốn cắt cho mát. Một sinh viên hết tiền nhưng hôm sau đi phỏng vấn part-time. Bảo cắt từng người như cắt cho khách hẹn lịch ở salon: khăn riêng, tông đơ xịt cồn, tóc quét gọn vào túi.",
        "Đến trưa, nắng xiên vào hẻm, mồ hôi chảy xuống lưng áo. Bảo ăn vội ổ bánh mì, chưa kịp uống nước thì bác Tư xuất hiện với ba người bạn xe ôm. Bác không nói lời xin lỗi nữa. Bác chỉ chỉ vào bảng: \"Thằng nhỏ này cắt đàng hoàng. Không khinh người nghèo.\"",
        "Câu đó kéo thêm khách hơn mọi quảng cáo. Chiều hôm ấy, Bảo cắt mười bảy đầu. Thu năm trăm mười nghìn. Trừ tiền khăn, cồn, dao lam, nước uống cho khách chờ, còn chưa bằng một ngày lương tốt ở Đẹp Plus. Nhưng khi nhìn hàng ghế nhựa mượn của hàng xóm, nhìn những người lao động ngồi chờ không ngại hỏi giá, Bảo thấy một thứ lâu rồi anh không thấy trong salon: người ta bước vào mà không sợ bị đánh giá.",
        "Tối, anh quét tóc vào bao, lau sạch nền hẻm, trả ghế cho hàng xóm. Chủ nhà trọ đứng trên cầu thang nhìn xuống, không còn nhăn mặt. \"Mai cắt nữa không?\"",
        "Bảo nhìn tấm bảng carton đã cong vì nắng. \"Dạ, nếu còn người cần.\"",
        "Đêm đó, anh không tìm việc salon nữa. Anh lên mạng đặt thêm hai ghế nhựa, một đèn kẹp, một áo choàng cắt tóc loại rẻ nhưng dễ giặt. Trong ô ghi chú giao hàng, anh viết địa chỉ hẻm thật kỹ. Lần đầu tiên sau khi bị đuổi, anh thấy vỉa hè không phải chỗ bị ném ra. Nó có thể là nơi bắt đầu.",
    ])),
    5796: ("Chương 3: Video Triệu View", html([
        "Video đầu tiên được quay bởi một tiktoker tên Mèo Mập, chuyên review đồ ăn vỉa hè. Hôm đó cô ghé hẻm mua bánh tráng trộn, thấy hàng người ngồi chờ cắt tóc trước cửa phòng trọ thì tò mò. Máy quay điện thoại lia qua tấm bảng carton, qua bình xịt cồn, qua đôi tay Bảo đang tỉa phần mái cho một cậu học sinh.",
        "\"Ba mươi nghìn mà kỹ vậy hả anh?\" cô hỏi.",
        "Bảo không nhìn camera. \"Tóc người ta mọc cả tháng mới dài. Cắt hỏng thì họ đội nón cả tháng. Kỹ là chuyện bình thường.\"",
        "Câu nói ấy được cắt thành đoạn mở đầu video. Mèo Mập quay rất gần bàn tay Bảo: ngón cái đẩy kéo, cổ tay xoay nhẹ, lược nâng tóc đúng góc. Không có phông nền đẹp, không nhạc sang, chỉ có tiếng xe máy, tiếng kéo lách tách và gương mặt khách thay đổi dần sau từng đường cắt.",
        "Video lên lúc chín giờ tối. Bảo đang giặt khăn trong nhà vệ sinh chung thì điện thoại rung liên tục. Hậu nhắn: <em>Anh ơi anh lên TikTok rồi.</em> Bác Tư gọi, giọng vui như trúng số: \"Con nổi tiếng rồi nha!\"",
        "Một giờ đầu, video có năm mươi nghìn view. Sáng hôm sau, nó vượt một triệu. Đến trưa, con hẻm nghẹt người. Có sinh viên, shipper, bảo vệ, nhân viên văn phòng tò mò, cả vài người đi ô tô đến chỉ để xem một thợ cắt tóc vỉa hè đang làm gì mà mạng xã hội khen dữ vậy.",
        "Bảo hoảng thật. Anh chỉ có hai tay, hai ghế nhựa, một ổ cắm điện kéo từ phòng trọ. Hàng xóm bắt đầu khó chịu vì xe đậu kín hẻm. Chủ nhà nhắc anh nếu lộn xộn sẽ phải dẹp. Bảo lập tức lấy giấy viết số thứ tự, giới hạn bốn mươi khách một ngày, ai đến sau thì hẹn hôm sau. Có người bực, bảo đang nổi mà chảnh. Anh chỉ cúi đầu xin lỗi.",
        "Mèo Mập quay lại phỏng vấn. \"Anh không tăng giá à? Mười triệu view rồi đó.\"",
        "Bảo đang xịt cồn tông đơ, đáp: \"View không làm bác xe ôm giàu hơn. Em tăng giá vì đông khách thì người đầu tiên bị đẩy ra ngoài vẫn là người cần em nhất.\"",
        "Câu trả lời lại viral lần nữa. Nhưng viral không chỉ mang khách. Nó mang cả soi mói. Một số salon để lại bình luận: <em>Cắt kiểu này vài bữa nấm da đầu.</em> <em>Vỉa hè bụi bặm mà tung hô.</em> <em>Thợ không có bằng gì chắc?</em> Có người chụp ảnh con hẻm đông, gửi lên phường phản ánh lấn chiếm.",
        "Bảo không cãi trên mạng. Anh dùng số tiền kiếm được mua thùng đựng dụng cụ có nắp, khăn trắng đủ đánh số từng chiếc, máy hấp khăn mini cũ và thảm cao su để tóc không bay xuống cống. Anh dán bảng quy trình ngay cạnh bảng giá: dụng cụ khử trùng trước mặt khách, lưỡi dao dùng một lần, khăn giặt riêng, nền quét sau mỗi lượt.",
        "Một cô điều dưỡng bệnh viện gần đó đến cắt tóc cho con trai. Thấy quy trình, cô góp ý thêm cách pha dung dịch sát khuẩn và tặng anh một hộp găng tay y tế. Bảo nhận, ghi tên cô vào sổ cảm ơn. Từ đó, khách chờ không chỉ thấy tay nghề, mà thấy một người thợ nghèo đang cố làm mọi thứ đàng hoàng nhất trong điều kiện ít ỏi.",
        "Đêm cuối tuần, khi hàng khách đã vãn, Bảo ngồi bệt xuống bậc cửa. Lưng đau, cổ tay mỏi, mắt cay vì nắng. Nhưng trên điện thoại, một bình luận làm anh ngồi thẳng lại: <em>Ba tôi làm bảo vệ, cả năm không dám vào salon vì sợ bị nhìn. Hôm nay ba về khoe tóc mới cả buổi.</em>",
        "Bảo đọc bình luận đó cho bác Tư nghe. Bác cười, rít điếu thuốc rồi nói: \"Thấy chưa, cái kéo của con đâu chỉ cắt tóc. Nó cắt bớt cái mặc cảm của người ta.\"",
        "Câu ấy làm Bảo im rất lâu. Anh từng nghĩ mình chỉ muốn kiếm sống sau khi bị đuổi. Nhưng hàng người ngoài hẻm đang cho anh thấy một nhu cầu lớn hơn: thành phố này có quá nhiều người muốn được chăm sóc tử tế mà không phải bước qua một cánh cửa sang trọng khiến họ thấy mình thấp bé.",
    ])),
    5797: ("Chương 4: Salon Đẹp Plus Phản Công", html([
        "Hiếu xem video của Bảo trong văn phòng kính tầng hai Đẹp Plus. Trên màn hình, hàng người xếp trước con hẻm Bình Thạnh dài đến tận quán nước đầu đường. Dưới phần bình luận, cái tên Đẹp Plus bị nhắc đi nhắc lại: <em>Salon đuổi người thợ có tâm.</em> <em>Hai trăm nghìn chưa chắc bằng ba mươi nghìn của anh này.</em>",
        "Hiếu ném điện thoại xuống bàn. Ba tháng trước, Bảo chỉ là một thợ giỏi nhưng dễ thay. Bây giờ, anh ta trở thành câu chuyện làm cả chuỗi salon cao cấp của Hiếu bị đem ra so sánh với một chiếc ghế nhựa.",
        "Đòn đầu tiên của Hiếu là bài Facebook dài. Ông ta không gọi tên Bảo, nhưng ai cũng hiểu. Bài viết nói về trách nhiệm vệ sinh ngành tóc, nguy cơ bệnh da đầu từ dịch vụ vỉa hè, những thợ không được đào tạo bài bản đang phá hỏng chuẩn mực chuyên nghiệp. Cuối bài là ảnh một chiếc khăn bẩn lấy từ mạng, kèm câu: <em>Đừng để giá rẻ đổi bằng sức khỏe.</em>",
        "Bài viết được vài chủ salon chia sẻ. Một số khách bắt đầu hỏi Bảo có giấy phép không, khăn giặt ở đâu, tông đơ có khử trùng thật không. Bảo trả lời từng người, nhưng anh biết trả lời miệng không đủ. Tin bẩn lan nhanh hơn tóc rơi trên nền gạch.",
        "Đòn thứ hai đến khi phường xuống kiểm tra. Hai cán bộ trật tự đô thị yêu cầu Bảo dọn ghế vì có phản ánh tụ tập đông, lấn hẻm, không đảm bảo vệ sinh. Hàng khách đang chờ xôn xao. Có người quay điện thoại. Bảo không chống đối. Anh mời cán bộ xem khu dụng cụ, sổ số thứ tự, bao tóc đã buộc kín, dung dịch sát khuẩn, danh sách khăn sạch và khăn đã dùng.",
        "\"Em biết cắt ở vỉa hè là nhạy cảm,\" Bảo nói. \"Nếu phường yêu cầu dừng, em dừng. Nhưng em xin được hướng dẫn làm đúng. Em không muốn trốn.\"",
        "Cách anh nói làm cán bộ bớt căng. Một chị cán bộ y tế phường đi cùng kiểm tra khăn, hỏi anh khử trùng bằng gì. Bảo đưa chai dung dịch, giấy hướng dẫn và hóa đơn mua máy hấp khăn cũ. Chị nhìn anh một lúc rồi bảo: \"Tự quay quy trình đăng lên. Và đừng để khách tràn ra lòng đường.\"",
        "Bảo làm ngay tối đó. Video không màu mè. Anh quay từ lúc giặt khăn bằng nước nóng, phơi riêng, hấp lại, xịt cồn dụng cụ, thay lưỡi lam, quét tóc, buộc rác. Anh nói trước camera: \"Vệ sinh không phụ thuộc salon hay vỉa hè. Vệ sinh phụ thuộc người thợ có tôn trọng khách không.\"",
        "Video ấy không bùng nổ như video đầu, nhưng nó đổi hướng câu chuyện. Một bác sĩ da liễu bình luận rằng quy trình của Bảo cơ bản ổn hơn nhiều tiệm nhỏ. Một giáo viên dạy nghề tóc đề nghị tặng anh buổi tư vấn chuẩn hóa. Khách hàng bắt đầu quay lại không chỉ vì rẻ, mà vì thấy anh không né câu hỏi khó.",
        "Hiếu không dừng. Ông ta thuê hai người đến gây rối, giả làm khách rồi phàn nàn bị ngứa da đầu sau khi cắt. Một người quay clip la lối ngay trước hẻm. Bảo nhận ra người đó vì sáng cùng ngày anh chưa hề cắt cho anh ta; số thứ tự trong tay là giấy photo sai màu.",
        "Thay vì cãi, Bảo mở sổ khách, chỉ camera an ninh của tiệm tạp hóa đối diện, mời người đó cùng lên phường làm việc. Đám đông im bặt. Người gây rối lắp bắp vài câu rồi bỏ đi. Clip chưa kịp hại Bảo đã bị khách trong hẻm đăng bản đầy đủ lên mạng.",
        "Tối đó, Hiếu gọi cho Bảo. Giọng ông ta không còn vẻ chủ salon nói với thợ. Nó sắc và thấp. \"Mày tưởng vài cái video là thắng à? Nghề này không phải trò từ thiện. Tao sẽ cho mày biết không có mặt bằng, không có thương hiệu, không có tiền thì mày chỉ là thằng cắt tóc ngoài đường.\"",
        "Bảo đứng trong hẻm, nhìn chiếc ghế nhựa đã bạc màu sau nhiều ngày nắng. \"Có thể,\" anh đáp. \"Nhưng khách của em biết họ được tôn trọng ở đâu.\"",
        "Hiếu cúp máy. Bảo biết trận này chưa xong. Muốn giữ khách, giữ hẻm, giữ nghề, anh không thể mãi dựa vào sự thương cảm của mạng xã hội. Anh cần một nơi tử tế hơn chiếc ghế nhựa, nhưng vẫn giữ được lời hứa ban đầu.",
    ])),
    5798: ("Chương 5: Tiệm Tóc Bảo Ra Đời", html([
        "Sáu tháng sau ngày bị đuổi, Bảo có một cuốn sổ tiết kiệm nhỏ và cổ tay đau mãn tính. Cắt ngoài hẻm giúp anh kiếm được tiền, nhưng mỗi ngày là một cuộc vật lộn với nắng, mưa, hàng xóm, trật tự đô thị và chính sức chịu đựng của cơ thể. Chiếc ghế nhựa đã giúp anh bắt đầu, nhưng không thể là mái nhà lâu dài cho lời hứa cắt tóc tử tế.",
        "Mặt bằng anh thuê nằm sát chợ Bà Chiểu, rộng hai mươi mét vuông, trước đây là tiệm sửa điện thoại. Tường loang lổ, nền gạch cũ, cửa cuốn kêu rít. Chủ nhà đòi cọc ba tháng. Bảo ngồi tính đi tính lại, nếu đóng cọc xong anh chỉ còn đủ mua hai ghế second-hand, một gương lớn bị xước góc, máy lạnh cũ và vài bóng đèn.",
        "Bác Tư nghe chuyện, mang đến một phong bì. Bên trong là ba triệu đồng, tiền bác góp cùng mấy anh em xe ôm. Bảo lập tức đẩy lại. Bác nhăn mặt: \"Không cho thì tụi bác giận. Hồi con cắt ngoài hẻm, tụi bác đâu chỉ mua tóc. Tụi bác mua lại cái quyền vô tiệm mà không bị khinh.\"",
        "Bảo nhận tiền, nhưng ghi từng người vào sổ, hẹn khi tiệm ổn sẽ trả bằng thẻ cắt tóc miễn phí. Bác Tư cười: \"Trả bằng cách đừng quên tụi bác là được.\"",
        "Tấm bảng hiệu đầu tiên rất giản dị: <strong>Tiệm Tóc Bảo - Cắt Gọn 50K</strong>. Dưới bảng giá, anh viết thêm: <em>Khăn sạch, dụng cụ khử trùng, không ép gội, không ép mua wax, không chê khách.</em> Dòng cuối làm thợ sơn bảng bật cười. Bảo không xóa.",
        "Ngày khai trương, khách xếp hàng từ bảy giờ sáng. Có bác xe ôm, chú bảo vệ, cô lao công, sinh viên, nhân viên văn phòng từng xem video, cả vài khách cũ của Đẹp Plus mặc áo sơ mi đắt tiền nhưng ngồi rất ngoan trên ghế nhựa chờ tới lượt. Bảo giữ chiếc ghế nhựa cũ đặt ở góc tiệm như ghế chờ đầu tiên.",
        "Giá tăng từ ba mươi lên năm mươi nghìn khiến anh mất ngủ. Anh biết có người sẽ thấy đắt hơn. Nhưng có mặt bằng, điện, nước, máy lạnh, thuế, khăn sạch, lương phụ việc. Nếu cố giữ ba mươi nghìn bằng cách vắt kiệt bản thân, lời hứa tử tế sẽ thành một kiểu tự hành hạ. Anh treo bảng giải thích chi phí ngay cạnh quầy thu ngân.",
        "Khách đọc bảng, đa phần gật đầu. Một chú bảo vệ còn nói: \"Năm mươi mà ngồi phòng mát, khăn sạch, không bị dụ mua dầu gội là quá được rồi.\"",
        "Bảo tuyển thêm hai thợ trẻ. Điều kiện đầu tiên không phải tay nghề mà là thái độ. Một cậu cắt khá nhưng nhăn mặt khi thấy khách công nhân tóc bết mồ hôi. Bảo cho nghỉ sau một ngày thử việc. \"Tóc bẩn gội được,\" anh nói. \"Cái nhìn khinh người khó sửa hơn.\"",
        "Quy trình tiệm được viết trên tường sau quầy: hỏi nghề khách để chọn kiểu dễ sống, báo giá trước khi cắt, dụng cụ khử trùng trước mặt khách, không upsell nếu khách không hỏi, ai không hài lòng được sửa miễn phí trong ba ngày. Những điều ấy nghe đơn giản, nhưng với nhiều khách, nó là lần đầu họ thấy mình được đối xử rõ ràng.",
        "Tháng đầu tiên, tiệm hòa vốn. Tháng thứ hai, bắt đầu có lãi. Nhưng thứ làm Bảo vui nhất không phải doanh thu. Đó là cảnh một bác xe ôm ngồi cạnh một giám đốc công ty, cả hai cùng chờ số, cùng được phủ khăn giống nhau, cùng được hỏi: \"Hôm nay mình muốn gọn kiểu nào?\"",
        "Một buổi tối, Hiếu lái xe ngang qua. Ông ta không vào, chỉ dừng bên kia đường nhìn bảng hiệu. Bảo thấy qua gương nhưng không bước ra. Anh đang cắt tóc cho một cậu học sinh chuẩn bị chụp ảnh thẻ thi tốt nghiệp. Đường kéo cuối cùng ôm gọn gáy, cậu bé soi gương rồi cười rất tươi.",
        "Khoảnh khắc ấy đủ cho Bảo. Tiệm nhỏ, bảng hiệu nhỏ, lợi nhuận chưa nhiều. Nhưng nó đã chứng minh một điều: giá rẻ không đồng nghĩa với rẻ rúng. Nếu người thợ đủ tử tế và quy trình đủ sạch, một tiệm bình dân vẫn có thể khiến khách bước ra với lưng thẳng hơn lúc bước vào.",
    ])),
    5799: ("Chương 6: Hiếu Xin Hợp Tác", html([
        "Sáu tháng sau khi Tiệm Tóc Bảo mở cửa, Đẹp Plus đóng chi nhánh đầu tiên ở quận 7. Lý do công bố là tái cơ cấu, nhưng nhân viên trong ngành đều biết lượng khách nam giảm mạnh. Không phải tất cả chạy sang Bảo, nhưng câu chuyện của anh khiến nhiều người bắt đầu hỏi: vì sao một lần cắt tóc cơ bản lại phải trả quá nhiều cho những thứ họ không cần?",
        "Hiếu gọi điện cho Bảo vào một chiều mưa. Bảo đang sửa lại mái cho một chú công nhân bị hói nhẹ hai bên thái dương. Anh để máy rung đến lần thứ ba mới bắt. Giọng Hiếu ngọt bất thường: \"Bảo à, anh em mình gặp nhau nói chuyện đi. Chuyện cũ bỏ qua.\"",
        "Họ gặp ở một quán cà phê sang gần trung tâm. Hiếu đến với áo sơ mi trắng, vẫn đồng hồ vàng, nhưng mắt thâm và cằm lún râu. Ông ta gọi espresso, Bảo gọi trà đá. Sự chênh lệch ấy làm Hiếu cười nhạt, như thể vẫn muốn giữ thế trên.",
        "\"Anh nói thẳng,\" Hiếu mở laptop, xoay màn hình về phía Bảo. \"Đẹp Plus còn thương hiệu, mặt bằng, hệ thống đặt lịch. Em có hình ảnh thợ tử tế, cộng đồng khách trung thành. Mình hợp tác. Anh mở nhánh bình dân dưới thương hiệu Đẹp Plus Express, em làm giám đốc vận hành.\"",
        "Trên slide, Hiếu đã thiết kế cả logo: kéo bạc, nền xanh, slogan <em>Đẹp chuẩn salon, giá bình dân</em>. Bảo nhìn đến bảng giá thì dừng. Cắt cơ bản: chín mươi chín nghìn. Gội thêm: bốn mươi chín. Tư vấn da đầu: sáu mươi chín. Gói combo khuyến nghị bắt buộc cho khách mới.",
        "\"Đây không phải bình dân,\" Bảo nói. \"Đây là cách lấy khách của tôi rồi bán thêm.\"",
        "Hiếu đặt tách cà phê xuống. \"Em ngây thơ vừa thôi. Năm mươi nghìn không scale được. Tiền mặt bằng, nhân sự, marketing, training, thuế. Em muốn mở rộng thì phải học cách kiếm tiền trên mỗi đầu khách.\"",
        "\"Kiếm tiền không sai. Ép khách mua thứ họ không cần mới sai.\"",
        "Hiếu cười, lần này không giấu được khó chịu. \"Em nghĩ mình là thánh à? Cắt tóc từ thiện cứu đời? Rồi vài năm nữa cổ tay em hỏng, tiền đâu sống?\"",
        "Câu đó đánh trúng nỗi sợ thật. Cổ tay Bảo đã đau. Anh biết mô hình của mình mong manh. Nhưng nếu vì sợ mà bán lại linh hồn tiệm cho Đẹp Plus, chiếc ghế nhựa trong góc kia sẽ chỉ còn là đạo cụ marketing.",
        "Bảo đóng laptop lại. \"Em có thể học cách mở rộng. Nhưng không học cách khinh khách nghèo.\"",
        "Hiếu nhìn anh rất lâu. \"Không có anh, em chỉ là thợ cắt tóc may mắn nổi trên mạng.\"",
        "\"Có thể. Nhưng may mắn không khiến khách quay lại suốt một năm.\"",
        "Cuộc gặp kết thúc trong mưa. Bảo về tiệm, thấy bác Tư đang ngồi sửa dây nón bảo hiểm, chờ tới lượt cắt. Bác hỏi gặp chủ cũ vui không. Bảo bật cười. \"Ông ấy muốn mua cái bảng giá của con.\"",
        "\"Bảng giá mua được. Cái bụng thì khó.\"",
        "Tối hôm đó, Bảo họp hai thợ trẻ. Anh nói rõ tiệm sẽ không nhận nhượng quyền của Đẹp Plus, nhưng nếu muốn giữ giá năm mươi nghìn, họ phải tối ưu: đặt lịch tốt hơn, đào tạo thợ nhanh hơn, mua vật tư chung, giảm lãng phí khăn và điện. Tử tế không có nghĩa là lơ mơ với con số.",
        "Ba tháng sau, Đẹp Plus Express tự mở một điểm cách tiệm Bảo hai cây số. Khuyến mãi rầm rộ, thuê KOL, bảng hiệu đẹp. Nhưng khách đến một lần rồi thôi vì cuối hóa đơn vẫn đội lên gần hai trăm nghìn. Trong khi đó, Tiệm Tóc Bảo vẫn đông, chậm hơn, giản dị hơn, nhưng rõ giá và giữ lời.",
        "Khi chi nhánh thứ ba của Đẹp Plus đóng cửa, Hiếu không gọi nữa. Bảo cũng không hả hê. Anh hiểu thất bại của Hiếu không phải vì salon sang xấu. Nó thất bại vì quên rằng dịch vụ càng chạm vào con người gần bao nhiêu, sự tôn trọng càng không thể giả bấy nhiêu.",
    ])),
    5800: ("Chương 7: Mẹ Và Chiếc Kéo Cũ", html([
        "Chiếc kéo cũ của mẹ Bảo nằm trong hộp kính nhỏ ở góc tiệm. Nó không đẹp. Cán đen tróc sơn, lưỡi kéo đã mòn, một bên ốc vặn thay bằng con vít không cùng màu. Khách mới thường tưởng đó là đồ trang trí kiểu vintage. Bảo luôn nói thật: \"Kéo của mẹ tôi. Cắt ở chợ Bà Chiểu hai mươi năm.\"",
        "Bà Võ Thị Lan từng có một ghế gỗ dưới mái bạt xanh cạnh chợ. Bảng giá viết phấn: nam năm nghìn, nữ mười nghìn, trẻ em ba nghìn. Hồi nhỏ, Bảo tan học thường ra ngồi cạnh mẹ, nhặt tóc rơi, đưa lược, nghe đủ chuyện đời từ những người ngồi trên ghế. Chú bốc vác muốn cắt thật ngắn vì nóng. Cô bán cá muốn tỉa mái để đi đám cưới. Một bà cụ chỉ cần cắt gọn để đỡ vướng khi nằm viện.",
        "Mẹ Bảo không học trường nghề. Bà học từ một bà dì, rồi học thêm bằng cách nhìn đầu người ta thật kỹ. Bà hay nói: \"Cắt tóc không phải cắt cho cái đầu. Cắt cho cái đời người ta đang sống.\" Hồi nhỏ Bảo không hiểu. Lớn lên vào salon sang, anh mới thấy nhiều thợ cắt theo trend, theo ảnh mẫu, theo gói dịch vụ, mà quên hỏi khách làm nghề gì, đội nón ra sao, có thời gian sấy tóc mỗi sáng không.",
        "Năm Bảo mười sáu tuổi, mẹ bị đau khớp tay. Những ngón tay từng cầm kéo cả ngày bắt đầu sưng, co lại khi trời mưa. Bà vẫn cố cắt vì khách quen không nỡ đi chỗ khác. Bảo cầm kéo thay mẹ từ đó. Đường cắt đầu tiên méo đến mức thằng bé hàng xóm đội nón cả tuần. Mẹ không mắng, chỉ bắt anh xin lỗi và cắt miễn phí lần sau.",
        "\"Tay nghề có thể học,\" bà nói. \"Nhưng cắt hỏng mà không dám nhận thì bỏ nghề.\"",
        "Khi Bảo được nhận vào Đẹp Plus, mẹ mừng lắm. Bà giặt áo sơ mi cho anh, bảo vào salon sang phải đứng thẳng, nói năng sạch sẽ, đừng để người ta coi thường thợ chợ. Tháng lương đầu tiên, Bảo mua cho mẹ một đôi dép êm và một chai dầu xoa khớp. Bà cười, nhưng vẫn giữ chiếc kéo cũ trong ngăn tủ.",
        "Ngày Bảo bị Đẹp Plus đuổi, anh không dám về gặp mẹ ngay. Anh sợ bà buồn, sợ bà nghĩ mình ngu vì đánh đổi công việc tốt cho một lần cắt rẻ. Nhưng mẹ nghe hàng xóm kể trước. Bà gọi điện chỉ hỏi: \"Con có cắt ẩu cho người ta không?\"",
        "\"Không.\"",
        "\"Có lấy kéo dơ cắt không?\"",
        "\"Không.\"",
        "\"Có khinh khách không?\"",
        "\"Không.\"",
        "\"Vậy thì ăn cơm đi rồi tính.\"",
        "Câu nói ấy cứu Bảo hơn mọi lời động viên. Mẹ không bảo anh đúng hoàn toàn. Bà chỉ kéo anh về với ba câu gốc của nghề: không ẩu, không bẩn, không khinh người.",
        "Khi tiệm đầu tiên có lãi, Bảo mang chiếc kéo cũ của mẹ về đặt trong hộp kính. Bà Lan đứng nhìn, ngại ngùng: \"Kéo cùn rồi, trưng làm gì?\"",
        "\"Để con nhớ mình bắt đầu từ đâu.\"",
        "Bà gõ nhẹ lên hộp kính. \"Nhớ thì nhớ cả cái đau tay nữa. Làm nghề tử tế thì cũng phải biết giữ thân. Đừng biến lòng tốt thành cách tự giết mình.\"",
        "Lời ấy khiến Bảo thay đổi cách vận hành. Anh bắt đầu giới hạn số khách mỗi thợ một ngày, bắt buộc nghỉ giữa ca, mua băng cổ tay, dạy thợ trẻ tập giãn cơ. Anh hiểu nếu muốn phục vụ người ít tiền lâu dài, người thợ cũng phải được bảo vệ. Một mô hình tốt không thể xây trên đôi tay kiệt sức.",
        "Thỉnh thoảng, mẹ vẫn ghé tiệm. Bà ngồi ở ghế chờ, nhìn những chú bảo vệ, cô giúp việc, sinh viên soi gương sau khi cắt. Có lần một bác lớn tuổi nhận ra bà, hỏi có phải mẹ của ông chủ không. Bà xua tay: \"Ông chủ gì. Nó vẫn là thằng nhỏ nhặt tóc ngoài chợ thôi.\"",
        "Bảo nghe thấy, cười. Anh không thấy bị hạ thấp. Trái lại, đó là danh hiệu anh muốn giữ nhất. Vì nếu quên mình từng nhặt tóc dưới chân mẹ, có lẽ một ngày nào đó anh cũng sẽ ngồi trong phòng kính và hỏi người nghèo rằng họ lấy quyền gì đòi đẹp.",
    ])),
    5801: ("Chương 8: Chuỗi Tiệm Tóc Bảo - Cắt Đẹp Giá Rẻ", html([
        "Hai năm sau, Tiệm Tóc Bảo có năm chi nhánh: Bình Thạnh, Gò Vấp, Tân Bình, Thủ Đức và quận 12. Không chi nhánh nào nằm ở mặt tiền sang. Bảo chọn gần chợ, gần khu trọ, gần bến xe buýt, nơi người lao động đi ngang có thể nhìn bảng giá mà không thấy ngại.",
        "Mỗi chi nhánh chỉ ba ghế. Tường trắng, gương lớn, đèn đủ sáng, khăn sạch xếp trong tủ kính. Bảng giá giống nhau: cắt gọn năm mươi nghìn, trẻ em ba mươi, người già neo đơn miễn phí sáng thứ Hai, không gội ép, không combo ẩn. Dưới bảng giá vẫn có dòng cũ: <em>Không chê khách.</em>",
        "Mở rộng không dễ. Chi nhánh Gò Vấp từng có thợ tự ý gợi ý khách mua wax để kiếm thêm. Bảo phát hiện qua phản ánh của một sinh viên. Anh không mắng trước khách, nhưng tối đó họp toàn bộ chi nhánh. \"Mình bán wax được thêm vài chục nghìn,\" anh nói, \"nhưng mất thứ khiến khách dám bước vào: cảm giác hóa đơn không phục kích họ.\" Người thợ bị nhắc lỗi, tái phạm lần hai thì nghỉ.",
        "Chi nhánh Thủ Đức từng bị phàn nàn cắt vội vì khách đông. Bảo đóng đặt lịch online một ngày, tự đến đứng cắt cùng thợ, đo lại thời gian từng kiểu tóc. Anh nhận ra nếu giữ giá thấp mà nhận quá nhiều khách, chất lượng sẽ tụt. Từ đó, mỗi thợ tối đa hai mươi lăm khách một ngày, hết số thì hẹn. Có khách bực, nhưng đa số hiểu.",
        "Bảo lập lớp đào tạo thợ trẻ vào mỗi tối thứ Tư. Học viên phần lớn là con em lao động, có người từng bỏ học, có người làm phụ quán tóc gội đầu nhiều năm mà không được cầm kéo. Bài học đầu tiên không phải kỹ thuật fade mà là cách hỏi khách: làm nghề gì, đội nón nhiều không, có muốn dễ tự vuốt không, ngân sách chăm tóc mỗi tháng bao nhiêu. Anh bắt học viên tập nói giá rõ ràng trước khi cắt.",
        "Một nhà đầu tư từng tìm đến, đề nghị rót vốn để mở ba mươi chi nhánh trong một năm. Họ muốn tăng giá lên tám mươi chín nghìn, thêm gói thành viên, bán dầu gội thương hiệu riêng và đặt KPI doanh thu trên mỗi khách. Bảo nghe hết, cảm ơn, rồi từ chối. Minh, quản lý vận hành mới, lo lắng vì đó là cơ hội lớn. Bảo chỉ hỏi: \"Nếu bác Tư vào chi nhánh mới, cuối hóa đơn bác trả bao nhiêu?\" Minh im lặng.",
        "Không nhận vốn lớn nghĩa là tăng chậm. Có tháng dòng tiền căng vì mua khăn, sửa máy lạnh, đào tạo thợ mới. Bảo vẫn giữ một cuốn sổ tay ghi lý do từng quyết định. Anh sợ một ngày mình cũng bị con số kéo đi xa khỏi chiếc ghế nhựa ban đầu.",
        "Nhưng tăng chậm có cái chắc. Mỗi chi nhánh đều có khách quen gọi tên thợ. Chú bảo vệ ở Tân Bình cắt đúng sáng thứ Ba. Cô công nhân ở Thủ Đức dẫn con trai đến mỗi cuối tháng. Một nhóm sinh viên rủ nhau cắt trước ngày bảo vệ đồ án. Họ không gọi đó là chuỗi salon. Họ gọi là tiệm của Bảo, dù Bảo không thể có mặt ở mọi nơi.",
        "Ngày chi nhánh thứ năm khai trương, bác Tư đến cắt băng bằng chiếc kéo cũ của mẹ Bảo. Báo địa phương chụp ảnh, hỏi Bảo bí quyết cạnh tranh với salon cao cấp. Anh đáp: \"Tôi không cạnh tranh bằng đèn đẹp hơn. Tôi cạnh tranh bằng việc để người ít tiền vẫn được hỏi muốn tóc mình trông ra sao.\"",
        "Bài báo ấy được chia sẻ nhiều. Có người khen, có người bảo mô hình này không làm giàu lớn. Bảo không phủ nhận. Lợi nhuận của Tiệm Tóc Bảo vừa đủ, không hào nhoáng. Nhưng mỗi ngày, hàng trăm người bước vào, trả số tiền họ trả được, và bước ra với gáy sạch, mái gọn, lưng thẳng hơn một chút.",
        "Với Bảo, như vậy là giàu theo cách khác.",
    ])),
    5802: ("Chương 9: Chiếc Ghế Nhựa Vẫn Còn", html([
        "Chiếc ghế nhựa xanh vẫn nằm trong kho chi nhánh Bình Thạnh, chân ghế trầy xước, lưng ghế bạc màu vì nắng. Nhân viên nhiều lần đề nghị thay bằng ghế chờ đẹp hơn. Bảo không cho. Anh bảo ghế ấy không dùng để ngồi hằng ngày nữa, nhưng mỗi Chủ nhật anh sẽ mang nó ra vỉa hè.",
        "Chủ nhật đầu tiên của mỗi tháng là ngày cắt miễn phí. Không quảng cáo rầm rộ, chỉ một tấm bảng nhỏ trước tiệm: <strong>Cắt tóc miễn phí cho người vô gia cư, bán vé số, xe ôm khó khăn - từ 7h đến 11h.</strong> Bảo tự đứng cắt, cùng vài thợ tự nguyện. Ai đến cũng được phát nước, khăn sạch, số thứ tự. Không ai bị chụp ảnh nếu không đồng ý.",
        "Có lần một nhóm sinh viên truyền thông xin quay phóng sự. Bảo từ chối quay cận mặt khách. \"Làm thiện nguyện không cần biến người nghèo thành hình minh họa,\" anh nói. Nếu họ muốn quay, chỉ quay tay thợ, dụng cụ, quy trình, hoặc phỏng vấn người đồng ý. Câu đó khiến cả team trẻ hơi ngượng, nhưng sau đó họ làm một video rất tử tế.",
        "Một sáng Chủ nhật, Hiếu xuất hiện. Ông ta đứng bên kia đường, không còn vest bóng bẩy, chỉ mặc áo polo xám. Đẹp Plus sau tái cơ cấu còn vài điểm cao cấp, nhưng hào quang cũ đã nhạt. Bảo thấy ông ta, gật đầu chào. Hiếu bước sang, nhìn chiếc ghế nhựa.",
        "\"Mày vẫn làm trò này à?\" giọng ông ta không còn sắc như trước.",
        "\"Không phải trò. Lịch cố định.\"",
        "Hiếu im lặng nhìn một ông cụ bán vé số ngồi xuống ghế. Ông cụ tóc bạc rối, áo sơ mi sờn cổ, tay ôm xấp vé. Bảo phủ khăn, hỏi ông muốn cắt gọn hay giữ mái che trán. Ông cụ cười ngại: \"Sao cũng được, miễn nhìn đỡ bị người ta đuổi khỏi quán cà phê.\"",
        "Bảo cắt rất chậm. Không phải vì kiểu khó, mà vì ông cụ cứ nghiêng đầu theo thói quen né tránh. Anh phải đặt tay thật nhẹ sau gáy, nói: \"Bác cứ ngồi thẳng. Ở đây không ai đuổi bác.\"",
        "Hiếu nghe câu đó, mặt thoáng đổi. Có lẽ ông ta nhớ chính mình từng đuổi bác Tư, từng ném tạp dề của Bảo xuống sàn vì một lần cắt năm mươi nghìn. Ông ta không xin lỗi. Người như Hiếu không dễ nói hai chữ ấy. Nhưng khi ông cụ bán vé số soi gương và bật cười như trẻ con, Hiếu quay mặt đi rất lâu.",
        "Cắt xong, ông cụ lục túi lấy mấy tờ vé số dúi cho Bảo. Anh không nhận tiền, nhưng mua hai tờ vé số. Ông cụ nắm tay anh: \"Cảm ơn con. Lâu rồi bác không dám soi gương lâu như vậy.\"",
        "Câu nói ấy làm cả nhóm thợ trẻ im lặng. Một người mới vào nghề sau đó hỏi Bảo: \"Anh ơi, có đáng không? Cả sáng mình cắt miễn phí, trong khi tiệm vẫn phải trả điện nước.\"",
        "Bảo không trả lời ngay. Anh tháo lưỡi dao đã dùng, bỏ vào hộp an toàn, xịt cồn tông đơ. Rồi anh chỉ chiếc ghế nhựa. \"Ngày xưa anh ngồi sau cái ghế đó và hiểu một chuyện: người ta không chỉ trả tiền để bớt tóc. Người ta trả tiền để được đối xử như mình còn giá trị. Có người không đủ tiền trả, thì thỉnh thoảng mình trả lại cho họ một chút.\"",
        "Buổi trưa, khi dọn đồ, Hiếu vẫn đứng đó. Ông ta nhìn Bảo, nói rất khẽ: \"Tao từng nghĩ mày phá giá.\"",
        "Bảo buộc bao tóc lại. \"Em từng nghĩ anh ghét em. Sau này em hiểu anh sợ. Sợ khách nhận ra vài thứ không cần đắt như mình nói.\"",
        "Hiếu cười nhạt, nhưng không phản bác. Trước khi đi, ông ta đặt một hộp kéo mới lên bàn. \"Hàng dư của chi nhánh đóng cửa. Nếu dùng được thì dùng.\"",
        "Bảo nhìn hộp kéo. Anh không từ chối. Có những lời xin lỗi không đủ đẹp để lên sân khấu, nhưng vẫn có thể dùng được nếu biến thành việc có ích. Anh đưa hộp kéo cho thợ phụ kiểm tra, rồi kéo chiếc ghế nhựa vào trong.",
        "Chiều xuống, tiệm trở lại nhịp bình thường. Khách đặt lịch đến cắt, trẻ con chạy quanh ghế chờ, bác Tư ghé uống trà đá. Bảo treo lại chiếc kéo của mẹ trong tủ kính, đặt chiếc ghế nhựa ngay bên dưới. Một cái nhắc anh về gốc nghề, một cái nhắc anh về ngày bị ném ra khỏi salon.",
        "Ngoài cửa, thành phố vẫn ồn, vẫn đắt đỏ, vẫn có những nơi người nghèo ngại bước vào. Nhưng ở một góc nhỏ sát chợ, tấm bảng năm mươi nghìn vẫn sáng đèn. Bảo gọi khách tiếp theo, giọng như ngày đầu ở vỉa hè: \"Mời bác ngồi. Hôm nay mình muốn gọn kiểu nào?\"",
    ])),
}


def main():
    updates = []
    editor.upload_helper()
    try:
        meta = editor.update_story_meta(STORY_ID, title=TITLE, intro=INTRO)
        if not meta.get("success"):
            raise RuntimeError(meta)
        updates.append({"story_id": STORY_ID, "meta": "updated"})
        for chapter_id, (title, content) in CHAPTERS.items():
            res = editor.update_chapter(chapter_id, title, content)
            if not res.get("success"):
                raise RuntimeError(f"Chapter update failed {chapter_id}: {res}")
            updates.append({"chapter_id": chapter_id, "title": title})
    finally:
        editor.remove_helper()

    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    headers = [c.value for c in ws[4]]
    col = {v: i + 1 for i, v in enumerate(headers)}
    for row in range(5, ws.max_row + 1):
        if ws.cell(row, col["ID Truyện"]).value == STORY_ID:
            ws.cell(row, col["Tên Truyện"]).value = TITLE
            ws.cell(row, col["Điểm Cần Sửa"]).value = "Đã đọc live và viết lại/mở rộng 9 chương lên chuẩn 1000-1500 từ; tăng phản công salon, kiểm tra vệ sinh, mở tiệm và kết dư âm."
            ws.cell(row, col["Trạng Thái Sửa"]).value = "☑️ Đã sửa"
            break
    wb.save(XLSX)

    OUT.write_text(json.dumps({"success": True, "updates": updates}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"success": True, "updates": updates}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
