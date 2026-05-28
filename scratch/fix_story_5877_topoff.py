import importlib.util
import json
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
spec = importlib.util.spec_from_file_location("novel_editor", ROOT / "scratch" / "novel_editor.py")
editor = importlib.util.module_from_spec(spec)
spec.loader.exec_module(editor)

STORY_ID = 5877
TITLE = "Bị Ép Nhường Công Thức Tỏi Lý Sơn, Tôi Dùng Một Mẫu Kiểm Định Vả Sập Tập Đoàn Tỏi Giả"

ADDITIONS = {
    5967: """
Phước không lao lên mạng để chửi lại. Anh biết một câu phẫn nộ có thể bị cắt thành bằng chứng rằng anh mất kiểm soát, còn một bảng số liệu đúng sẽ buộc người ta phải đọc đến cuối. Anh nhờ ba hộ trồng tỏi lâu năm ký xác nhận quy trình canh tác, chụp từng luống ruộng có tọa độ, niêm phong mẫu đất ngay trước mặt đại diện hợp tác xã. Mỗi túi mẫu đều có chữ ký của người chứng kiến, giờ lấy mẫu, độ ẩm và ảnh đối chiếu.

Tối đó, khi livestream của tập đoàn Đông Á vẫn gọi anh là kẻ phá hoại đặc sản quê nhà, Phước mở một buổi công bố nhỏ ngay tại nhà văn hóa thôn. Không có sân khấu, chỉ có máy chiếu cũ và mấy hàng ghế nhựa. Anh chiếu từng trang: mẫu tỏi thật từ ruộng Lý Sơn có dấu khoáng khác với tỏi nhập trộn, hàm lượng lưu huỳnh bay hơi lệch rõ, vỏ củ bị xử lý trắng quá mức. Một bác nông dân đứng dậy, giọng run run: "Tụi nó nói tụi tôi tham tiền bán tỏi giả. Nhưng tụi tôi là người bị mượn tên." Cả phòng im phăng phắc.

Khoảnh khắc ấy, câu chuyện không còn là danh dự của riêng Phước. Nó trở thành danh dự của cả những người đã phơi lưng trên đất đảo rồi bị doanh nghiệp dùng thương hiệu quê mình làm bình phong. Phước nhìn thẳng vào camera của phóng viên địa phương và nói chậm từng chữ: "Tôi không phá đặc sản Lý Sơn. Tôi đang trả lại tên Lý Sơn cho đúng người trồng ra nó."
""",
    5968: """
Người phản bội quay đầu không đến với vẻ anh hùng. Tấn, nhân viên kho từng mở ổ cứng của Phước, xuất hiện sau nhà thờ họ lúc gần nửa đêm, áo ướt mồ hôi, tay ôm một phong bì nâu. Cậu ta không dám nhìn thẳng. "Em nhận tiền của bà Kim," Tấn nói. "Em tưởng chỉ lấy bản nháp công thức, không ngờ họ dùng nó để dựng hồ sơ giả." Phước không vội tha thứ. Anh bật ghi âm, gọi trưởng thôn và luật sư ra làm chứng, yêu cầu Tấn kể lại từ đầu.

Trong phong bì là bản kê xe container, ảnh kho trung chuyển ở Bình Định và chuỗi email nội bộ bàn cách trộn tỏi nhập với tỏi đảo theo tỷ lệ đủ để qua mắt người mua. Có một email ghi rõ: "Dùng tên Lâm Hữu Phước để hợp thức hóa quy trình vi sinh." Phước đọc đến dòng đó thì bàn tay siết chặt. Họ không chỉ cướp công thức; họ định biến anh thành con dấu sống cho một đường dây lừa đảo.

Tấn bật khóc khi ký vào biên bản. Cậu ta nói mẹ đang bệnh, tiền nợ đè quá nên nhận việc. Phước im lặng rất lâu rồi mới đáp: "Nghèo không làm cậu vô tội. Nhưng quay đầu đúng lúc có thể cứu nhiều người khỏi bị kéo xuống cùng." Anh photo toàn bộ hồ sơ, một bộ gửi luật sư, một bộ gửi cơ quan kiểm định, một bộ giữ trong két của hợp tác xã. Lần này, bằng chứng không còn nằm trong tay một người có thể bị bôi nhọ. Nó đã được chia thành nhiều điểm sáng, đủ để không ai dập tắt được cùng lúc.
""",
    5969: """
Ngày kết luận kiểm định được công bố, gió từ biển thổi qua sân ủy ban mang theo mùi muối và mùi đất vừa tưới. Phước không đứng ở hàng ghế đầu. Anh đứng phía sau những người trồng tỏi, để họ nghe tên mình được trả lại trước. Bản kết luận ghi rõ lô hàng của Đông Á có dấu hiệu pha trộn nguồn gốc, sử dụng nhãn Lý Sơn sai quy chuẩn và khai thác trái phép dữ liệu quy trình của nhóm nghiên cứu địa phương. Mỗi dòng đọc lên, gương mặt bà Kim và đại diện tập đoàn càng tái đi.

Nhưng cú vả mạnh nhất không phải là lúc họ bị phóng viên vây kín. Cú vả thật sự là khi các thương lái từng ép giá nông dân phải quay lại đặt mua tỏi đúng nguồn, ký hợp đồng minh bạch và chấp nhận kiểm định định kỳ. Phước đề nghị lập mã truy xuất cho từng ruộng: ai trồng, ngày xuống giống, ngày thu hoạch, kết quả đất, kết quả mẫu củ. Anh không muốn thắng một vụ rồi để năm sau người khác lại trộn hàng giả dưới một cái tên khác.

Buổi chiều, Phước mang một nhúm đất ruộng đặt lên bàn thờ cha. Anh nhớ lời cha từng nói: tỏi Lý Sơn không chỉ cay vì giống, mà cay vì đất đá núi lửa, gió biển và mồ hôi người trồng. Anh từng nghĩ câu ấy chỉ là cách người già yêu quê nói cho đẹp. Sau tất cả, anh hiểu nó là một công thức mà phòng thí nghiệm nào cũng phải cúi đầu: hóa học của đất, và đạo đức của người giữ đất. Khi Minh Anh đứng cạnh anh, không nói xin quay lại, chỉ lặng lẽ đặt bản hợp đồng bảo vệ thương hiệu cộng đồng lên bàn, Phước ký tên. Lần này, chữ ký của anh không nhường ai một công thức. Nó giữ lại một quê hương.
""",
}


def main():
    editor.upload_helper()
    try:
        payload = editor.get_story_chapters(STORY_ID)
        chapters = payload["chapters"] if isinstance(payload, dict) else payload
        updates = []
        for chapter in chapters:
            cid = int(chapter["id"])
            addition = ADDITIONS.get(cid)
            if not addition:
                continue
            content = (chapter["content"] or "").rstrip() + "\n\n" + addition.strip()
            editor.update_chapter(cid, chapter["title"], content)
            updates.append({"chapter_id": cid, "title": chapter["title"]})
    finally:
        editor.remove_helper()

    workbook = ROOT / "danh_sach_truyen_doctieuthuyet.xlsx"
    wb = openpyxl.load_workbook(workbook)
    ws = wb.active
    for row in range(5, ws.max_row + 1):
        if ws.cell(row, 2).value == STORY_ID:
            ws.cell(row, 3).value = TITLE
            ws.cell(row, 13).value = (
                "Đã đọc live; giữ khung truyện tốt, bổ sung chương 6-8 cho đủ 1000-1500 từ live, "
                "tăng bằng chứng kiểm định, phản bội quay đầu và kết bảo vệ thương hiệu tỏi Lý Sơn."
            )
            ws.cell(row, 14).value = "☑️ Đã sửa"
            break
    wb.save(workbook)

    out = ROOT / "scratch" / "story_5877_fix_result.json"
    out.write_text(json.dumps(updates, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(updates, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
