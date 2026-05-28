import json
import urllib.request

story_data = {
    "secret_token": "ZEN_TRUYEN_2026_BYPASS",
    "story_id": 1948,
    "title": "Thực Tập Sinh Bị Đuổi Giữa Phòng Họp, Không Ai Biết Cô Là Con Gái Chủ Tịch",
    "intro": "<p>Trần Bảo Ngọc, con gái duy nhất của Chủ tịch Tập đoàn Bất động sản Đại Phong, cắt tóc ngắn, mặc đồ giản dị, xin vào làm thực tập sinh ở chính công ty của cha mình — với cái tên giả: Nguyễn Mai Anh.</p>\n<p>Cô muốn hiểu công ty từ bên trong, trước khi nhận quyền kế thừa. Nhưng điều cô không ngờ là trưởng phòng kinh doanh Phạm Quốc Huy — người được cha cô tin tưởng nhất — đang cùng đối tác ngoài bòn rút dự án trăm tỷ.</p>\n<p>Khi Bảo Ngọc vô tình phát hiện chứng cứ, Huy ra lệnh đuổi cô giữa phòng họp: \"Thực tập sinh thì biết gì về kinh doanh?\"</p>\n<p>Hắn không biết rằng cô gái hắn vừa nhục mạ sẽ là người ký quyết định kỷ luật hắn.</p>",
    "author": "Nguyễn Hồng Phúc",
    "seo": {
        "focus_keyword": "thực tập sinh, con gái chủ tịch, vả mặt trưởng phòng",
        "seo_title": "Thực Tập Sinh Bị Đuổi Giữa Phòng Họp — Truyện Sảng Văn | Đọc Tiểu Thuyết",
        "seo_description": "Con gái Chủ tịch tập đoàn giả thực tập sinh, phát hiện trưởng phòng cướp dự án trăm tỷ. Khi bị đuổi giữa phòng họp, cô lật bài."
    },
    "chapters": []
}

ch1 = {
    "title": "Chương 1: Nguyễn Mai Anh — Thực Tập Sinh Quèn",
    "content": """<p>Trần Bảo Ngọc đứng trước gương trong nhà vệ sinh tầng trệt Tập đoàn Đại Phong, nhìn gương mặt mình lần cuối trước khi bước vào vai diễn.</p>

<p>Tóc cắt bob ngang vai, không nhuộm. Áo sơ mi trắng Uniqlo mua sale, quần tây đen Zara giá giảm, giày bệt bata. Trên ngực áo gắn thẻ nhân viên mới in sáng nay: "Nguyễn Mai Anh — Thực tập sinh — Phòng Kinh doanh."</p>

<p>Cô hai mươi bốn tuổi, tốt nghiệp thạc sĩ Quản trị Kinh doanh tại RMIT Melbourne, nói được bốn thứ tiếng, từng thực tập tại Cushman & Wakefield ở Sydney hai mùa hè. Cha cô — Trần Đình Khoa, Chủ tịch HĐQT Tập đoàn Bất động sản Đại Phong — sở hữu mười hai dự án bất động sản trải dài từ Hà Nội đến Phú Quốc, tổng giá trị danh mục hơn hai nghìn tỷ đồng.</p>

<p>Nhưng hôm nay, Bảo Ngọc không phải con gái Chủ tịch. Cô là Nguyễn Mai Anh, sinh viên mới ra trường, không quen biết ai, không có mối quan hệ, xin vào thực tập bằng CV gửi qua email chung.</p>

<p>Cô hít một hơi sâu, đẩy cửa nhà vệ sinh, và bước vào ngày đầu tiên.</p>

<hr>

<p>Phòng kinh doanh Đại Phong nằm ở tầng tám, chiếm gần nửa diện tích sàn. Ba mươi hai nhân viên ngồi theo hình chữ U quanh một bàn lớn ở giữa — nơi trưởng phòng Phạm Quốc Huy thường đứng mỗi sáng thứ Hai để "truyền cảm hứng".</p>

<p>Bảo Ngọc — hay đúng hơn, Mai Anh — được chỉ ngồi ở góc cuối phòng, bàn nhỏ nhất, kế bên máy photocopy. Cô nhận ra ngay: đây là cái bàn dành cho người không ai muốn nhìn thấy.</p>

<p>"Mai Anh đúng không? Ngồi đây nhé. Tuần đầu em hỗ trợ phòng — photocopy tài liệu, đặt cà phê cho anh chị, và nhập dữ liệu khách hàng vào CRM."</p>

<p>Người nói là Hoàng Thị Phương Linh, phó phòng kinh doanh, ba mươi mốt tuổi, tóc nhuộm nâu đỏ, móng tay gel, đeo kính Chanel. Chị ta nói giọng Hà Nội nhẹ, nhưng ánh mắt nhìn Mai Anh như nhìn một món đồ trang trí không cần thiết.</p>

<p>"Dạ, em hiểu. Nhưng em cũng muốn hỏi — em có được tham gia họp dự án không ạ?"</p>

<p>Phương Linh hơi ngạc nhiên, rồi cười nhẹ — kiểu cười mà người ta dùng khi nghe một đứa trẻ hỏi liệu nó có thể lái xe.</p>

<p>"Em mới vào tuần đầu, Mai Anh. Cứ quen với công việc đã rồi tính."</p>

<p>Bảo Ngọc gật đầu, ngồi xuống, mở laptop. Cô không nói gì thêm.</p>

<p>Nhưng trong lúc ngồi nhập dữ liệu CRM, cô đã bắt đầu quan sát.</p>

<p>Trưởng phòng Phạm Quốc Huy bước vào lúc chín giờ mười lăm — muộn mười lăm phút, nhưng không ai nhắc. Anh ta ba mươi tám tuổi, cao ráo, mặc vest xanh đen, đeo đồng hồ Omega Seamaster. Anh ta bắt tay từng người, vỗ vai, cười rạng rỡ — kiểu lãnh đạo biết cách làm mọi người cảm thấy được coi trọng.</p>

<p>Khi đi ngang bàn Mai Anh, anh ta dừng lại.</p>

<p>"Thực tập sinh mới hả? Chào em. Anh là Huy, trưởng phòng. Cần gì cứ hỏi anh."</p>

<p>"Dạ, cảm ơn anh Huy."</p>

<p>Anh ta mỉm cười rồi đi vào phòng riêng — cái phòng kính trong suốt ở góc, có ghế da, bàn gỗ sồi, và một bức ảnh chụp chung với Chủ tịch Trần Đình Khoa treo trên tường.</p>

<p>Bảo Ngọc nhìn bức ảnh đó. Cha cô đang bắt tay Huy, cả hai cười tươi. Ảnh chụp tại lễ ký kết dự án Đại Phong Marina năm ngoái — dự án hai trăm tỷ đồng ở Nha Trang.</p>

<p>Cô biết dự án đó. Cô cũng biết rằng ba tháng qua, tiến độ giải ngân dự án Marina đang có vấn đề — cha cô đã đề cập trong bữa tối gia đình, nhưng không nói chi tiết.</p>

<p>Đó chính là lý do cô ở đây.</p>

<p>Không phải để pha cà phê và photocopy. Mà để tìm hiểu xem hai trăm tỷ đồng đang chảy đi đâu.</p>"""
}

ch2 = {
    "title": "Chương 2: Hóa Đơn Không Khớp",
    "content": """<p>Tuần đầu tiên, Mai Anh làm đúng những gì được giao: photocopy, nhập CRM, đặt trà sữa cho phòng mỗi chiều thứ Sáu. Cô không hỏi nhiều, không tỏ vẻ thông minh, không gây chú ý.</p>

<p>Nhưng cô quan sát.</p>

<p>Cô để ý rằng mỗi thứ Ba và thứ Năm, Phạm Quốc Huy đều có cuộc họp kín trong phòng riêng với một người đàn ông tên Vũ Hoàng Sơn — Giám đốc Công ty TNHH Đầu tư Sơn Hải, đối tác thi công chính của dự án Đại Phong Marina.</p>

<p>Cô để ý rằng sau mỗi cuộc họp, Huy thường gọi phó phòng Phương Linh vào, nói gì đó rồi Phương Linh soạn email gửi cho phòng tài chính. Những email đó luôn có tiêu đề: "Thanh toán đợt [X] — Dự án Marina."</p>

<p>Cô để ý rằng tiến độ thi công trên báo cáo tuần — file được gửi cho cả phòng kinh doanh qua email — luôn ghi "đạt 95% kế hoạch." Nhưng khi cô kiểm tra ảnh công trường trên trang Zalo nội bộ của công ty, cô thấy hiện trường vẫn còn là bãi đất trống với vài cọc bê tông.</p>

<p>Tuần thứ hai, cô quyết định đi một bước xa hơn.</p>

<hr>

<p>Chiều thứ Tư, khi phòng kinh doanh đi ăn trưa, Mai Anh ở lại văn phòng với lý do "em cần hoàn thành nhập CRM." Cô đợi cho đến khi người cuối cùng ra khỏi phòng, rồi đi đến máy in chung.</p>

<p>Trên khay giấy máy in có một tập tài liệu vừa được in ra và chưa ai lấy — thói quen xấu mà cô đã quan sát suốt tuần qua: nhân viên thường in tài liệu rồi quên nhấc.</p>

<p>Tập tài liệu này ghi: "Hợp đồng phụ kiện — Dự án Đại Phong Marina — Gói thầu số 07: Hệ thống PCCC."</p>

<p>Mai Anh lật qua từng trang. Hợp đồng ký giữa Đại Phong và một công ty tên Công ty TNHH Phòng cháy An Tâm, trụ sở tại quận Tân Phú, TP.HCM. Giá trị hợp đồng: mười bốn tỷ năm trăm triệu đồng.</p>

<p>Cô chụp ảnh trang đầu và trang cuối bằng điện thoại cá nhân, rồi tra cứu Công ty Phòng cháy An Tâm trên Cổng thông tin Doanh nghiệp.</p>

<p>Kết quả: Công ty thành lập cách đây tám tháng. Vốn điều lệ: năm trăm triệu. Giám đốc: Vũ Thị Hồng Nhung.</p>

<p>Vũ Thị Hồng Nhung. Cô gõ tên này vào Facebook.</p>

<p>Kết quả đầu tiên: một phụ nữ ba mươi hai tuổi, sống ở quận 7, TP.HCM. Trong danh sách bạn bè — Vũ Hoàng Sơn, Giám đốc Sơn Hải.</p>

<p>Mai Anh kiểm tra thêm: Vũ Thị Hồng Nhung là em gái ruột của Vũ Hoàng Sơn.</p>

<p>Một công ty PCCC vừa thành lập tám tháng, vốn năm trăm triệu, do em gái đối tác thi công làm giám đốc, nhận hợp đồng mười bốn tỷ rưỡi từ dự án Marina.</p>

<p>Bảo Ngọc đặt điện thoại xuống, nhìn ra cửa sổ. Trời chiều Sài Gòn nắng gắt, ánh sáng phản chiếu trên kính tòa nhà đối diện, chói mắt.</p>

<p>Cô hiểu rồi. Đây không phải chuyện chậm tiến độ. Đây là chuyện tiền công ty đang chảy vào túi ai.</p>

<p>Cô lưu tất cả ảnh chụp vào một thư mục mã hóa trên điện thoại, xóa khỏi thư viện ảnh, và quay lại bàn ngồi nhập CRM như không có gì xảy ra.</p>

<p>Lúc mọi người trở về sau bữa trưa, Phương Linh đặt một ly trà sữa Gong Cha trước mặt cô.</p>

<p>"Mai Anh, chị mua cho em luôn. Trân châu đường đen nhé."</p>

<p>"Dạ, cảm ơn chị."</p>

<p>Bảo Ngọc mỉm cười, nhấp một ngụm trà sữa, và tiếp tục gõ phím.</p>

<p>Ngón tay cô không run. Nhưng tim cô đập nhanh hơn bình thường.</p>"""
}

ch3 = {
    "title": "Chương 3: Cuộc Họp Mà Thực Tập Sinh Không Được Dự",
    "content": """<p>Tuần thứ ba, Phạm Quốc Huy triệu tập một cuộc họp toàn phòng kinh doanh lúc hai giờ chiều thứ Hai.</p>

<p>"Tất cả nhân viên chính thức vào phòng họp lớn. Thực tập sinh ở ngoài."</p>

<p>Mai Anh gật đầu, ngồi yên tại bàn. Cô nhìn ba mươi mốt người đứng dậy đi vào phòng họp — cánh cửa kính đóng lại, rèm che được kéo xuống.</p>

<p>Nhưng Bảo Ngọc không phải người ngồi yên chờ.</p>

<p>Cô biết phòng họp lớn tầng tám có hệ thống hội nghị truyền hình — được kết nối với phần mềm Zoom của công ty. Và cô biết tài khoản Zoom công ty dùng chung một mật khẩu cho tất cả nhân viên, bao gồm thực tập sinh — vì chính cô đã được gửi thông tin đăng nhập vào ngày đầu tiên.</p>

<p>Cô mở laptop, đăng nhập Zoom, tìm cuộc họp phòng kinh doanh. Cuộc họp đang diễn ra — nhưng không ai bật camera từ phía phòng họp, chỉ có micro.</p>

<p>Cô cắm tai nghe và lắng nghe.</p>

<p>Tiếng Phạm Quốc Huy vang lên rõ ràng:</p>

<p>"Dự án Marina đã giải ngân được bảy mươi phần trăm — tương đương một trăm bốn mươi tỷ. Tiến độ thi công thực tế là bốn mươi lăm phần trăm, nhưng trên báo cáo gửi cho HĐQT, chúng ta vẫn giữ con số chín mươi lăm phần trăm. Không ai được thay đổi số liệu này mà không có sự đồng ý của tôi."</p>

<p>Cả phòng im lặng. Không ai phản đối.</p>

<p>Huy tiếp tục: "Tháng tới, Chủ tịch sẽ đến Nha Trang kiểm tra tiến độ. Trước ngày đó, tôi cần anh Sơn ở Sơn Hải hoàn thành ít nhất mặt tiền khối A để chụp ảnh. Mặt tiền thôi — không cần nội thất."</p>

<p>Phương Linh hỏi: "Nếu Chủ tịch vào trong khối A thì sao?"</p>

<p>"Không vào được. Tôi sẽ nói rằng khu vực đang trong giai đoạn nghiệm thu an toàn lao động, chưa cho người ngoài vào."</p>

<p>Bảo Ngọc tháo tai nghe ra, nhìn xuống bàn tay mình.</p>

<p>Cha cô sắp bị lừa. Một trăm bốn mươi tỷ đồng đã được giải ngân, nhưng công trình chỉ hoàn thành bốn mươi lăm phần trăm. Phần còn lại — gần sáu mươi tỷ — đang ở đâu đó giữa các hợp đồng phụ kiện, các công ty vỏ bọc, và các tài khoản mà cô chưa biết.</p>

<p>Cô cần bằng chứng cứng. Không phải ghi âm Zoom — vì ghi âm qua Zoom có thể bị phản bác về tính pháp lý. Cô cần hồ sơ giải ngân, hợp đồng phụ kiện, và sao kê tài khoản.</p>

<p>Cô nhắn tin cho một người — không phải cha cô, mà là cô Trần Minh Tâm, Giám đốc Tài chính Đại Phong, người duy nhất trong ban lãnh đạo mà cô tin tưởng.</p>

<p>"Cô Tâm, cháu là Bảo Ngọc. Cháu cần gặp cô tối nay. Ngoài công ty."</p>

<p>Tin nhắn gửi đi lúc ba giờ chiều. Trả lời lúc ba giờ mười phút: "Quán cà phê Phúc Long Nguyễn Du. 7 giờ tối."</p>

<p>Bảo Ngọc khóa điện thoại, quay lại màn hình CRM, và tiếp tục nhập dữ liệu.</p>

<p>Bên ngoài phòng họp, ba mươi mốt người vẫn đang nghe Phạm Quốc Huy chỉ đạo cách che mắt Chủ tịch.</p>

<p>Không ai biết rằng con gái Chủ tịch đang ngồi cách họ mười hai mét, và đã nghe được mọi thứ.</p>"""
}

ch4 = {
    "title": "Chương 4: Giám Đốc Tài Chính",
    "content": """<p>Trần Minh Tâm ngồi đối diện Bảo Ngọc trong quán Phúc Long trên đường Nguyễn Du, khuôn mặt nghiêm lại khi nghe xong những gì cháu gái kể.</p>

<p>Cô Tâm năm mươi hai tuổi, là em gái cha Bảo Ngọc, từng làm kiểm toán tại KPMG mười hai năm trước khi về Đại Phong giữ vị trí Giám đốc Tài chính. Cô là người duy nhất trong gia đình biết Bảo Ngọc đang giả thực tập sinh — vì chính cô đề xuất ý tưởng này.</p>

<p>"Cháu nói giải ngân bảy mươi phần trăm nhưng tiến độ chỉ bốn mươi lăm phần trăm?" cô Tâm hỏi, giọng trầm.</p>

<p>"Dạ. Và cháu phát hiện ít nhất một hợp đồng phụ kiện đáng ngờ — gói PCCC mười bốn tỷ rưỡi, ký với công ty của em gái đối tác thi công."</p>

<p>Cô Tâm nhấp cà phê, mắt nhìn xuống mặt bàn.</p>

<p>"Cháu ơi, cô biết dự án Marina có vấn đề từ ba tháng trước. Số liệu giải ngân không khớp với tiến độ — nhưng mỗi lần cô yêu cầu phòng kinh doanh giải trình, Huy luôn đưa ra báo cáo tiến độ rất đẹp. Cô không có cơ sở để nghi ngờ vì báo cáo do bên thi công Sơn Hải cung cấp."</p>

<p>"Bây giờ cô có cơ sở rồi."</p>

<p>"Đúng. Nhưng cháu cần hiểu: nếu cháu tố cáo Huy mà không đủ bằng chứng, hắn sẽ quay ngược lại — nói rằng cháu lợi dụng thân phận con gái Chủ tịch để trả thù cá nhân. Hắn có mối quan hệ rất tốt với ba cháu."</p>

<p>"Cháu biết. Nên cháu không muốn nói với ba cho đến khi có hồ sơ hoàn chỉnh."</p>

<p>Cô Tâm gật đầu.</p>

<p>"Được. Cô sẽ giúp cháu. Phòng tài chính có quyền truy cập toàn bộ hồ sơ giải ngân và hợp đồng. Cô sẽ chuẩn bị một bản đối chiếu chi tiết: giải ngân thực tế so với tiến độ thi công. Cháu cung cấp ảnh công trường từ Zalo nội bộ — đó là bằng chứng tiến độ thực."</p>

<p>"Còn hợp đồng phụ kiện thì sao?"</p>

<p>"Cô sẽ kiểm tra tất cả hợp đồng phụ kiện của dự án Marina — bao nhiêu công ty, thành lập khi nào, ai làm giám đốc, có liên quan đến Sơn Hải không."</p>

<p>Bảo Ngọc nhìn cô Tâm. Trong ánh đèn vàng ấm của quán cà phê, gương mặt cô mình trông mệt mỏi nhưng kiên quyết — gương mặt của người đã nhìn thấy quá nhiều con số giả trong đời.</p>

<p>"Cô ơi, cháu cần bao lâu?"</p>

<p>"Một tuần. Nhưng cháu phải cẩn thận — đừng để Huy nghi. Tiếp tục làm thực tập sinh bình thường, photocopy, pha trà, nhập CRM. Đừng hỏi bất cứ điều gì về dự án Marina."</p>

<p>"Dạ."</p>

<p>Cô Tâm nhìn cháu gái, ánh mắt thoáng chút xót xa.</p>

<p>"Bảo Ngọc, ba cháu sẽ rất đau khi biết chuyện này. Huy là người ông ta tin tưởng nhất — hơn cả cô."</p>

<p>"Cháu biết. Nhưng cháu không thể để người ta cướp tiền của ba trước mặt cháu mà cháu không làm gì."</p>

<p>Cô Tâm gật đầu, không nói thêm.</p>

<p>Họ ngồi im một lúc, mỗi người với suy nghĩ riêng, trong khi quán cà phê xung quanh vẫn ồn ào với tiếng cười nói của những người không biết rằng hai người phụ nữ ở góc quán đang chuẩn bị lật tẩy một vụ gian lận sáu mươi tỷ đồng.</p>"""
}

ch5 = {
    "title": "Chương 5: Bị Đuổi Giữa Phòng Họp",
    "content": """<p>Tuần thứ tư. Thứ Năm.</p>

<p>Phạm Quốc Huy triệu tập cuộc họp khẩn vào lúc mười giờ sáng. Lần này, anh ta không nói "thực tập sinh ở ngoài" — vì Mai Anh đang photocopy tài liệu cho phòng nhân sự ở tầng khác.</p>

<p>Nhưng cô quay lại sớm hơn dự kiến.</p>

<p>Cô đẩy cửa phòng kinh doanh, thấy mọi người đang đứng quanh bàn họp giữa phòng. Trên bàn là một tập hồ sơ dày — Huy đang giơ lên và chỉ vào từng trang.</p>

<p>"Gói thầu số 09 — Hệ thống điện thông minh — hai mươi hai tỷ. Ký với Công ty Điện lực Tương Lai. Giải ngân đợt một: mười một tỷ. Đợt hai thanh toán trước cuối tháng."</p>

<p>Mai Anh ngồi về bàn mình, lặng lẽ. Nhưng cô đã nghe được tên: Công ty Điện lực Tương Lai. Cô ghi nhanh vào Notes trên điện thoại.</p>

<p>Cuộc họp kéo dài hai mươi phút. Khi kết thúc, Huy đi ngang bàn Mai Anh và dừng lại.</p>

<p>"Mai Anh, em có nghe gì trong cuộc họp vừa rồi không?"</p>

<p>"Dạ không ạ. Em mới vào, đang sắp xếp tài liệu."</p>

<p>Huy nhìn cô một lúc, ánh mắt lạnh hơn nụ cười.</p>

<p>"Tốt. Thực tập sinh không cần quan tâm đến chuyện dự án. Em cứ làm tốt phần việc của mình."</p>

<p>"Dạ."</p>

<hr>

<p>Chiều hôm đó, cô Tâm gửi cho Bảo Ngọc một file Excel qua Signal — ứng dụng nhắn tin mã hóa đầu cuối.</p>

<p>File có bốn sheet:</p>

<p>Sheet 1: Danh sách tất cả hợp đồng phụ kiện dự án Marina — mười bảy hợp đồng, tổng giá trị sáu mươi ba tỷ đồng.</p>

<p>Sheet 2: Thông tin đăng ký kinh doanh của mười bảy nhà thầu phụ — bốn trong số đó có giám đốc là người thân của Vũ Hoàng Sơn (em gái, anh vợ, cháu ruột, và một người bạn thân từ thời đại học).</p>

<p>Sheet 3: Đối chiếu giải ngân và tiến độ — một trăm bốn mươi tỷ đã giải ngân, tiến độ thực tế bốn mươi lăm phần trăm, chênh lệch ước tính: năm mươi tám tỷ đồng.</p>

<p>Sheet 4: Ảnh chụp công trường từ Zalo nội bộ (Bảo Ngọc cung cấp) đặt cạnh ảnh trong báo cáo tiến độ — sự khác biệt rõ ràng đến mức không cần chuyên gia xây dựng cũng thấy.</p>

<p>Bảo Ngọc đọc xong, đóng điện thoại, nằm xuống giường trong căn hộ thuê ở quận 2.</p>

<p>Năm mươi tám tỷ đồng. Chảy qua bốn công ty vỏ bọc, về túi Vũ Hoàng Sơn và những người liên quan. Phạm Quốc Huy là người duyệt, Phương Linh là người thực hiện thanh toán, và cả phòng kinh doanh im lặng vì sợ mất việc.</p>

<p>Ngày mai, Bảo Ngọc sẽ gặp cha.</p>

<hr>

<p>Nhưng ngày mai đến sớm hơn cô nghĩ.</p>

<p>Sáng thứ Sáu, khi Mai Anh bước vào phòng kinh doanh, Phạm Quốc Huy đã đứng ở giữa phòng, tay cầm một chiếc điện thoại.</p>

<p>"Nguyễn Mai Anh."</p>

<p>Cả phòng quay lại nhìn.</p>

<p>"Em giải thích cho tôi tại sao em chụp ảnh hợp đồng phụ kiện gói PCCC trên máy in tuần trước?"</p>

<p>Bảo Ngọc cảm thấy máu dồn lên mặt. Camera an ninh. Cô đã quên camera ở góc phòng nhìn thẳng vào khu vực máy in.</p>

<p>"Em chụp để tham khảo mẫu hợp đồng, phục vụ việc học," cô nói, giọng bình tĩnh nhất có thể.</p>

<p>"Học? Thực tập sinh chụp ảnh hợp đồng trị giá mười bốn tỷ để 'học'?" Huy cười nhạt. "Em nghĩ tôi tin à?"</p>

<p>Anh ta bước đến gần hơn, giọng hạ thấp nhưng đủ để cả phòng nghe:</p>

<p>"Nguyễn Mai Anh, kể từ bây giờ em bị chấm dứt thực tập. Nộp thẻ nhân viên, thu dọn đồ, và ra khỏi văn phòng trước mười hai giờ trưa."</p>

<p>"Anh không có quyền đuổi thực tập sinh," Bảo Ngọc nói. "Quyết định đó thuộc phòng nhân sự."</p>

<p>"Tôi là trưởng phòng kinh doanh. Em thực tập ở phòng tôi. Tôi có toàn quyền quyết định ai ở lại, ai ra đi."</p>

<p>Huy nhìn cô từ trên xuống, ánh mắt khinh bỉ.</p>

<p>"Thực tập sinh thì biết gì về kinh doanh? Biết gì về hợp đồng? Em chỉ là đứa ngồi cạnh máy photocopy, đừng cố đóng vai thám tử."</p>

<p>Ba mươi mốt người trong phòng im lặng. Không ai đứng lên bênh vực cô.</p>

<p>Bảo Ngọc nhìn quanh — những gương mặt cúi xuống, những ánh mắt lảng tránh. Cô hiểu: họ sợ Huy hơn sợ sự thật.</p>

<p>Cô thu dọn đồ vào túi xách, nộp thẻ nhân viên lên bàn Huy, và bước ra.</p>

<p>Khi cánh cửa thang máy đóng lại, cô lấy điện thoại ra và gọi một cuộc gọi.</p>

<p>Không phải cho cô Tâm. Không phải cho luật sư.</p>

<p>Mà cho cha cô.</p>

<p>"Ba ơi, con cần gặp ba. Ngay bây giờ."</p>"""
}

ch6 = {
    "title": "Chương 6: Cha Và Con Gái",
    "content": """<p>Trần Đình Khoa ngồi trong phòng làm việc ở tầng hai mươi mốt — tầng cao nhất của trụ sở Đại Phong — nhìn con gái đặt tập hồ sơ trước mặt.</p>

<p>Ông sáu mươi mốt tuổi, tóc bạc, gương mặt sạm nắng của người từng chạy xe máy đi bán xi măng ở Bình Dương ba mươi năm trước. Bàn tay ông thô ráp, móng tay vẫn ngắn cũn — thói quen từ thời làm thợ xây, dù bây giờ ông ngồi trong phòng gắn máy lạnh trị giá hai mươi nghìn đô.</p>

<p>"Con nói Huy đang cướp tiền của ba?"</p>

<p>"Không phải con nói, ba. Số liệu nói." Bảo Ngọc mở file trên laptop. "Dự án Marina giải ngân một trăm bốn mươi tỷ, tiến độ thực tế bốn mươi lăm phần trăm. Ít nhất năm mươi tám tỷ đang ở đâu đó giữa bốn công ty vỏ bọc do người nhà đối tác thi công làm giám đốc."</p>

<p>Trần Đình Khoa nhìn bảng số liệu. Ông không phải dân tài chính — ông là dân xây dựng, đọc bản vẽ giỏi hơn đọc bảng cân đối. Nhưng ông hiểu một điều rất đơn giản: khi tiền ra nhiều hơn công trình có, ai đó đang lấy cắp.</p>

<p>Ông im lặng rất lâu.</p>

<p>"Huy... ba tin thằng Huy. Ba cho nó quyền ký duyệt thanh toán dưới hai mươi tỷ. Ba nghĩ nó trung thành."</p>

<p>"Ba ơi, người trung thành không cần bốn công ty vỏ bọc."</p>

<p>Ông nhìn con gái. Đôi mắt ông ướt — không phải vì buồn, mà vì tức. Kiểu tức của một người xây dựng cả đời mà bị chính người mình tin tưởng móc túi.</p>

<p>"Con muốn ba làm gì?"</p>

<p>"Con muốn ba không làm gì. Chưa."</p>

<p>Ông ngạc nhiên.</p>

<p>"Con muốn ba cho con thêm một tuần. Con sẽ phối hợp với cô Tâm hoàn thiện hồ sơ, thuê luật sư độc lập rà soát, và chuẩn bị phương án xử lý — bao gồm khởi kiện dân sự và chuyển hồ sơ sang cơ quan điều tra nếu cần."</p>

<p>"Tại sao không xử lý ngay?"</p>

<p>"Vì nếu ba gọi Huy lên ngay bây giờ, hắn sẽ xóa bằng chứng. Hắn sẽ liên lạc Sơn Hải để thống nhất lời khai. Và hắn sẽ nói rằng con gái Chủ tịch đang trả thù cá nhân vì bị đuổi việc."</p>

<p>Trần Đình Khoa nhìn con gái. Cô bé mà ông từng bế trên tay, đưa đi công trường, cho đội mũ bảo hộ quá rộng — bây giờ đang dạy ông cách đánh một ván cờ.</p>

<p>"Con lớn rồi, Bảo Ngọc."</p>

<p>"Dạ. Nhưng con vẫn cần ba."</p>

<p>Ông gật đầu.</p>

<p>"Một tuần. Ba sẽ không nói với ai."</p>

<p>Bảo Ngọc đứng dậy, ôm cha, rồi bước ra.</p>

<p>Ngoài cửa phòng Chủ tịch, cô gặp thư ký riêng của cha — chị Nguyễn Thu Hương, người biết mặt cô từ nhỏ.</p>

<p>"Bảo Ngọc? Sao con lại ở đây? Ba con nói con đang thực tập?"</p>

<p>"Dạ, con vừa kết thúc thực tập, chị Hương." Bảo Ngọc mỉm cười. "Sớm hơn dự kiến."</p>

<p>Chị Hương nhìn cô, không hiểu, nhưng cũng không hỏi thêm.</p>

<p>Bảo Ngọc bước vào thang máy, nhấn nút tầng trệt. Trong gương thang máy, cô nhìn thấy mình — không còn là Mai Anh, thực tập sinh quèn. Nhưng cũng chưa phải Trần Bảo Ngọc, con gái Chủ tịch.</p>

<p>Cô đang ở giữa — và đó là nơi nguy hiểm nhất.</p>"""
}

ch7 = {
    "title": "Chương 7: Phiên Họp Hội Đồng Quản Trị",
    "content": """<p>Một tuần sau.</p>

<p>Hội đồng Quản trị Tập đoàn Đại Phong họp phiên bất thường vào sáng thứ Hai, tại phòng họp lớn tầng hai mươi mốt. Thành phần: năm thành viên HĐQT, Giám đốc Tài chính Trần Minh Tâm, luật sư tư vấn độc lập Lê Hoàng Long, và — bất ngờ — Trần Bảo Ngọc.</p>

<p>Khi Bảo Ngọc bước vào phòng họp, một thành viên HĐQT hỏi:</p>

<p>"Cô này là ai?"</p>

<p>Trần Đình Khoa đáp: "Đây là con gái tôi, Trần Bảo Ngọc. Cô ấy sẽ trình bày báo cáo đặc biệt về dự án Đại Phong Marina."</p>

<p>Phạm Quốc Huy — cũng được mời dự — ngồi ở đầu bàn bên kia, mặt biến sắc khi nhận ra "thực tập sinh Mai Anh" đang đứng trước mặt mình trong bộ vest đen, tóc búi cao, bảng tên ghi: "Trần Bảo Ngọc — Đại diện HĐQT."</p>

<p>"Cô... cô là..." Huy lắp bắp.</p>

<p>"Chào anh Huy," Bảo Ngọc nói, giọng bình tĩnh. "Tuần trước anh đuổi em ra khỏi phòng kinh doanh. Hôm nay em quay lại."</p>

<p>Cả phòng họp im lặng.</p>

<p>Bảo Ngọc mở laptop, kết nối với máy chiếu, và bắt đầu trình bày.</p>

<p>Slide 1: Tổng quan dự án Đại Phong Marina — tổng mức đầu tư hai trăm tỷ, giải ngân một trăm bốn mươi tỷ, tiến độ thi công thực tế bốn mươi lăm phần trăm.</p>

<p>Slide 2: Bảng đối chiếu giải ngân — chi tiết từng đợt thanh toán, so sánh với ảnh công trường thực tế. Hình ảnh phân chia hai cột: bên trái là ảnh trong báo cáo tiến độ (công trình gần hoàn thiện), bên phải là ảnh thực tế từ Zalo nội bộ (bãi đất trống, vài cọc bê tông).</p>

<p>Slide 3: Mười bảy hợp đồng phụ kiện — bốn công ty có giám đốc là người thân Vũ Hoàng Sơn. Tổng giá trị: sáu mươi ba tỷ. Trong đó ít nhất bốn công ty thành lập dưới một năm, vốn điều lệ dưới một tỷ.</p>

<p>Slide 4: Sơ đồ dòng tiền — tiền từ Đại Phong → Sơn Hải → bốn công ty vỏ bọc → quay ngược lại tài khoản cá nhân (dựa trên phân tích của cô Tâm và luật sư).</p>

<p>Phạm Quốc Huy ngồi cứng đờ. Mặt anh ta trắng bệch, tay nắm chặt dưới bàn.</p>

<p>"Đây là vu khống," anh ta nói, giọng run. "Cô này mới hai mươi bốn tuổi, mới ra trường, cô ấy biết gì—"</p>

<p>"Anh Huy," Bảo Ngọc cắt lời, giọng không lớn nhưng sắc, "em không cần biết nhiều. Em chỉ cần biết đếm. Và con số không biết nói dối."</p>

<p>Luật sư Lê Hoàng Long lên tiếng:</p>

<p>"Thưa HĐQT, dựa trên hồ sơ mà Giám đốc Tài chính cung cấp, có cơ sở để cho rằng đã xảy ra hành vi chiếm đoạt tài sản thông qua hợp đồng khống. Tôi đề xuất: một, đình chỉ chức vụ trưởng phòng kinh doanh Phạm Quốc Huy; hai, yêu cầu kiểm toán độc lập; ba, chuyển hồ sơ sang Cơ quan Cảnh sát Điều tra nếu kết quả kiểm toán xác nhận sai phạm."</p>

<p>HĐQT biểu quyết. Năm phiếu thuận, không phiếu chống.</p>

<p>Phạm Quốc Huy đứng dậy, ghế đẩy ra sau kêu rít trên sàn gỗ.</p>

<p>"Các anh sẽ hối hận. Tôi đã cống hiến cho công ty này tám năm—"</p>

<p>"Cống hiến bao nhiêu thì cướp bấy nhiêu," Trần Đình Khoa nói, giọng trầm, nặng như xi măng. "Ra đi, Huy."</p>

<p>Huy nhìn quanh phòng, không ai nhìn lại anh ta. Anh ta cầm điện thoại, bước ra, và cánh cửa đóng lại.</p>

<p>Bảo Ngọc đứng yên, nhìn cánh cửa đóng lại, rồi nhìn cha. Ông Khoa gật đầu nhẹ — không mỉm cười, nhưng mắt ông nói: "Ba tự hào."</p>"""
}

ch8 = {
    "title": "Chương 8: Cô Gái Không Còn Ngồi Cạnh Máy Photocopy",
    "content": """<p>Ba tháng sau.</p>

<p>Kiểm toán độc lập do KPMG thực hiện xác nhận: dự án Đại Phong Marina bị thất thoát năm mươi tám tỷ bốn trăm triệu đồng thông qua mười một hợp đồng phụ kiện khống. Tiền được chuyển qua bốn công ty vỏ bọc, cuối cùng về tài khoản cá nhân của Vũ Hoàng Sơn và hai người thân.</p>

<p>Phạm Quốc Huy bị khởi tố về tội "Lạm dụng chức vụ, quyền hạn chiếm đoạt tài sản" theo Điều 355 Bộ luật Hình sự. Anh ta nhận vai trò người duyệt thanh toán, thông đồng với Sơn Hải để rút ruột dự án.</p>

<p>Hoàng Thị Phương Linh — phó phòng kinh doanh — hợp tác điều tra và khai nhận đã thực hiện thanh toán theo chỉ đạo của Huy, dù biết các hợp đồng có vấn đề. Cô ta bị kỷ luật sa thải nhưng không bị khởi tố do tự nguyện cung cấp bằng chứng.</p>

<p>Vũ Hoàng Sơn và em gái Vũ Thị Hồng Nhung bị bắt tạm giam.</p>

<hr>

<p>Trần Bảo Ngọc được HĐQT bổ nhiệm vào vị trí mới: Giám đốc Kiểm soát Nội bộ Tập đoàn Đại Phong — phòng ban mới thành lập, báo cáo trực tiếp cho HĐQT, không qua bất kỳ giám đốc điều hành nào.</p>

<p>Phòng làm việc của cô ở tầng mười chín — không phải tầng hai mươi mốt của Chủ tịch, và không phải tầng tám của phòng kinh doanh cũ. Đó là quyết định của cô: cô muốn ở giữa, nơi cô có thể nhìn thấy cả trên lẫn dưới.</p>

<p>Phòng có cửa sổ lớn nhìn ra sông Sài Gòn, bàn gỗ sáng, ghế lưng cao, và một cây xanh nhỏ trên góc bàn. Không có máy photocopy trong phòng.</p>

<hr>

<p>Một buổi sáng, Bảo Ngọc đi thang máy xuống tầng tám — tầng phòng kinh doanh cũ — để lấy một tập hồ sơ.</p>

<p>Khi cô bước ra, phòng kinh doanh đang hoạt động bình thường. Trưởng phòng mới — một người do cô Tâm giới thiệu, từng làm ở PwC — đang họp nhóm. Nhân viên đi lại, gõ phím, nói chuyện điện thoại.</p>

<p>Cô đi ngang qua khu vực máy photocopy — cái góc quen thuộc, cái bàn nhỏ nhất phòng, nơi cô từng ngồi suốt bốn tuần.</p>

<p>Bây giờ, ở cái bàn đó, ngồi một cô gái trẻ — khoảng hai mươi hai tuổi, tóc dài, áo sơ mi trắng, đang nhập dữ liệu CRM. Thẻ nhân viên ghi: "Thực tập sinh."</p>

<p>Cô gái ngẩng lên, thấy Bảo Ngọc, rồi cúi xuống tiếp tục gõ.</p>

<p>Bảo Ngọc dừng lại một giây. Cô nhìn cái bàn, cái máy photocopy, cái ghế — và nhớ lại bốn tuần mình đã ngồi ở đây, bị gọi là "đứa photocopy", bị đuổi giữa phòng họp.</p>

<p>Cô bước đến gần.</p>

<p>"Em tên gì?"</p>

<p>Cô gái thực tập ngẩng lên, hơi ngạc nhiên.</p>

<p>"Dạ, em tên Phương. Lê Ngọc Phương."</p>

<p>"Em thực tập ở đây lâu chưa?"</p>

<p>"Dạ, mới tuần đầu."</p>

<p>Bảo Ngọc nhìn cô gái — nhìn thấy chính mình của bốn tháng trước: trẻ, háo hức, và hoàn toàn vô hình trong mắt mọi người.</p>

<p>"Phương, chị có một lời khuyên. Em đang ngồi ở vị trí này, mọi người sẽ gọi em là 'đứa photocopy'. Sẽ không ai hỏi em tốt nghiệp trường nào, giỏi gì, muốn gì. Nhưng em nhớ một điều: người ngồi cạnh máy photocopy là người nhìn thấy mọi tài liệu đi qua phòng. Đôi khi, đó là vị trí tốt nhất để hiểu một công ty."</p>

<p>Phương nhìn Bảo Ngọc, mắt tròn xoe.</p>

<p>"Dạ... chị là ai ạ?"</p>

<p>Bảo Ngọc mỉm cười — nụ cười mà cô đã giấu suốt bốn tuần khi còn là Mai Anh.</p>

<p>"Chị là người từng ngồi ở chỗ em."</p>

<p>Cô quay người, bước vào thang máy, nhấn nút tầng mười chín.</p>

<p>Cánh cửa thang máy đóng lại, phản chiếu gương mặt cô trên lớp thép bóng — gương mặt của Trần Bảo Ngọc, không phải Nguyễn Mai Anh.</p>

<p>Phía dưới tầng tám, máy photocopy vẫn chạy rè rè. Nhưng lần này, nó không còn là âm thanh của sự vô hình.</p>

<p>Mà là âm thanh của một câu chuyện đang bắt đầu.</p>"""
}

story_data["chapters"] = [ch1, ch2, ch3, ch4, ch5, ch6, ch7, ch8]

# Save & upload
with open("scratch/story_1948_rewrite.json", "w", encoding="utf-8") as f:
    json.dump(story_data, f, ensure_ascii=False, indent=2)

print("✅ Truyện 1948 viết xong!")
for i, ch in enumerate(story_data['chapters']):
    print(f"  Ch{i+1}: {ch['title']} — {len(ch['content'])} chars")

# Upload
print("\n📤 Uploading to WordPress...")
url = "https://doctieuthuyet.com/overwrite_story_v13.php"
payload = json.dumps(story_data).encode('utf-8')
req = urllib.request.Request(url, data=payload, headers={'Content-Type': 'application/json'})
try:
    resp = urllib.request.urlopen(req, timeout=120)
    result = json.loads(resp.read().decode('utf-8'))
    print(f"  ✅ Success! Story ID={result['story_id']}, Chapters={result['chapters_count']}")
except Exception as e:
    print(f"  ❌ Error: {e}")

# Update Excel
print("\n📊 Updating Excel...")
import openpyxl
wb = openpyxl.load_workbook("danh_sach_truyen_doctieuthuyet.xlsx")
ws = wb.active
row = 191  # STT 187 = row 191 (header at row 4, data starts row 5, so STT 187 = row 4+187=191)
# Find exact row
for r in range(5, ws.max_row+1):
    if ws.cell(row=r, column=2).value and str(ws.cell(row=r, column=2).value).strip() == "1948":
        row = r
        break

ws.cell(row=row, column=3).value = story_data["title"]
ws.cell(row=row, column=4).value = story_data["author"]
ws.cell(row=row, column=5).value = "Sảng Văn"
ws.cell(row=row, column=12).value = "Con gái Chủ tịch tập đoàn bất động sản giả thực tập sinh, phát hiện trưởng phòng kinh doanh cùng đối tác rút ruột dự án 200 tỷ. Bị đuổi giữa phòng họp vì chụp ảnh hợp đồng, cô lật bài trước HĐQT."
ws.cell(row=row, column=13).value = "REWRITE TOÀN BỘ: Tên TQ, AI template text, meta-breaking → ĐÃ HOÀN THÀNH"
ws.cell(row=row, column=14).value = "☑️ Đã sửa"
wb.save("danh_sach_truyen_doctieuthuyet.xlsx")
print(f"  ✅ Excel updated for story 1948 (row {row})")
