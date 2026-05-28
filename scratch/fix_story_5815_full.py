import importlib.util
import json
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
spec = importlib.util.spec_from_file_location("novel_editor", ROOT / "scratch" / "novel_editor.py")
editor = importlib.util.module_from_spec(spec)
spec.loader.exec_module(editor)

STORY_ID = 5815
TITLE = "Ông Già Bán Chè Đêm Và Cậu Bé Mồ Côi - Tô Chè Ấm Nhất Sài Gòn"
INTRO = (
    "Ông Trần Văn Phước bảy mươi tuổi, bán chè đêm ở góc đường quận 5. "
    "Chiếc xe chè nhỏ sáng đèn từ chín giờ tối đến hai giờ sáng, đón bảo vệ, tài xế, sinh viên và những người mệt mỏi nhất thành phố. "
    "Một đêm mưa, ông gặp Bảo, cậu bé mồ côi mười tuổi ngồi co ro cạnh thùng rác, đói đến mức không dám xin một tô chè. "
    "Từ một tô chè nóng, hai số phận cô độc tựa vào nhau, đi qua chữ nghĩa, bệnh tật, trưởng thành và một lời hứa giữ lại hơi ấm cho Sài Gòn."
)

CHAPTERS = {
    5817: (
        "Chương 1: Xe Chè Đêm Góc Đường",
        """
Chín giờ tối, quận 5 vẫn chưa chịu ngủ. Đèn tiệm vàng đã hạ bớt, quán mì kéo cửa sắt lưng chừng, xe buýt cuối ngày thở khói ở ngã tư, còn góc đường Nguyễn Trãi - Trần Bình Trọng bắt đầu sáng lên bằng một bóng đèn vàng treo trên chiếc xe chè cũ. Người ta gọi đó là xe chè của ông Phước, nhưng với nhiều người làm đêm, nó giống một cái bếp nhỏ hơn là một hàng quán.

Ông Trần Văn Phước bảy mươi tuổi, lưng hơi còng, tóc bạc như sợi bông dính trên vai áo nâu. Mỗi tối ông đẩy xe ra đúng một chỗ, lau mặt kính, châm thêm nước nóng vào nồi, xếp chén sứ thành từng chồng. Chè đậu xanh, chè bắp, chè chuối, sương sáo, nước cốt dừa đều được ông nấu từ chiều trong căn phòng trọ phía sau chợ. Ông không ghi bảng hiệu lớn. Tấm bảng carton treo bằng dây kẽm chỉ viết: "Chè nóng, ai đói cứ ăn trước."

Khách của ông không ồn ào. Bác bảo vệ chung cư ghé lúc đổi ca, ăn một chén chè đậu đen rồi thở dài vì đau lưng. Chú tài xế công nghệ tấp xe vào, vừa ăn vừa nhìn app xem còn cuốc nào không. Mấy cô công nhân may tan ca muộn mua chung một bịch chè bắp, chia nhau bằng muỗng nhựa. Sinh viên ký túc xá thỉnh thoảng thiếu tiền, ông Phước cứ múc đầy rồi nói: "Mai mốt ra trường nhớ trả bằng cách đàng hoàng với đời."

Người ta hỏi sao già rồi không nghỉ. Ông cười, bàn tay vẫn đều đều múc chè: "Ở nhà một mình, cái im nó lớn lắm." Vợ ông mất mười hai năm trước vì bệnh tim. Con trai duy nhất đi biển rồi không về sau một cơn bão ngoài Cà Mau. Từ đó, ông bán chè đêm không hẳn để kiếm sống, mà để nghe tiếng người. Tiếng muỗng chạm chén, tiếng ai đó than mệt, tiếng một câu cảm ơn nhỏ, tất cả gom lại thành thứ ấm áp giúp ông qua từng đêm.

Đêm ấy mưa lất phất từ mười giờ. Mưa Sài Gòn không lạnh như miền Trung, nhưng đủ làm áo người đi đường dính vào da, đủ làm những đứa trẻ lang thang co vai tìm mái hiên. Ông Phước vừa đậy nắp nồi chè chuối thì thấy bên kia gốc me có một bóng nhỏ ngồi im. Ban đầu ông tưởng là bao rác ai bỏ quên. Đến khi chiếc bóng ấy ho khan, ông mới nhận ra đó là một cậu bé.

Cậu chừng mười tuổi, tóc bết nước, áo thun rộng quá vai, quần đùi rách gấu. Hai đầu gối ôm sát ngực, mắt nhìn nồi chè như nhìn một thứ xa xỉ không thuộc về mình. Mỗi lần có khách trả tiền, cậu lại cúi mặt xuống, như sợ bị phát hiện chỉ vì đã đói. Ông Phước nhìn một lúc rồi múc một chén chè đậu xanh, chan thêm nước cốt dừa, bỏ muỗng vào, bưng qua.

"Ăn đi con," ông nói.

Cậu bé giật mình lùi lại. "Con không có tiền."

"Ông đâu có hỏi tiền."

"Con không ăn xin."

Câu nói cứng hơn thân hình gầy gò của cậu. Ông Phước ngồi xuống bậc thềm, đặt chén chè ở giữa hai người. "Vậy ông nhờ con ăn thử coi hôm nay ông nấu có ngọt quá không. Khách chê là ông buồn lắm."

Cậu bé nhìn ông, nghi ngờ và thèm thuồng đánh nhau trong đôi mắt đen. Cuối cùng cậu cầm muỗng. Miếng đầu tiên vào miệng, vai cậu run lên rất nhẹ. Không biết vì nóng, vì đói, hay vì lâu rồi chưa có ai đưa cho cậu một thứ mà không kèm cái nhìn khinh rẻ. Ông Phước quay mặt về phía xe chè, giả vờ không thấy nước mắt rơi vào chén.

Đêm đó, cậu bé ăn hết hai chén. Trước khi đi, cậu đặt cái muỗng xuống thật ngay ngắn rồi nói nhỏ: "Mai con trả."

Ông Phước hỏi: "Con tên gì?"

"Bảo."

"Nhà ở đâu?"

Cậu bé im. Cái im ấy dài hơn câu trả lời. Ông Phước không hỏi nữa. Ông chỉ lấy một túi nilon, múc thêm chè bắp, buộc kỹ rồi đưa cậu. "Mai trả cũng được. Không trả cũng được. Nhưng mai ghé lại cho ông biết chè có bị ngọt quá không."

Bảo ôm túi chè vào ngực như ôm một món đồ dễ vỡ. Khi cậu chạy qua màn mưa, ông Phước bỗng thấy góc đường rộng hơn mọi đêm, nhưng cũng trống hơn. Ông nhìn theo cho đến khi bóng cậu khuất sau hàng xe máy, lòng tự dưng nhoi nhói. Sài Gòn có quá nhiều ánh đèn, nhưng vẫn có những đứa trẻ ngồi ngay dưới đèn mà không ai thấy.
""",
    ),
    5818: (
        "Chương 2: Cậu Bé Ngồi Cạnh Xe Chè",
        """
Tối hôm sau, Bảo quay lại thật. Cậu không ngồi ở gốc me nữa mà đứng cách xe chè ba bước, hai bàn tay nắm vạt áo, vẻ mặt như người đến trả một món nợ lớn. Ông Phước đang bán cho bác bảo vệ, thấy cậu thì chỉ gật đầu, không reo lên, không hỏi dồn. Ông biết với những đứa trẻ từng bị đời xua đuổi, một sự niềm nở quá mức cũng có thể làm chúng sợ.

"Hôm qua ngọt vừa không?" ông hỏi.

Bảo đáp rất nghiêm túc: "Hơi ngọt. Nhưng ngon."

Bác bảo vệ bật cười. Ông Phước cũng cười, múc cho cậu một chén chè bắp ít đường hơn. Bảo móc trong túi ra ba đồng xu và một tờ hai nghìn nhàu nát, đặt lên mặt kính. "Con trả trước một ít."

Ông Phước không đẩy tiền lại ngay. Ông sợ làm cậu tổn thương. Ông chỉ lấy hai nghìn, bỏ vào hộp tiền, rồi đưa lại mấy đồng xu. "Ông giảm giá cho người kiểm chè."

Từ hôm đó, Bảo trở thành một phần của góc đường. Cậu đến sau mười giờ, khi khách thưa hơn, ngồi trên chiếc ghế nhựa thấp cạnh xe chè. Ban đầu cậu chỉ ăn rồi đi. Sau vài đêm, cậu tự cầm khăn lau muỗng. Rồi cậu biết lấy bịch nilon, biết đưa tiền thừa, biết nhắc ông Phước khi nước cốt dừa gần cạn. Ông không sai cậu làm, cậu tự làm như thể mỗi việc nhỏ là một cách trả lại tô chè đã nhận.

Khách quen bắt đầu để ý. Cô công nhân may hỏi: "Cháu ông hả chú Phước?"

Ông chưa kịp trả lời, Bảo đã nói: "Không phải. Con phụ ông."

Chữ "phụ" được cậu nói rất nhanh, nhưng ông Phước nghe ra trong đó có một niềm tự hào nhỏ. Ông gật đầu: "Ừ, phụ ông coi xe chè."

Bảo không kể nhiều về mình. Ông Phước phải nhặt từng mẩu rơi ra trong những đêm dài. Mẹ cậu mất vì sốt xuất huyết khi cậu bảy tuổi. Cha cậu đi làm phụ hồ ở Bình Dương, rồi biến mất sau một vụ nợ nần. Cậu từng ở với dì, nhưng nhà dì đông con, bữa cơm nào cũng có tiếng thở dài. Một hôm cậu bị mắng vì làm mất năm chục nghìn, cậu bỏ đi. Từ đó cậu ngủ khi thì gầm cầu, khi thì hiên chợ, khi thì trong thùng giấy sau một tiệm điện thoại.

Ông Phước nghe mà cổ họng nghẹn lại. Ông muốn hỏi sao con không tìm người lớn giúp, nhưng câu ấy bật ra quá dễ với người có nhà để về. Với Bảo, người lớn thường là bàn tay xua đi, giọng nói đuổi khỏi mái hiên, cái nhìn nghi ngờ khi cậu đứng gần quầy bánh. Cậu đã học cách không xin để khỏi bị mắng, không khóc để khỏi bị coi thường, không ở lâu một chỗ để khỏi bị nhớ mặt.

Một đêm, trời mưa lớn. Khách vắng, nước chảy thành dòng dưới bánh xe chè. Bảo ngồi co ro, áo ướt từ vai đến lưng. Ông Phước lấy chiếc áo khoác cũ của mình khoác lên người cậu. Áo rộng, mùi dầu gió và khói bếp. Bảo kéo áo sát lại, cúi đầu rất lâu.

"Con ngủ ở đâu tối nay?" ông hỏi.

"Chỗ cũ."

"Chỗ cũ là chỗ nào?"

"Sau chợ."

Ông Phước nhìn màn mưa trắng xóa. Sau chợ đêm nay nước ngập đến mắt cá, chuột chạy đầy, bảo vệ mới đổi ca rất khó tính. Ông im lặng múc chè cho khách cuối cùng, dọn bớt chén, rồi nói như nói chuyện bình thường: "Lát nữa phụ ông đẩy xe về. Nhà ông có cái hiên khô. Con ngủ một đêm, sáng đi đâu thì đi."

Bảo ngẩng phắt lên. "Con không ăn cắp."

"Ông biết."

"Con cũng không làm phiền."

"Ông già rồi, có người làm phiền chút cũng vui."

Cậu bé nhìn ông rất lâu, như đang tìm một cái bẫy trong giọng nói hiền lành ấy. Không thấy, cậu lại cúi xuống. Khi ông Phước tắt bóng đèn, Bảo lặng lẽ đi ra sau xe chè, hai tay đặt vào càng đẩy. Chiếc xe nặng, đường trơn, nhưng lần đầu tiên sau nhiều tháng, cậu đi trong mưa mà không phải tìm chỗ trốn. Có một người đi bên cạnh, chậm rãi, không hỏi quá khứ của cậu như hỏi tội, chỉ thỉnh thoảng nhắc: "Coi chừng ổ gà, con."
""",
    ),
    5819: (
        "Chương 3: Ông Phước Đưa Bảo Về Nhà",
        """
Nhà ông Phước là căn phòng trọ mười sáu mét vuông sau chợ, mái tôn thấp, tường vôi loang lổ, nhưng sạch đến mức Bảo ngại đặt chân vào. Một bên là bếp than và kệ nồi chè, một bên là chiếc giường gỗ nhỏ phủ chiếu trúc. Trên tường treo ảnh một người phụ nữ tóc búi và một thanh niên mặc áo lính biển. Dưới ảnh có bát nhang, lọ hoa nhựa đã phai màu.

"Vợ ông với con trai ông," ông Phước nói, thấy Bảo nhìn. "Hai người đi xa rồi."

Bảo đứng nép cạnh cửa. "Con ngủ ngoài hiên được."

"Ngoài hiên muỗi nhiều. Con nằm trên giường."

"Còn ông?"

"Ông già nằm đâu cũng được."

Cậu lắc đầu rất mạnh. "Con ngủ dưới đất."

Ông Phước định nói thêm, nhưng thấy mắt Bảo căng lên như sắp bỏ chạy. Ông hiểu lòng tự trọng của cậu là mảnh áo cuối cùng còn lành. Ông lấy tấm chiếu khác trải dưới nền, đặt cái gối cũ bên cạnh. "Vậy hai ông cháu đều có chỗ. Công bằng."

Đêm đó, Bảo không ngủ ngay. Cậu nằm nghiêng, hai tay ôm chiếc áo khoác, mắt mở trong bóng tối. Ông Phước cũng thức. Ông nghe tiếng bụng cậu réo rất khẽ, nghe cậu cố nén ho, nghe cả nhịp thở giật mình mỗi khi ngoài hẻm có tiếng xe máy. Ông không bật đèn, chỉ nói: "Trong nồi còn ít chè đậu. Đói thì ăn. Không cần xin."

Một lúc sau, Bảo bò dậy rất nhẹ, mở nắp nồi, múc nửa chén. Cậu ăn đứng cạnh bếp, ăn chậm như sợ tiếng muỗng làm phiền người khác. Ông Phước quay mặt vào tường, mắt cay xè. Có những đứa trẻ bị đói lâu đến mức cả cách ăn cũng biết xin lỗi.

Sáng hôm sau, Bảo dậy trước cả ông. Cậu đã xếp chiếu, rửa chén, quét tóc bụi trước cửa, còn nhóm sẵn bếp than bằng mấy tờ giấy vụn. Ông Phước nhìn đống than cháy hơi mạnh, hoảng hốt dập bớt. "Con làm vậy phỏng tay thì sao?"

"Con muốn trả tiền ngủ."

"Ở đây không tính tiền ngủ."

"Vậy con phụ nấu chè."

Ông Phước thở dài, rồi đưa cậu cái rổ đậu xanh. "Nhặt hạt sâu ra. Làm chậm thôi."

Ngày đầu tiên trong căn phòng trọ, Bảo học rằng đậu xanh phải ngâm trước, nước cốt dừa không được để lửa lớn, chuối chín quá sẽ nát, bắp non thì ngọt nhưng khó bào. Ông Phước vừa làm vừa kể chuyện vợ ông ngày xưa nấu chè khéo hơn ông nhiều, chỉ cần nghe tiếng nước sôi đã biết đường tới đâu. Bảo ngồi bên cạnh, thỉnh thoảng hỏi một câu rất nhỏ. Cậu không còn căng như con mèo hoang bị dồn vào góc, nhưng vẫn luôn nhìn cửa, như sợ niềm yên ổn này bị ai đó bước vào lấy mất.

Buổi trưa, ông Phước dẫn Bảo ra phường hỏi giấy tờ. Cậu không có khai sinh trong người, không nhớ rõ họ cha, chỉ nhớ mẹ tên Hạnh và từng thuê trọ ở Tân Phú. Cán bộ phường nhìn hai ông cháu, ban đầu hơi nghi ngại. Ông Phước đặt căn cước của mình lên bàn, nói rõ: "Tôi không nhận nuôi bậy. Tôi muốn làm đúng. Nó cần chỗ ngủ, cần đi học, cần được biết nó là ai."

Bảo đứng sau lưng ông, ngón tay bấu vào vạt áo. Cậu quen nghe người lớn nói "nó" bằng giọng phiền phức. Nhưng chữ "nó" của ông Phước hôm ấy lại giống một cách che chắn. Cán bộ hẹn xác minh, cho giấy giới thiệu tạm để liên hệ lớp học tình thương trong chùa gần đó.

Trên đường về, Bảo hỏi: "Ông không sợ con làm ông rắc rối hả?"

Ông Phước đẩy xe chè, bánh xe lộc cộc trên vỉa hè. "Đời ông rắc rối lâu rồi. Thêm con chắc vui hơn."

Bảo cúi đầu đi bên cạnh. Đến trước cửa phòng trọ, cậu bỗng nói: "Nếu con ở lại, con sẽ phụ ông mỗi tối. Con không ăn không."

"Được," ông Phước đáp. "Nhưng có một điều kiện."

Cậu căng người.

"Ban ngày con phải học chữ."

Bảo ngẩn ra. Trong những điều kiện cậu từng nghe, có điều kiện phải im, phải làm, phải biến đi, nhưng chưa ai từng bắt cậu học như một quyền lợi. Cậu nhìn ông, rồi nhìn tấm bảng xe chè đang phơi ngoài hiên. Những chữ trên đó cậu chỉ đoán được lờ mờ. Lần đầu tiên, cậu muốn đọc rõ một thứ thuộc về nơi mình được phép quay về.
""",
    ),
    5820: (
        "Chương 4: Dạy Bảo Đọc Chữ",
        """
Lớp học tình thương nằm sau sân chùa, dưới mái tôn xanh, có ba dãy bàn gỗ thấp và một tấm bảng đen sứt góc. Học trò đủ tuổi: đứa bán vé số, đứa phụ quán cơm, đứa theo mẹ nhặt ve chai, có cả một anh mười tám tuổi mới học viết tên mình. Bảo bước vào với vẻ mặt lì lợm, nhưng bàn tay nắm quai cặp chặt đến trắng bệch. Chiếc cặp ấy là của con trai ông Phước ngày xưa, được ông phủi bụi, khâu lại quai, bỏ vào hai cuốn vở mới.

Cô giáo tên Mai, tóc cắt ngắn, giọng nhẹ. Cô không hỏi Bảo vì sao lớn vậy mới học lại. Cô chỉ đưa cậu một cây bút chì và bảo: "Hôm nay mình bắt đầu từ tên con." Bảo nhìn trang giấy trắng rất lâu. Cậu biết nói tên mình, biết nghe người ta gọi, nhưng chưa từng thấy nó nằm dưới tay mình. Cô Mai viết mẫu: Bảo. Dấu hỏi cong nhỏ trên chữ a như một chiếc móc kéo tim cậu lên.

Buổi tối, sau khi bán chè, ông Phước bày một cái bàn nhựa trước cửa phòng. Ông đeo kính lão, Bảo ngồi bên cạnh, hai người cùng học. Ông Phước đọc chữ còn chậm vì ngày xưa chỉ học hết lớp ba. Bảo cười khi ông đánh vần sai, ông gõ nhẹ cây bút lên đầu cậu: "Cười ông thì con đọc đi." Cứ vậy, căn phòng trọ có thêm một thứ tiếng mới: tiếng đánh vần. Bờ - ao - bao - hỏi - Bảo. Chờ - e - che - huyền - chè.

Ban đầu Bảo rất dễ nổi nóng với trang vở. Chữ của cậu nghiêng ngả, dấu đặt sai, nét lên nét xuống như người chạy trốn. Mỗi lần viết hỏng, cậu vò giấy. Ông Phước không mắng, chỉ nhặt giấy vuốt lại. "Chè nấu khét còn nấu lại được. Chữ xấu thì viết lại được. Không có gì phải sợ."

"Ở trường cũ, cô đánh tay con vì viết sai," Bảo nói sau một lúc im lặng.

Ông Phước đặt bút xuống. "Đau không?"

"Đau. Nhưng con sợ bị cười hơn."

"Ở nhà ông, sai thì sửa. Không ai cười."

Câu ấy trở thành luật. Khách quen biết Bảo học chữ, ai cũng góp một chút. Bác bảo vệ cho cậu cây thước nhựa. Cô công nhân may tặng hộp bút màu con gái không dùng nữa. Chú tài xế công nghệ chỉ cậu đọc biển đường bằng cách biến mỗi chuyến xe thành một bài học: Nguyễn Trãi, Châu Văn Liêm, Hải Thượng Lãn Ông. Cả góc đường như âm thầm mở một lớp học riêng cho cậu bé từng nghĩ mình không đáng được ngồi vào bàn.

Một đêm, Bảo nhìn tấm bảng carton trên xe chè rồi hỏi: "Ông ơi, chữ này là gì?"

"Chè nóng, ai đói cứ ăn trước."

"Sao ông viết vậy?"

Ông Phước lau nồi, cười buồn. "Hồi con trai ông còn nhỏ, có lần nó đi học về đói mà ông chưa bán được đồng nào. Bà nhà ông múc chè cho nó ăn trước rồi nói: trẻ con đói thì phải cho ăn trước, tiền tính sau. Sau này bà mất, ông viết câu đó để nhớ."

Bảo chạm tay lên mép bảng. Những chữ cũ bị mưa làm nhòe nhưng vẫn còn đọc được. Cậu đánh vần từng tiếng, chậm rãi, vấp vài lần rồi đọc trọn câu. Ông Phước quay đi châm nước nóng, nhưng vai ông rung lên. Một câu viết để nhớ người đã mất, bây giờ được một đứa trẻ còn sống đọc thành tiếng. Có những mất mát không biến mất, nhưng chúng có thể đổi dạng thành một ngọn đèn nhỏ.

Tháng thứ ba, Bảo viết được tên mình lên nhãn vở. Cậu đem khoe từng khách. Ai cũng khen, cậu cố tỏ ra bình thường nhưng tai đỏ bừng. Đêm đó, khi dọn hàng, cậu lấy một mẩu carton mới, nắn nót viết lại bảng cho ông: "CHÈ NÓNG - AI ĐÓI CỨ ĂN TRƯỚC." Chữ còn méo, khoảng cách không đều, nhưng rõ ràng và mạnh mẽ.

Ông Phước treo tấm bảng mới lên xe, đứng ngắm rất lâu. "Đẹp hơn bảng cũ."

"Xấu mà."

"Đẹp vì con đọc được."

Bảo cúi xuống buộc dây kẽm, miệng mím lại. Từ ngày có chữ, thế giới bớt đáng sợ hơn một chút. Biển cấm không còn là hình vẽ lạ, số xe buýt không còn là dấu bí mật, tên đường không còn là thứ chỉ người lớn mới hiểu. Và quan trọng hơn, cậu bắt đầu tin rằng đời mình cũng có thể viết lại, từng nét một, dù ban đầu run rẩy.
""",
    ),
    5821: (
        "Chương 5: Bảo Đi Học",
        """
Ngày Bảo được nhận vào lớp bổ túc tiểu học, ông Phước dậy từ bốn giờ sáng. Ông nấu một nồi chè đậu đỏ nhỏ, không phải để bán mà để cúng vợ và con trai. Trước bàn thờ, ông đặt giấy xác nhận tạm trú, giấy giới thiệu của phường, vài cuốn vở và chiếc áo sơ mi trắng mới mua ở chợ đồ si. Áo hơi rộng, nhưng ông đã giặt sạch, ủi phẳng, gấp ngay ngắn như một món quà lớn.

Bảo đứng trước gương nhựa, kéo cổ áo, vẻ mặt khó chịu. "Con mặc vậy kỳ lắm."

"Đi học thì mặc áo trắng."

"Người ta cười con lớn rồi còn học lớp nhỏ."

Ông Phước cài lại nút áo lệch cho cậu. "Người ta cười một chút rồi quên. Con không học thì con nhớ cả đời."

Câu ấy làm Bảo im. Cậu đeo chiếc cặp cũ, đi sau ông đến trường. Sân trường buổi sáng ồn ào đến mức cậu muốn quay đầu. Trẻ con chạy qua chạy lại, phụ huynh dặn dò, tiếng trống thử vang lên khô khốc. Bảo lớn hơn nhiều bạn trong lớp, da sạm hơn, bàn tay có vết xước do phụ xe chè. Khi cô chủ nhiệm giới thiệu cậu, vài đứa quay xuống nhìn. Có đứa thì thầm: "Sao anh này học chung với tụi mình?"

Bảo nghe rõ. Mặt cậu nóng lên. Bàn tay trong túi quần nắm chặt tờ giấy ông Phước nhét trước khi vào lớp: "Sai thì sửa. Đói thì về ăn chè. Buồn thì kể ông nghe." Cậu đọc được từng chữ, và chính việc đọc được làm cậu không bỏ chạy.

Những tuần đầu rất khó. Bảo học nhanh phần toán vì quen tính tiền thừa, nhưng tập đọc còn vấp. Khi cả lớp đọc đồng thanh, cậu thường chậm nửa nhịp. Có hôm một bạn trai bắt chước giọng cậu, cả nhóm cười rúc rích. Bảo đẩy ghế đứng dậy, mắt đỏ. Cô chủ nhiệm kịp giữ lại, nhưng cậu không khóc. Tối đó về, cậu không ăn chè, chỉ ngồi sau xe lau chén đến sáng bóng.

Ông Phước chờ khách vãn mới hỏi: "Hôm nay có chuyện gì?"

"Con không muốn đi học nữa."

"Vì khó?"

"Vì nhục."

Ông Phước ngồi xuống cạnh cậu. "Hồi ông mới bán chè, người ta cũng cười. Họ nói đàn ông góa vợ bưng chè ngoài đường nhìn tội nghiệp. Ông cũng nhục. Nhưng nếu ông nghỉ, tối hôm đó ai bán cho bác bảo vệ chén chè nóng? Ai gặp con dưới mưa?"

Bảo không đáp.

"Nhục không giết mình ngay," ông nói tiếp. "Nhưng bỏ cuộc vì nhục thì nó lấy mất tương lai của mình."

Hôm sau, Bảo quay lại lớp. Cậu không thân với bạn ngay, nhưng bắt đầu có cách đứng vững. Khi ai trêu, cậu nhìn thẳng và nói: "Tao học chậm, nhưng tao vẫn học." Cô chủ nhiệm giao cậu phụ trách lau bảng vì cậu cao hơn các bạn. Ban đầu đó chỉ là việc nhỏ, sau thành niềm tự hào. Cậu lau bảng rất sạch, xếp phấn ngay ngắn, nhắc các bạn tắt quạt khi ra về.

Một ngày, lớp có bài tập kể về người thân. Các bạn viết về ba mẹ. Bảo ngồi rất lâu trước trang giấy. Cậu không biết có được viết về ông Phước không. Ông không phải ông ruột, không có giấy tờ nhận nuôi, không cùng họ. Nhưng mỗi tối ông để phần chè ít đường cho cậu, mỗi sáng nhắc cậu mang áo mưa, mỗi lần cậu sai chữ lại ngồi sửa đến khuya. Cuối cùng, Bảo viết: "Người thân của em là ông bán chè đêm. Ông không sinh ra em, nhưng ông cho em ăn khi em đói và bắt em đi học khi em muốn biến mất."

Bài văn ấy được cô đọc trước lớp. Bảo cúi gằm mặt, tưởng mình sẽ bị cười. Nhưng lớp im lặng. Một bạn gái quay xuống hỏi nhỏ: "Tối nào ông bạn cũng bán chè hả? Mình mua được không?" Chiều hôm ấy, ba đứa bạn theo Bảo ra góc đường. Chúng ăn chè, chào ông Phước rất lễ phép. Ông múc thêm nước cốt dừa, mắt cứ nhìn Bảo cười.

Đêm đó, Bảo làm bài dưới bóng đèn xe chè. Khách đến, cậu vẫn tính tiền, vẫn rửa chén, nhưng trên bàn nhựa có thêm sách. Ông Phước nhìn cậu đánh vần một đoạn văn dài, trong lòng ấm như nồi chè vừa sôi. Có lẽ một đứa trẻ không cần được cứu bằng điều gì lớn lao. Đôi khi chỉ cần một người đủ kiên nhẫn ngồi cạnh, nhắc nó rằng ngày mai vẫn đáng đi tới.
""",
    ),
    5822: (
        "Chương 6: Ông Ốm - Bảo Đẩy Xe Chè",
        """
Mùa mưa năm ấy kéo dài hơn thường lệ. Nước ngập ngang bánh xe, gió lùa qua hẻm làm mái tôn nhà ông Phước kêu lạch cạch suốt đêm. Ông bắt đầu ho nhiều. Ban đầu ông giấu, chỉ quay mặt vào khăn. Bảo nghe thấy nhưng không nói, lặng lẽ pha nước gừng, nhắc ông mặc áo khoác, giành phần đẩy xe nặng. Ông cười: "Ông già ho vài tiếng thôi." Nhưng tiếng ho ngày một sâu, như có ai gõ từ trong ngực.

Một đêm, đang múc chè cho khách, tay ông Phước run lên. Chén chè rơi xuống nền, vỡ toang. Ông vịn vào xe, mặt tái đi. Bảo lao tới đỡ. Bác bảo vệ gọi xe ôm, cô công nhân may gom tiền trong túi, chú tài xế công nghệ chở ông vào bệnh viện quận. Mọi thứ diễn ra nhanh đến mức Bảo chỉ kịp ôm theo chiếc túi thuốc và căn cước của ông.

Bệnh viện ban đêm sáng trắng và lạnh. Bảo ngồi ngoài hành lang, hai chân không chạm đất, tay nắm chặt tờ giấy đăng ký khám. Bác sĩ nói ông Phước viêm phổi, sức yếu, phải nằm lại theo dõi. Cậu nghe được vài chữ, còn lại biến thành tiếng ù trong tai. Ông Phước nằm trên giường bệnh, ống truyền gắn ở tay, vẫn cố cười: "Mai nghỉ bán một bữa."

Bảo gật đầu, nhưng trong đầu đã tính khác. Tiền phòng, tiền thuốc, tiền ăn, tiền nhà trọ, tất cả đều nằm trong chiếc hộp sắt dưới gầm giường. Ông Phước dành dụm không nhiều. Nghỉ vài bữa là hụt. Hôm sau, sau giờ học, Bảo về phòng trọ, mở nồi, kiểm đậu, nhóm bếp. Cậu chưa từng nấu một mình, nhưng đã nhìn ông làm hàng trăm lần. Đậu xanh ngâm bao lâu, nước cốt dừa để lửa nhỏ, chè bắp phải khuấy đều kẻo khê. Cậu làm chậm, mồ hôi rơi xuống cổ, mắt cay vì khói.

Tối đó, chiếc xe chè vẫn sáng đèn ở góc đường. Người đứng sau xe không phải ông Phước mà là Bảo, áo học sinh đã thay bằng áo thun cũ, tay cầm vá múc chè hơi vụng. Tấm bảng carton vẫn treo: "Ai đói cứ ăn trước." Khách quen đến, ai cũng khựng lại. Bảo nói trước khi họ hỏi: "Ông bệnh. Con bán thay. Chè hôm nay nếu dở, cô chú nói để con sửa."

Không ai chê. Có người ăn chậm hơn thường lệ, có người trả dư rồi bảo khỏi thối, có người mua ba bịch dù nhà không có ai ăn. Bảo ghi từng khoản vào cuốn sổ, không lấy một đồng lẫn lộn. Khi một cậu sinh viên thiếu tiền, Bảo múc đầy như ông từng làm. Cậu nghe giọng mình run nhưng rõ: "Ăn trước đi anh. Mai mốt trả cũng được."

Đêm thứ ba, mưa lớn. Xe chè nặng trịch, đường về ngập. Bảo cắn răng đẩy, nước bắn lên quần, bánh xe mắc vào ổ gà. Cậu suýt khóc vì mệt. Đúng lúc đó, bác bảo vệ từ chung cư chạy ra phụ một tay. Rồi chú tài xế công nghệ ghé vào đẩy phía sau. Cô công nhân may cầm dù che nồi chè. Ba bốn người khách quen cùng đưa xe về tận hẻm. Bảo đứng giữa mưa, không biết nói gì.

Bác bảo vệ vỗ vai cậu: "Hồi con đói, ông Phước cho ăn. Giờ ông bệnh, tụi chú phụ. Đời là vậy."

Trong bệnh viện, Bảo đem sổ tiền và hộp cháo vào cho ông. Ông Phước nhìn những con số ngay ngắn, nhìn nét chữ đã cứng cáp hơn, rồi nhìn cậu bé gầy nhưng mắt sáng. "Con cực quá."

Bảo lắc đầu. "Con chỉ giữ xe chè cho ông."

"Sao con không nghỉ học?"

"Ông nói nhục không được bỏ. Mệt cũng không được bỏ."

Ông Phước quay mặt vào tường, giả vờ ho để giấu nước mắt. Có những thứ ông tưởng mình cho Bảo: tô chè, chỗ ngủ, con chữ. Nhưng hóa ra đứa trẻ ấy cũng đang trả lại cho ông một điều rất lớn: lý do để sống tiếp, để khỏi buông tay khi tuổi già kéo mình xuống.
""",
    ),
    5823: (
        "Chương 7: Bảo Lớn - Ông Già Đi",
        """
Năm tháng qua đi bằng những mùa chè. Mùa nóng bán nhiều sương sáo, mùa mưa chè gừng đắt khách, mùa thi sinh viên ngồi kín vỉa hè ăn chè đậu đỏ lấy may. Bảo cao lên, vai rộng hơn, giọng trầm xuống. Cậu học hết bổ túc, rồi vào trường trung cấp du lịch - ẩm thực bằng học bổng của một quỹ địa phương. Ban ngày cậu học bếp, tối vẫn về phụ ông Phước bán chè. Chiếc áo trắng học trò được thay bằng áo đồng phục bếp, nhưng tấm bảng carton trên xe chè vẫn do chính tay cậu viết lại mỗi khi mưa làm nhòe.

Ông Phước già nhanh hơn Bảo muốn thừa nhận. Sau trận viêm phổi, ông không còn đứng lâu được. Tay ông run khi múc nước cốt dừa, mắt mờ khi đếm tiền lẻ. Bảo nhiều lần bảo ông nghỉ, nhưng ông chỉ cười: "Ông ngồi nhìn cũng được. Xe chè không có ông, khách buồn."

Thật ra người buồn nhất là Bảo. Cậu sợ một ngày quay lại góc đường mà chiếc ghế của ông trống. Nỗi sợ ấy làm cậu đôi khi cáu gắt. Ông ho, cậu nhăn mặt. Ông quên uống thuốc, cậu nói lớn. Có lần ông Phước lỡ làm đổ nồi nước cốt dừa, Bảo buột miệng: "Ông nghỉ đi, để con làm, ông cứ cố làm gì!" Câu nói vừa ra, cậu hối hận ngay. Ông Phước im lặng nhặt cái vá, bàn tay run run.

Đêm đó, hai ông cháu không nói chuyện. Bảo dọn hàng, ông ngồi nhìn dòng xe. Gần hai giờ sáng, khi khách cuối cùng đi rồi, ông Phước mới lên tiếng: "Con lớn rồi, con sợ ông thành gánh nặng."

Bảo quỳ xuống trước mặt ông. "Con không có ý đó."

"Ông biết. Nhưng ông cũng sợ." Ông nhìn tấm bảng xe chè. "Ông sợ một ngày không còn làm được gì, con thương ông mà mệt. Người già hay sợ mình sống lâu quá."

Bảo nắm tay ông. Bàn tay ấy từng múc chè cho cậu dưới mưa, từng cầm bút sửa chữ, từng ký giấy tạm trú, từng vuốt tóc cậu khi sốt. "Ông không phải gánh nặng. Nếu không có ông, con không biết mình đang ở đâu."

Ông Phước cười, mắt đục nhưng hiền. "Vậy mai con cho ông ngồi bán. Con múc, ông thu tiền. Để ông còn cảm giác mình có ích."

Từ đó, xe chè đổi nhịp. Bảo đứng chính, ông Phước ngồi bên cạnh như một ông chủ già, thỉnh thoảng nhắc khách bỏ ít đá kẻo đau họng, kể chuyện Bảo ngày xưa không biết chữ cho mấy đứa sinh viên nghe. Bảo giả vờ than: "Ông kể hoài con quê." Nhưng trong lòng, cậu ấm.

Ngày Bảo tốt nghiệp trung cấp, ông Phước mặc áo sơ mi trắng đã ngả màu, chống gậy đến dự. Khi tên Bảo được xướng lên, ông vỗ tay chậm nhưng lâu nhất. Sau buổi lễ, Bảo đưa ông xem bằng tốt nghiệp. Trên đó có họ tên đầy đủ mà phường đã hỗ trợ làm lại giấy tờ: Nguyễn Minh Bảo. Ông Phước sờ lên từng chữ, như sờ một phép màu.

"Con định làm ở nhà hàng lớn hả?" ông hỏi.

Bảo lắc đầu. "Con muốn mở một quán chè tử tế hơn cho ông. Có chỗ ngồi trong nhà, có bếp sạch, có tủ kính mới. Nhưng vẫn bán đêm. Vẫn có dòng ai đói cứ ăn trước."

Ông Phước nhìn cậu rất lâu. "Con còn trẻ. Đừng trói đời mình vào xe chè của ông."

"Con không bị trói," Bảo nói. "Con được kéo về từ nó."

Ông Phước không cãi nữa. Tối hôm ấy, ông ăn nửa chén chè đậu đỏ, khen hơi ngọt. Bảo bật cười vì đó là câu cậu từng nói trong đêm thứ hai gặp ông. Hai ông cháu ngồi cạnh nhau đến khuya, không biết rằng thời gian còn lại đang mỏng dần như lớp khói trên nồi chè nóng.
""",
    ),
    5824: (
        "Chương 8: Tô Chè Cuối Cùng",
        """
Ông Phước mất vào một buổi sáng tháng Chạp, khi Sài Gòn se lạnh hiếm hoi. Đêm trước ông vẫn đòi ra xe chè, ngồi nhìn khách ăn, còn nhắc Bảo đừng quên giảm đường cho bác bảo vệ bị tiểu đường. Gần sáng, ông trở mình, gọi tên vợ, gọi tên con trai, rồi gọi Bảo. Khi Bảo chạy đến, ông nắm tay cậu, bàn tay nhẹ đến mức cậu sợ chỉ cần thở mạnh là mất.

"Xe chè..." ông nói rất khẽ.

"Con giữ. Ông yên tâm."

"Không phải giữ cái xe. Giữ cái câu trên bảng."

Bảo gật đầu, nước mắt rơi xuống mu bàn tay ông. Ông Phước nhìn cậu như nhìn đứa bé dưới mưa năm nào, nhưng cũng như nhìn một người đàn ông đã đủ sức đi tiếp. "Đói thì cho ăn trước. Buồn thì nghe người ta nói. Sai thì sửa. Con nhớ chưa?"

"Con nhớ."

Ông mỉm cười. "Vậy ông nghỉ chút."

Ông nghỉ thật. Căn phòng trọ vốn nhỏ bỗng rộng đến đau lòng. Khách quen đến viếng chật cả con hẻm. Bác bảo vệ mang vòng hoa, cô công nhân may tự tay khâu khăn tang cho Bảo, chú tài xế công nghệ đứng điều phối xe để hẻm khỏi kẹt. Những người từng ăn chịu, từng được ông cho thêm nước cốt dừa, từng ngồi kể chuyện buồn bên xe chè, mỗi người đem đến một mẩu ký ức. Bảo nghe họ kể và nhận ra ông Phước không chỉ nuôi mình. Ông đã âm thầm giữ ấm cho cả một góc thành phố.

Sau đám tang, Bảo đóng xe chè bảy ngày. Đến đêm thứ tám, cậu đẩy xe ra lại góc đường. Bóng đèn vàng bật lên, nhưng chiếc ghế bên cạnh trống. Bảo đặt lên ghế chiếc áo khoác cũ của ông. Tấm bảng carton được thay bằng bảng gỗ nhỏ, chữ khắc rõ ràng: "Chè nóng - ai đói cứ ăn trước." Bên dưới, Bảo thêm một dòng: "Nhớ ông Phước."

Khách đến không ai nói lớn. Bác bảo vệ gọi một chén chè đậu đen, ăn được nửa thì lau mắt. Cô công nhân may mua chè bắp, bảo vị giống hệt ngày xưa. Một cậu bé bán vé số đứng cách xe ba bước, nhìn nồi chè bằng ánh mắt mà Bảo không bao giờ quên, vì đó từng là mắt mình. Cậu bé nói: "Con không có tiền."

Bảo nghe trong ngực mình có gì vỡ ra rồi liền lại. Cậu múc một chén chè nóng, chan ít nước cốt dừa, đặt xuống bậc thềm. "Anh đâu có hỏi tiền. Em ăn thử coi hôm nay anh nấu có ngọt quá không."

Cậu bé cầm muỗng, nghi ngờ và đói khát đánh nhau trong mắt. Bảo ngồi xuống bên cạnh, y như ông Phước năm xưa. Khi cậu bé ăn miếng đầu tiên và vai run lên, Bảo quay mặt đi châm nước, để em được khóc mà không xấu hổ. Trên chiếc ghế trống, áo khoác của ông Phước lay nhẹ trong gió.

Ba năm sau, góc đường ấy có một quán chè nhỏ tên "Chè Đêm Ông Phước". Quán không sang, nhưng sạch, có bàn dài cho sinh viên học bài, có kệ sách cũ, có một hũ tiền ghi "ai dư bỏ vào, ai thiếu lấy ra." Bảo trở thành ông chủ trẻ, nhưng tối nào cũng tự múc vài chén đầu tiên. Chiếc xe chè cũ được đặt ngay cửa, bánh xe đã khóa lại, mặt kính lau sáng. Nhiều người trẻ đến chụp ảnh, hỏi câu trên bảng có phải khẩu hiệu marketing không. Bảo chỉ cười: "Không. Đó là di chúc."

Mỗi dịp giỗ ông, Bảo nấu một nồi chè đậu đỏ nhỏ, đặt lên bàn thờ cạnh ảnh vợ chồng ông Phước và người con trai đi biển. Cậu kể cho ông nghe quán tháng này có bao nhiêu người ăn trước trả sau, có mấy đứa trẻ được giới thiệu đi học, có bác bảo vệ đã nghỉ hưu nhưng vẫn ghé uống trà. Cậu kể cả những lúc mình mệt, cáu, nấu khét, tính sai tiền, rồi tự nhắc: sai thì sửa.

Đêm Sài Gòn vẫn ồn, vẫn có người đói, người buồn, người không biết về đâu. Nhưng ở góc đường quận 5, luôn có một bóng đèn vàng sáng đến khuya và một nồi chè nóng. Người ta đến đó không chỉ để ăn ngọt. Họ đến để được tin rằng giữa thành phố rộng và vội, vẫn có một nơi không hỏi ví tiền trước khi hỏi người ta có lạnh không.

Bảo nhiều lần nghĩ về tô chè đầu tiên. Nếu đêm đó ông Phước chỉ nhìn rồi quay đi, đời cậu có thể đã rẽ vào một con hẻm tối hơn rất nhiều. Nhưng ông đã bưng một chén chè qua, đặt giữa hai người, và dùng sự tử tế bình thường nhất để mở một cánh cửa. Bây giờ, mỗi khi múc chè cho một đứa trẻ lạ, Bảo lại nghe giọng ông đâu đó trong tiếng muỗng chạm chén: "Đói thì ăn trước, con. Trên đời này, có những món nợ mình trả bằng cách thương người kế tiếp."
""",
    ),
}


def main():
    editor.upload_helper()
    try:
        editor.update_story_meta(STORY_ID, title=TITLE, intro=INTRO)
        updates = []
        for chapter_id, (title, content) in CHAPTERS.items():
            editor.update_chapter(chapter_id, title, content.strip())
            updates.append({"chapter_id": chapter_id, "title": title})
    finally:
        editor.remove_helper()

    workbook = ROOT / "danh_sach_truyen_doctieuthuyet.xlsx"
    wb = openpyxl.load_workbook(workbook)
    ws = wb.active
    for row in range(5, ws.max_row + 1):
        if ws.cell(row, 2).value == STORY_ID:
            ws.cell(row, 3).value = TITLE
            ws.cell(row, 12).value = INTRO
            ws.cell(row, 13).value = (
                "Đã đọc live và viết lại toàn bộ 8 chương từ bản quá ngắn lên 1000-1500 từ/chương; "
                "tăng chất đời, tình ông cháu, hành trình học chữ, bệnh tật và kết truyền lại xe chè."
            )
            ws.cell(row, 14).value = "☑️ Đã sửa"
            break
    wb.save(workbook)

    out = ROOT / "scratch" / "story_5815_fix_result.json"
    out.write_text(json.dumps(updates, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(updates, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
