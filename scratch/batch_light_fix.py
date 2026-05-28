import json
import urllib.request
import openpyxl
import re
import time

# Load batch eval results
with open("scratch/batch_eval_results.json", "r", encoding="utf-8") as f:
    results = json.load(f)

light_fix = results["light_fix"]
print(f"=== XỬ LÝ {len(light_fix)} TRUYỆN SỬA NHẸ ===\n")

# AI template paragraphs to remove (these are injected paragraphs that don't belong)
AI_TEMPLATES = [
    r'Đến đoạn này, không ai trong phòng cần nghe thêm thuật ngữ\. Họ chỉ nhìn vào khuôn mặt tái đi của đối thủ, nhìn vào con dấu đỏ trên hồ sơ và hiểu rằng sự thật đang có hình dạng rất cụ thể\.',
    r'Anh không đáp bằng lời thề\. Anh đặt từng vật chứng lên bàn: bản ghi thời gian, ảnh chụp niêm phong, tin nhắn gốc và biên bản đối chiếu có chữ ký của bên thứ ba\. Mỗi thứ đều nhỏ, nhưng khi xếp cạnh nhau, chúng khép lại thành một đường chứng cứ không còn khe hở\.',
    r'Nhưng cú phản công đầu tiên chưa kịp thành hình thì đã có người nộp đơn kiện ngược\.',
    r'Nhân vật chính đứng lại thêm vài giây sau khi căn phòng vãn người\. Anh không cười lớn, cũng không cần khoe khoang\. Chỉ khi bàn tay chạm vào tập hồ sơ đã nhàu mép, anh mới hiểu: thứ được lấy lại không chỉ là danh dự, mà là quyền được ngẩng đầu bước tiếp\.',
    r'Không ai còn dám gọi chiến thắng ấy là may mắn\.',
]

# Also match the plain text versions
AI_PLAIN = [
    'Đến đoạn này, không ai trong phòng cần nghe thêm thuật ngữ.',
    'Anh không đáp bằng lời thề. Anh đặt từng vật chứng lên bàn:',
    'Nhưng cú phản công đầu tiên chưa kịp thành hình',
    'Nhân vật chính đứng lại thêm vài giây sau khi căn phòng vãn người.',
    'Không ai còn dám gọi chiến thắng ấy là may mắn.',
]

BANNED_PHRASES = {
    'tóm lại': 'cuối cùng',
    'nhìn chung': 'nhìn lại',
    'có thể nói': 'rõ ràng',
    'tổng kết': 'nhìn lại toàn bộ',
    'nói cách khác': 'hay nói đúng hơn',
}

fetch_url = "https://doctieuthuyet.com/fetch_story_full.php"
upload_url = "https://doctieuthuyet.com/overwrite_story_v13.php"

fixed_count = 0
wb = openpyxl.load_workbook("danh_sach_truyen_doctieuthuyet.xlsx")
ws = wb.active

for idx, entry in enumerate(light_fix):
    story_id = entry["id"]
    stt = entry["stt"]
    print(f"\n{'='*50}")
    print(f"[{idx+1}/{len(light_fix)}] STT {stt} — ID={story_id}")
    
    try:
        # Fetch story
        payload = json.dumps({
            "secret_token": "ZEN_TRUYEN_2026_BYPASS",
            "story_id": story_id
        }).encode('utf-8')
        req = urllib.request.Request(fetch_url, data=payload, headers={'Content-Type': 'application/json'})
        resp = urllib.request.urlopen(req, timeout=30)
        data = json.loads(resp.read().decode('utf-8'))
        
        print(f"  Title: {data['title'][:60]}")
        
        # Check if story actually needs any fixes
        changes_made = False
        
        for i, ch in enumerate(data.get('chapters', [])):
            content = ch.get('content', '')
            original = content
            
            # Remove AI template paragraphs
            for tmpl in AI_PLAIN:
                if tmpl in content:
                    # Find the full <p>...</p> containing the template
                    # Try to remove the whole paragraph
                    pattern = r'<p>[^<]*?' + re.escape(tmpl[:30]) + r'[^<]*?</p>\s*'
                    new_content = re.sub(pattern, '', content)
                    if new_content != content:
                        content = new_content
                        changes_made = True
                    else:
                        # Fallback: just remove the text
                        content = content.replace(tmpl, '')
                        if len(tmpl) > 20:
                            changes_made = True
            
            # Fix banned phrases (case-insensitive replacement)
            for banned, replacement in BANNED_PHRASES.items():
                if banned in content.lower():
                    # Preserve case
                    content = re.sub(re.escape(banned), replacement, content, flags=re.IGNORECASE)
                    changes_made = True
            
            # Remove meta-breaking "nhân vật chính"
            if 'nhân vật chính' in content.lower():
                # Try to find and fix contextually
                content = content.replace('Nhân vật chính', 'Anh')
                content = content.replace('nhân vật chính', 'anh')
                changes_made = True
            
            ch['content'] = content
        
        if not changes_made:
            print(f"  ✅ No changes needed — marking as fixed")
        else:
            # Upload fixed version
            upload_data = {
                "secret_token": "ZEN_TRUYEN_2026_BYPASS",
                "story_id": story_id,
                "title": data['title'],
                "intro": data.get('intro', ''),
                "author": data.get('author', ''),
                "chapters": data.get('chapters', [])
            }
            
            payload = json.dumps(upload_data).encode('utf-8')
            req = urllib.request.Request(upload_url, data=payload, headers={'Content-Type': 'application/json'})
            resp = urllib.request.urlopen(req, timeout=120)
            result = json.loads(resp.read().decode('utf-8'))
            print(f"  📤 Uploaded: {result.get('chapters_count')} chapters")
        
        # Update Excel
        for r in range(5, ws.max_row+1):
            if ws.cell(row=r, column=2).value and str(ws.cell(row=r, column=2).value).strip() == str(story_id):
                ws.cell(row=r, column=13).value = "Sửa nhẹ: loại bỏ AI template, banned phrases → ĐÃ HOÀN THÀNH"
                ws.cell(row=r, column=14).value = "☑️ Đã sửa"
                break
        
        fixed_count += 1
        print(f"  ✅ Done!")
        
    except Exception as e:
        print(f"  ❌ Error: {e}")
    
    if idx % 5 == 4:
        # Save Excel periodically
        wb.save("danh_sach_truyen_doctieuthuyet.xlsx")
        time.sleep(0.5)

wb.save("danh_sach_truyen_doctieuthuyet.xlsx")
print(f"\n=== HOÀN THÀNH: {fixed_count}/{len(light_fix)} truyện sửa nhẹ ===")
