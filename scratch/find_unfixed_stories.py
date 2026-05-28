import openpyxl

wb = openpyxl.load_workbook("danh_sach_truyen_doctieuthuyet.xlsx")
ws = wb.active

print("=== TRUYỆN CHƯA SỬA ===\n")
unfixed = []
for row in range(5, ws.max_row + 1):
    fix_status = ws.cell(row=row, column=14).value
    if fix_status and "chưa" in str(fix_status).lower():
        stt = ws.cell(row=row, column=1).value
        story_id = ws.cell(row=row, column=2).value
        title = ws.cell(row=row, column=3).value
        status = ws.cell(row=row, column=8).value
        issues = ws.cell(row=row, column=13).value
        unfixed.append({
            'row': row,
            'stt': stt,
            'id': story_id,
            'title': str(title)[:80] if title else 'N/A',
            'status': status,
            'issues': str(issues)[:120] if issues else 'N/A'
        })

for i, s in enumerate(unfixed):
    print(f"{i+1}. STT {s['stt']} — ID={s['id']} — [{s['status']}]")
    print(f"   {s['title']}")
    print(f"   Lỗi: {s['issues']}")
    print()

print(f"Tổng: {len(unfixed)} truyện chưa sửa")
