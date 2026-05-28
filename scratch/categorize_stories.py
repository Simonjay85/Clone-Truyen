import openpyxl
import json

wb = openpyxl.load_workbook("danh_sach_truyen_doctieuthuyet.xlsx")
ws = wb.active

# Categorize all unfixed stories
categories = {
    "mark_only": [],      # "Đã tối ưu" but not marked as fixed
    "light_fix": [],      # Banned phrases, short chapters, weak plot
    "rewrite": [],        # Chinese names, supernatural, template text
    "need_fetch": []       # Need to fetch to determine
}

for row in range(5, ws.max_row + 1):
    fix_status = ws.cell(row=row, column=14).value
    if not fix_status or "chưa" not in str(fix_status).lower():
        continue
    
    stt = ws.cell(row=row, column=1).value
    story_id = ws.cell(row=row, column=2).value
    title = str(ws.cell(row=row, column=3).value or '')
    status = ws.cell(row=row, column=8).value
    issues = str(ws.cell(row=row, column=13).value or '')
    
    entry = {
        'row': row, 'stt': stt, 'id': story_id,
        'title': title[:80], 'status': status,
        'issues': issues[:200]
    }
    
    # Categorize based on issues text
    if "Đã tối ưu" in issues and "10/10" in issues:
        categories["mark_only"].append(entry)
    elif any(kw in issues.lower() for kw in ['tên tq', 'tên trung quốc', 'siêu nhiên', 'rewrite']):
        categories["rewrite"].append(entry)
    elif any(kw in title for kw in ['Tô Khanh', 'Lâm Uyển', 'Thẩm Triệt']):
        categories["rewrite"].append(entry)
    else:
        categories["need_fetch"].append(entry)

print("=== PHÂN LOẠI TRUYỆN CHƯA SỬA ===\n")

print(f"📗 CHỈ CẦN ĐÁNH DẤU (đã tối ưu 10/10): {len(categories['mark_only'])} truyện")
for s in categories['mark_only'][:5]:
    print(f"   STT {s['stt']} — ID={s['id']} — {s['title']}")
if len(categories['mark_only']) > 5:
    print(f"   ... và {len(categories['mark_only'])-5} truyện nữa")

print(f"\n🔴 CẦN REWRITE (đã xác định): {len(categories['rewrite'])} truyện")
for s in categories['rewrite']:
    print(f"   STT {s['stt']} — ID={s['id']} — {s['title']}")

print(f"\n🟡 CẦN FETCH ĐỂ ĐÁNH GIÁ: {len(categories['need_fetch'])} truyện")
for s in categories['need_fetch'][:10]:
    print(f"   STT {s['stt']} — ID={s['id']} — [{s['status']}] — {s['title']}")
if len(categories['need_fetch']) > 10:
    print(f"   ... và {len(categories['need_fetch'])-10} truyện nữa")

print(f"\n=== TỔNG: {sum(len(v) for v in categories.values())} truyện ===")

# Save categories for pipeline
with open("scratch/story_categories.json", "w", encoding="utf-8") as f:
    json.dump(categories, f, ensure_ascii=False, indent=2)
print("\nSaved to scratch/story_categories.json")
