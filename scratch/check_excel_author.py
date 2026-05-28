import openpyxl

wb = openpyxl.load_workbook("/Users/aaronnguyen/TN/App/doctieuthuyet/danh_sach_truyen_doctieuthuyet.xlsx")
sheet = wb.active

for row in range(1, sheet.max_row + 1):
    val = sheet.cell(row, 2).value
    if val and "Nghệ Nhân Trà Sen" in str(val):
        print(f"Row {row}:")
        print(f"  Title:  {sheet.cell(row, 2).value}")
        print(f"  Author: {sheet.cell(row, 3).value}")
        print(f"  Status: {sheet.cell(row, 4).value}")
        break
