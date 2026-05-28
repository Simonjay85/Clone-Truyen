import json
import urllib.request
import openpyxl

story_data = {
    "secret_token": "ZEN_TRUYEN_2026_BYPASS",
    "story_id": 5618,
    "title": "Bị Mẹ Bỏ Rơi Ở Cổng Chùa Từ Lúc Sơ Sinh, Tôi Lớn Lên Thành CEO Tập Đoàn May Mặc Lớn Nhất Đông Nam Á",
    "intro": "<p>Một đêm mưa tháng Mười, ai đó đặt em bé gái trong chiếc rổ tre trước cổng chùa Linh Ứng, Đà Nẵng. Không tên, không tuổi, không cha mẹ — chỉ có tấm khăn len đỏ quấn quanh và một mảnh giấy nhàu: \"Xin hãy nuôi con giúp.\"</p>\n<p>Sư thầy Thích Minh Đức đặt tên cô là Trần Thị Ngọc — lấy họ Trần theo pháp danh, đặt tên Ngọc vì \"trong bùn có ngọc.\" Cô lớn lên trong chùa, bị trêu chọc vì không có nguồn gốc, bị từ chối ở mọi nơi vì hai chữ \"mồ côi.\"</p>\n<p>Nhưng đôi bàn tay khéo léo may vá từ năm mười tuổi, một ý chí thép không gì bẻ gãy, và tấm khăn len đỏ luôn mang theo bên mình — đã đưa cô từ cổng chùa đến sân khấu thời trang quốc tế, từ cô bé không gia đình đến CEO tập đoàn may mặc lớn nhất Đông Nam Á.</p>",
    "author": "Trần Thị Ngọc",
    "seo": {
        "focus_keyword": "bị bỏ rơi cổng chùa, CEO may mặc, sảng văn truyền cảm hứng",
        "seo_title": "Bị Mẹ Bỏ Rơi Cổng Chùa — Lớn Lên Thành CEO May Mặc | Đọc Tiểu Thuyết",
        "seo_description": "Câu chuyện cô bé bị bỏ rơi ở cổng chùa Linh Ứng, lớn lên bằng nghề may, vượt qua nghèo khổ thành CEO tập đoàn may mặc Đông Nam Á. Truyện sảng văn truyền cảm hứng."
    },
    "chapters": []
}

ch1 = {
    "title": "Chương 1: Đêm Mưa Cổng Chùa",
    "content": """<p>Sư thầy Thích Minh Đức nghe tiếng khóc lúc ba giờ sáng.</p>

<p>Mưa xối xả, gió Đà Nẵng giật từng hồi, lay cửa gỗ chùa Linh Ứng kêu cọt kẹt. Thầy vừa xong thời kinh khuya, đang gấp y hậu, thì tiếng khóc trẻ sơ sinh xuyên qua tiếng mưa — yếu ớt nhưng đều đặn, như thể đứa bé đã khóc cả đêm cho đến khi cổ họng khan hết, chỉ còn lại tiếng nấc.</p>

<p>Thầy chống gậy ra cổng. Mắt thầy nhìn xuống: một chiếc rổ tre đặt ngay trước bậc tam cấp, bên trong là em bé gái đỏ hỏn, quấn trong tấm khăn len đỏ đã ướt sũng. Nước mưa chảy dọc má em bé, hòa lẫn với nước mắt, trông không biết đâu là mưa đâu là khóc.</p>

<p>Mảnh giấy kẹp trong khăn, mực bút bi đã nhòe gần hết, chỉ đọc được mấy chữ: "Xin hãy nuôi con giúp."</p>

<p>Sư thầy đứng dưới mưa rất lâu. Thầy bảy mươi hai tuổi, tu hành bốn mươi năm, từng chứng kiến nhiều nghiệp đời — nhưng đứng trước một đứa bé bị bỏ rơi giữa đêm mưa, thầy vẫn thấy trái tim quặn lại.</p>

<p>Thầy bế em bé lên, đưa vào chánh điện, lấy khăn khô lau người. Em bé ngừng khóc khi chạm tay thầy — có thể vì hơi ấm, có thể vì mùi trầm hương thoảng nhẹ trong điện, hoặc đơn giản vì lần đầu tiên từ khi ra đời, có người ôm cô ấy đủ lâu để cô ngừng run.</p>

<p>"A Di Đà Phật. Phật không bỏ ai. Và thầy cũng không."</p>

<p>Sáng hôm sau, thầy đặt tên cho em bé: Trần Thị Ngọc. Họ Trần theo pháp danh. Tên Ngọc vì "trong bùn có ngọc, trong khổ đau có giác ngộ."</p>

<p>Thầy giữ lại tấm khăn len đỏ — tấm khăn duy nhất mà mẹ em bé để lại. Thầy giặt sạch, phơi khô, gấp gọn, đặt dưới gối Ngọc.</p>

<p>"Ngọc ơi, con không cần biết mình đến từ đâu. Con chỉ cần biết mình sẽ đi đến đâu."</p>

<p>Câu nói đó, hai mươi tám năm sau, Trần Thị Ngọc khắc lên tấm bảng đồng ở sảnh chính tập đoàn NGỌC Fashion — tập đoàn may mặc bền vững lớn nhất Đông Nam Á, doanh thu bốn nghìn tỷ đồng, ba mươi nghìn nhân viên, xuất khẩu hai mươi sáu quốc gia.</p>

<p>Nhưng đó là hai mươi tám năm sau. Còn bây giờ, cô mới chỉ là một đứa trẻ nặng ba ký hai, khóc trong đêm mưa, và không biết rằng ngoài kia, một người phụ nữ mười bảy tuổi đang ngồi trong xe khách Đà Nẵng — Sài Gòn, tay bấu chặt thanh ghế, khóc không thành tiếng, mang theo nỗi đau mà suốt hai mươi năm bà không dám kể cho ai.</p>"""
}

ch2 = {
    "title": "Chương 2: Con Bé Chùa Linh Ứng",
    "content": """<p>Ngọc lớn lên trong chùa như cây bồ đề con giữa sân — không ai trồng, nhưng vẫn vươn lên.</p>

<p>Năm tuổi, cô đã biết quét sân, rửa bát, phụ sư thầy nấu cơm chay. Bảy tuổi, cô vào lớp 1 trường tiểu học Hòa Vang, đi bộ ba cây số mỗi sáng, mang theo cặp vải do sư thầy may bằng tay.</p>

<p>Ngày đầu tiên đi học, cô giáo chủ nhiệm gọi tên từng học sinh và hỏi "bố mẹ con tên gì?" Khi đến lượt Ngọc, cô im lặng.</p>

<p>"Ngọc, bố con tên gì?"</p>

<p>"Dạ, con không có bố."</p>

<p>"Vậy mẹ con?"</p>

<p>"Dạ, con không có mẹ. Con ở chùa."</p>

<p>Ba mươi sáu đứa trẻ trong lớp quay sang nhìn cô. Có đứa tò mò, có đứa thương hại, và có ba đứa — Hùng, Đạt, Phương — cười rúc rích ở bàn cuối.</p>

<p>"Con bé chùa! Mẹ nó bỏ nó vì nó xấu!"</p>

<p>Cô giáo quát, nhưng Ngọc đã nghe. Cô không khóc — cô chỉ cúi đầu, tay bấu chặt mép bàn, móng tay trắng bệch.</p>

<p>Về chùa, cô kể cho sư thầy. Thầy ngồi bên bậc tam cấp, tay lần tràng hạt, im lặng một lúc lâu.</p>

<p>"Ngọc ơi, người ta cười con vì họ chưa hiểu con. Nhưng con không cần họ hiểu — con chỉ cần con hiểu mình."</p>

<p>"Thưa thầy, vậy con là ai?"</p>

<p>"Con là Trần Thị Ngọc. Con là người Phật nhặt lên từ mưa. Và Phật không nhặt ai vô ích."</p>

<p>Ngọc ghi câu đó vào vở, bằng nét chữ lớp 1 nguệch ngoạc. Hai mươi năm sau, cô vẫn giữ cuốn vở đó trong két sắt văn phòng CEO.</p>

<hr>

<p>Chín tuổi, Ngọc bắt đầu may. Sư thầy dạy cô khâu y phục cho các sư trong chùa — những chiếc áo vạt hò đơn giản, vải nâu, chỉ vàng. Tay cô nhỏ nhưng khéo — đường kim đều tăm tắp, thẳng hơn cả những sư huynh đã may mười năm.</p>

<p>Mười tuổi, Ngọc may chiếc áo dài đầu tiên. Cô tháo một chiếc áo dài cũ mà phật tử tặng chùa, đo từng đường cắt, rồi may lại từ đầu. Chiếc áo dài hoàn thành lúc hai giờ sáng — cô ngồi dưới ánh đèn dầu trong phòng gia công nhỏ sau chánh điện, kim đâm vào ngón tay ba lần, máu thấm lên vải trắng.</p>

<p>Cô giặt vết máu, may lại đường chỉ, và khi trời sáng, chiếc áo dài trắng nằm trên bàn — đẹp đến mức sư thầy đứng nhìn rất lâu.</p>

<p>"Ngọc ơi, bàn tay con có duyên với vải."</p>

<p>"Dạ, thưa thầy, con thích may. Khi con may, con không nghĩ đến chuyện buồn."</p>

<p>Từ đó, Ngọc may áo cho tất cả phật tử đến chùa. Bà Sáu bán bánh mì đầu xóm, bác Tám chạy xe ôm, dì Năm bán rau ở chợ Hòa Vang — ai cần sửa áo, Ngọc đều may miễn phí. Đổi lại, họ cho cô vải vụn, chỉ thừa, và đôi khi một bữa cơm ngoài chùa.</p>

<p>Ngọc không biết rằng đôi bàn tay may miễn phí đó đang xây nền móng cho một đế chế thời trang. Cô chỉ biết rằng khi kim xuyên qua vải, khi đường chỉ chạy đều, cô cảm thấy mình có giá trị — một cảm giác mà không ai cho cô, trừ sợi chỉ và miếng vải.</p>"""
}

ch3 = {
    "title": "Chương 3: Cô Giáo Lương — Ân Nhân Đầu Tiên",
    "content": """<p>Mười hai tuổi, Ngọc vào lớp 6 trường THCS Hòa Vang.</p>

<p>Cô giáo dạy Công nghệ — Lương Thị Hạnh — là người đầu tiên ngoài sư thầy nhìn thấy tài năng của Ngọc. Bài tập đầu tiên của môn Công nghệ lớp 6: khâu một chiếc túi vải đơn giản. Ba mươi sáu học sinh nộp bài — ba mươi lăm chiếc túi méo mó, đường chỉ lệch, góc không vuông.</p>

<p>Chiếc túi của Ngọc: đường khâu đều như máy may công nghiệp, bốn góc vuông vắn, miệng túi thêu hoa sen bằng chỉ vàng — chi tiết mà cô giáo Lương không yêu cầu.</p>

<p>"Em học may ở đâu?"</p>

<p>"Dạ, em may áo cho sư thầy ở chùa, cô."</p>

<p>Cô Lương cầm chiếc túi lên, xoay từng góc, nhìn đường kim. Bà dạy may hai mươi năm — và bà nhận ra: đường kim của đứa trẻ mười hai tuổi này ngang bằng thợ may có năm năm kinh nghiệm.</p>

<p>Từ đó, mỗi chiều thứ Bảy, cô Lương gọi Ngọc đến nhà — một căn nhà cấp bốn ở xóm 3 Hòa Vang, đầy vải vụn và máy khâu. Bà dạy Ngọc những thứ không có trong sách giáo khoa: cách đọc rập (pattern), cách cắt vải đúng canh sợi, cách may bằng máy Singer công nghiệp, cách phối màu vải, cách đo cơ thể chính xác đến từng mili.</p>

<p>"Cô Lương ơi, sao cô dạy em nhiều vậy?"</p>

<p>"Vì cô thấy bàn tay em khác. Em may không phải bằng tay — em may bằng trái tim."</p>

<p>Lớp 7, cô Lương cho Ngọc mượn máy Singer cũ mang về chùa. Ngọc may quần áo cho trẻ em nghèo trong xóm — áo đồng phục cho con bác Tám, quần cho con dì Năm, váy cho con chị Bảy. Miễn phí, chỉ cần vải.</p>

<p>Lớp 8, Ngọc thắng giải nhất cuộc thi Khéo tay Hay làm cấp huyện. Lớp 9, cô thắng giải nhất cấp thành phố. Tấm bằng khen đầu tiên Ngọc mang về treo trên tường phòng gia công sau chánh điện — cạnh bức tượng Quan Âm.</p>

<p>Cuối năm lớp 9, cô Lương gọi Ngọc đến nhà.</p>

<p>"Ngọc ơi, cô có tin vui."</p>

<p>Bà đặt trước mặt Ngọc một phong bì: thư mời nhập học từ Trường Trung cấp Nghề May và Thiết kế Thời trang TP.HCM — với học bổng toàn phần.</p>

<p>"Cô đã gửi portfolio của em — tất cả ảnh chụp sản phẩm em may trong ba năm qua. Họ chấp nhận em ngay."</p>

<p>Ngọc nhìn phong bì, rồi nhìn cô Lương.</p>

<p>"Nhưng cô ơi, Sài Gòn xa lắm. Em không có tiền..."</p>

<p>"Học bổng lo học phí. Còn tiền ăn ở, cô lo. Mỗi tháng cô gửi em hai triệu — không nhiều, nhưng đủ sống."</p>

<p>"Cô ơi, lương cô—"</p>

<p>"Lương cô đủ sống, Ngọc. Cô không có con — em là học trò cô tự hào nhất. Cô đầu tư vào em."</p>

<p>Ngọc khóc. Đó là lần thứ hai trong đời cô khóc trước mặt người khác — lần đầu là đêm mưa cổng chùa, nhưng lần đó cô không nhớ.</p>"""
}

ch4 = {
    "title": "Chương 4: Sài Gòn — Cô Bé Chùa Giữa Thành Phố Không Ngủ",
    "content": """<p>Mười lăm tuổi, Ngọc rời chùa Linh Ứng vào Sài Gòn.</p>

<p>Sư thầy tiễn cô ở cổng chùa, tay trao tấm khăn len đỏ — tấm khăn đêm mưa hai mươi tám năm trước, giờ đã cũ sờn, màu đỏ phai thành hồng nhạt.</p>

<p>"Con giữ cái này. Để nhớ mình đi từ đâu."</p>

<p>"Thưa thầy, con sẽ về."</p>

<p>"Thầy biết. Nhưng con đừng vội về — con hãy đi cho đủ xa, rồi con sẽ biết đường về."</p>

<p>Xe khách Đà Nẵng — Sài Gòn, mười tám tiếng đường. Ngọc ngồi ghế cuối, ba lô vải trên lòng, tấm khăn len đỏ gấp gọn trong túi áo. Cô nhìn qua cửa kính: đèo Hải Vân, ruộng lúa Quảng Ngãi, biển Quy Nhơn, đồi cát Ninh Thuận — cả Việt Nam trải ra trước mắt cô, rộng lớn và xa lạ.</p>

<p>Sài Gòn ập vào cô như một bức tường âm thanh. Xe cộ, còi, tiếng người, đèn neon, mùi xăng trộn mùi phở — mọi thứ nhanh hơn, ồn hơn, đông hơn bất kỳ nơi nào cô từng sống.</p>

<p>Phòng trọ ở quận Tân Bình: mười hai mét vuông, tường nứt, trần thấp, một cái quạt máy và một chiếc giường đơn. Tiền trọ: một triệu hai trăm nghìn đồng mỗi tháng — hết một nửa số tiền cô Lương gửi.</p>

<p>Trường nghề nằm ở quận 3 — Ngọc đi xe buýt mỗi sáng, hai tuyến, bốn mươi phút. Cô là học sinh trẻ nhất lớp, gầy nhất lớp, và nghèo nhất lớp. Các bạn cùng lớp hầu hết từ gia đình có điều kiện — đi học có bố mẹ đưa đón, mang theo bộ dụng cụ may hiệu Fiskars trị giá năm triệu đồng.</p>

<p>Bộ dụng cụ của Ngọc: một cây kéo cũ cô Lương tặng, thước dây nhựa, phấn may, và cuộn chỉ mua ở chợ Tân Bình — tổng cộng hai trăm nghìn đồng.</p>

<p>Nhưng khi bài tập đầu tiên trả về, Ngọc đạt điểm cao nhất lớp.</p>

<p>Bài tập thứ hai — cao nhất lớp.</p>

<p>Bài tập thứ ba — cô giáo dạy thực hành gọi cô ở lại sau giờ học.</p>

<p>"Ngọc, em may đường nào cũng đều. Em học ở đâu trước khi vào đây?"</p>

<p>"Dạ, em may áo cho phật tử ở chùa. Và cô giáo Lương ở Hòa Vang dạy em."</p>

<p>Cô giáo nhìn cô, rồi nói: "Em có năng khiếu tự nhiên. Nhưng năng khiếu không đủ — em cần học thiết kế, học kỹ thuật cao, học cả kinh doanh thời trang. Nếu em chịu khó, tôi sẽ giới thiệu em thực tập ở xưởng may."</p>

<p>Năm thứ hai, Ngọc bắt đầu thực tập tại xưởng may gia công ở Bình Dương — chuyên may hàng xuất khẩu cho các thương hiệu quốc tế. Cô làm việc từ sáu giờ sáng đến mười hai giờ đêm, lương hai triệu đồng mỗi tháng. Tay cô chai sạn, lưng cô đau nhức, mắt cô đỏ quạch vì thức khuya — nhưng cô không nghỉ một ngày nào.</p>

<p>Có lần, cô giao hàng chậm một ngày vì máy may hỏng. Khách hàng — một bà chủ shop online — xông vào xưởng, chửi cô trước mặt hai mươi công nhân.</p>

<p>"Mày là con mồ côi không ai dạy à? Giao hàng chậm một ngày, tao mất cả đơn! Đồ vô dụng!"</p>

<p>Ngọc đứng im, tay nắm chặt, môi mím. Cô không trả lời — vì cô biết nếu mở miệng, cô sẽ khóc.</p>

<p>Tối hôm đó, sư thầy gọi điện.</p>

<p>"Con ơi, con có muốn về chùa không?"</p>

<p>"Dạ không, thầy. Con chưa tạo xong nguồn gốc cho mình."</p>"""
}

ch5 = {
    "title": "Chương 5: Rào Cản — Bị Từ Chối Vì Hai Chữ Mồ Côi",
    "content": """<p>Mười tám tuổi, Ngọc tốt nghiệp trường nghề — thủ khoa. Portfolio của cô dày một trăm trang, ghi lại từng sản phẩm cô may trong ba năm: áo dài, vest, đầm dạ hội, đồng phục, áo khoác — đủ thể loại, đủ chất liệu.</p>

<p>Cô nộp đơn xin việc tại bốn công ty thời trang lớn nhất Việt Nam.</p>

<p>Công ty đầu tiên: Ivy Moda. Phỏng vấn xuất sắc. Portfolio ấn tượng. Nhưng khi HR hỏi "gia đình", Ngọc nói thật:</p>

<p>"Em lớn lên trong chùa Linh Ứng, Đà Nẵng. Em không có gia đình."</p>

<p>HR nhìn cô bằng ánh mắt thương hại — loại thương hại mà Ngọc ghét nhất, vì nó không phải sự cảm thông mà là sự đánh giá.</p>

<p>"Em ơi, vị trí này cần người có nền tảng gia đình ổn định. Chúng tôi rất tiếc."</p>

<p>Công ty thứ hai: Elise. Cùng câu hỏi, cùng câu trả lời, cùng ánh mắt thương hại, cùng lời từ chối.</p>

<p>Công ty thứ ba: NEM Fashion. Lần này, Ngọc thử nói dối: "Bố mẹ em ở Đà Nẵng, bố làm thợ mộc, mẹ bán rau." Nhưng khi HR yêu cầu số điện thoại phụ huynh, cô im lặng — vì không có số nào để cho.</p>

<p>Công ty thứ tư: Canifa. Từ chối qua email, không cần giải thích.</p>

<p>Ngọc nộp đơn thêm mười hai công ty. Bảy cái từ chối. Năm cái không trả lời.</p>

<p>Cô về phòng trọ mười hai mét vuông ở Tân Bình, ngồi trước chiếc máy khâu Singer cũ mà cô Lương tặng khi tốt nghiệp. Tay cô đặt lên bàn đạp máy khâu, ngón tay vuốt mặt kim — cái lạnh của kim thép khiến cô tỉnh lại.</p>

<p>Cô khóc. Lần này cô khóc lớn — vì cô mười tám tuổi, giỏi nhất lớp, và không ai nhận cô chỉ vì cô không có gia đình.</p>

<p>Rồi cô lau nước mắt, bật máy khâu, và may. Cô may đến ba giờ sáng — may một chiếc áo dài trắng, thêu hoa sen bằng chỉ vàng, và ở gấu áo, cô thêu một dòng chữ nhỏ mà chỉ ai lật áo mới thấy:</p>

<p>"Cô bé cổng chùa, con sẽ tự tạo nguồn gốc."</p>

<p>Chiếc áo dài đó — chiếc áo đầu tiên Ngọc may cho chính mình — sau này được trưng bày trong Bảo tàng Phụ nữ Việt Nam tại Hà Nội, với bảng ghi: "Chiếc áo dài của Trần Thị Ngọc — may lúc 18 tuổi, trong đêm bị từ chối bởi mọi nhà tuyển dụng."</p>"""
}

ch6 = {
    "title": "Chương 6: Cuộc Thi — Bước Ngoặt",
    "content": """<p>Hai mươi tuổi. Ngọc vẫn may gia công ở xưởng Bình Dương, vẫn sống trong phòng trọ mười hai mét vuông, vẫn gom từng đồng.</p>

<p>Nhưng ban đêm, cô làm một việc khác: cô thiết kế. Trên sàn phòng trọ, cô trải giấy A0 ra, vẽ bằng bút chì 2B — những mẫu áo dài, những chiếc đầm, những bộ sưu tập mà cô chưa có tiền may nhưng có thừa trí tưởng tượng.</p>

<p>Tháng Ba năm 2015, cô nhìn thấy poster cuộc thi Vietnam Young Fashion Designer trên bảng tin trường nghề cũ. Giải thưởng: năm mươi triệu đồng và học bổng du học thiết kế thời trang tại Parsons School of Design, New York.</p>

<p>Điều kiện tham gia: nộp bộ sưu tập năm bộ trang phục, chủ đề tự chọn.</p>

<p>Ngọc có kỹ năng. Ngọc có ý tưởng. Nhưng Ngọc không có tiền mua vải.</p>

<p>Cô gom hết tiền tiết kiệm: hai triệu ba trăm nghìn đồng. Giá vải ở chợ Soái Kình Lâm: lụa Hà Đông bốn trăm nghìn đồng mỗi mét, thổ cẩm Cơ Tu sáu trăm nghìn, tơ tằm Quảng Nam năm trăm nghìn. Năm bộ cần ít nhất hai mươi mét vải — tức là mười triệu đồng.</p>

<p>Không đủ. Thiếu bảy triệu bảy.</p>

<p>Ngọc đi vay. Bác Tám — bác chạy xe ôm ở xóm chùa ngày xưa, giờ chạy Grab ở Sài Gòn — cho cô mượn hai triệu. Dì Năm gửi qua ngân hàng một triệu. Cô Lương gọi điện: "Cô gửi thêm ba triệu." Sư thầy gửi một triệu — tiền công đức phật tử dành cho Ngọc.</p>

<p>Bảy triệu bảy — đủ.</p>

<p>Ngọc mua vải, đóng cửa phòng trọ, và may liên tục mười bốn ngày đêm. Cô ngủ ba tiếng mỗi đêm, ăn mì gói, uống cà phê đen — và may.</p>

<p>Bộ sưu tập cô đặt tên: "Nguồn Gốc" (Roots).</p>

<p>Năm bộ trang phục: áo dài trắng kết hợp thổ cẩm Cơ Tu của đồng bào dân tộc miền Trung; váy lụa Hà Đông phối streetwear; áo khoác tơ tằm Quảng Nam thêu hoa sen; đầm dạ hội từ vải lanh dệt tay; và bộ cuối cùng — chiếc áo dài trắng, thắt eo bằng tấm khăn len đỏ cũ kỹ — tấm khăn đêm mưa cổng chùa.</p>

<p>"Bộ sưu tập này kể câu chuyện của những người không có nguồn gốc," Ngọc viết trong thuyết minh. "Của những đứa trẻ bị bỏ rơi, của những phụ nữ bị quên lãng, của những nghệ nhân dệt vải mà không ai biết tên. Nguồn gốc không phải nơi bạn sinh ra — nguồn gốc là nơi bạn tự tạo ra."</p>

<p>Ngày thi: Nhà hát Bến Thành, quận 1, TP.HCM. Hai trăm thí sinh từ cả nước. Ban giám khảo: bốn nhà thiết kế hàng đầu Việt Nam và một giám đốc sáng tạo từ Kering Group (Pháp).</p>

<p>Khi người mẫu mặc bộ cuối — chiếc áo dài trắng với tấm khăn len đỏ — bước ra sân khấu, khán phòng im lặng.</p>

<p>Rồi vỗ tay. Vỗ tay rất lâu.</p>

<p>Ngọc đứng trên bục, tay cầm cúp, mắt tìm một người trong khán phòng — sư thầy Thích Minh Đức, ngồi hàng cuối, tay lần tràng hạt, mắt ướt.</p>

<p>Giải nhất. Năm mươi triệu đồng. Và học bổng toàn phần Parsons School of Design.</p>"""
}

ch7 = {
    "title": "Chương 7: New York — Bốn Năm Rực Cháy",
    "content": """<p>Parsons School of Design, Manhattan, New York.</p>

<p>Ngọc đến Mỹ lần đầu tiên vào mùa thu năm 2015, mang theo một vali, tấm khăn len đỏ, và tiếng Anh trình độ IELTS 5.5 — đủ để nghe giảng nhưng chưa đủ để tranh luận.</p>

<p>Lớp thiết kế thời trang có ba mươi hai sinh viên đến từ mười bảy quốc gia. Ngọc là sinh viên duy nhất từ Việt Nam, duy nhất từ trường nghề (không phải đại học), và duy nhất sống bằng học bổng hoàn toàn — không có cha mẹ gửi tiền.</p>

<p>Cô vừa học vừa làm thêm: ban ngày ở Parsons, ban đêm may gia công cho tiệm sửa quần áo ở East Village — một bà chủ người Hàn Quốc trả mười hai đô la mỗi giờ. Ngọc may từ bảy giờ tối đến mười hai giờ đêm, kiếm sáu mươi đô một ngày — đủ trả tiền ăn và metro.</p>

<p>Năm thứ nhất: GPA 3.8. Cô vượt qua rào cản ngôn ngữ bằng cách vẽ — vì bản vẽ không cần phiên dịch.</p>

<p>Năm thứ hai: GPA 3.9. Cô bắt đầu thu hút sự chú ý khi bài tập "cultural identity in fashion" của cô — một chiếc áo dài deconstructed kết hợp denim — được giáo sư trưng bày trong phòng triển lãm của khoa.</p>

<p>Năm thứ ba: cô thực tập tại studio của một nhà thiết kế Việt kiều ở Brooklyn, chuyên thiết kế sustainable fashion. Lần đầu tiên cô tiếp xúc với khái niệm "thời trang bền vững" — dùng vải tự nhiên, quy trình sản xuất xanh, trả lương công bằng cho thợ may.</p>

<p>"Thời trang bền vững không phải xu hướng," nhà thiết kế Việt kiều nói. "Nó là tương lai. Và Việt Nam — với tơ tằm, thổ cẩm, lụa, lanh — là mỏ vàng của sustainable fashion."</p>

<p>Câu nói đó thay đổi hướng đi của Ngọc mãi mãi.</p>

<p>Năm thứ tư — bộ sưu tập tốt nghiệp: "Orphan" (Mồ Côi). Áo dài hiện đại kết hợp streetwear New York, lấy cảm hứng từ trẻ mồ côi Việt Nam. Mỗi bộ đều có một chi tiết ẩn: tên và tuổi của một đứa trẻ mồ côi thực sự, thêu bên trong gấu áo.</p>

<p>Show diễn tại Parsons gây chấn động nhỏ. Tạp chí Vogue Mỹ viết: "Trần Thị Ngọc brings Vietnamese soul to global fashion — raw, emotional, unapologetic."</p>

<p>Ba nhà mốt lớn liên hệ: Dior mời thực tập, Kenzo mời cộng tác, và một quỹ đầu tư thời trang đề nghị tài trợ cô mở thương hiệu riêng tại Paris.</p>

<p>Ngọc từ chối tất cả.</p>

<p>"Em muốn về Việt Nam. Em muốn may cho người Việt."</p>

<p>Giáo sư hướng dẫn nhìn cô, lắc đầu: "You're crazy."</p>

<p>"Dạ, em biết."</p>"""
}

ch8 = {
    "title": "Chương 8: Xưởng May Mười Lăm Mét Vuông",
    "content": """<p>Hai mươi ba tuổi, Ngọc về Đà Nẵng.</p>

<p>Cô mở xưởng may đầu tiên — mười lăm mét vuông, trong con hẻm nhỏ ở quận Thanh Khê. Hai máy may cũ (một chiếc Singer, một chiếc Juki), một bàn cắt, một gương soi, và mười cuộn vải tơ tằm Quảng Nam.</p>

<p>Vốn: mười lăm triệu đồng — tiền Ngọc gom từ tiền thưởng cuộc thi, tiền làm thêm ở New York, và năm triệu sư thầy cho.</p>

<p>Khách hàng đầu tiên: một cô giáo ở trường tiểu học Hòa Vang — đặt may chiếc áo dài cho lễ Tổng kết năm học. Giá: bốn trăm nghìn đồng. Ngọc may trong hai ngày — mỗi đường kim đều tăm tắp, lót bên trong là lụa Hà Đông, cổ áo thêu chỉ vàng.</p>

<p>Cô giáo nhận áo, mặc thử, nhìn mình trong gương, rồi khóc.</p>

<p>"Ngọc ơi, chị chưa bao giờ mặc áo dài đẹp vậy."</p>

<p>"Dạ, đẹp là vì chị đẹp, không phải vì áo đẹp."</p>

<p>Cô giáo kể cho đồng nghiệp. Đồng nghiệp kể cho bạn bè. Bạn bè kể cho hàng xóm. Trong một tháng, Ngọc nhận hai mươi đơn đặt hàng — tất cả đều áo dài.</p>

<p>Tháng thứ hai: bốn mươi đơn. Ngọc thuê thêm một thợ may — chị Bảy, bốn mươi tuổi, mẹ đơn thân, từng may gia công ở khu công nghiệp mười lăm năm.</p>

<p>Tháng thứ ba: tám mươi đơn. Thuê thêm hai thợ. Xưởng chật, Ngọc chuyển sang phòng lớn hơn — ba mươi mét vuông, cùng con hẻm.</p>

<p>Tháng thứ sáu: Ngọc mở website, đăng ảnh sản phẩm. Đơn hàng từ Sài Gòn, Hà Nội, Huế bắt đầu đổ về. Cô không quảng cáo — chỉ dựa vào chất lượng và truyền miệng.</p>

<p>Đặc biệt: mỗi chiếc áo dài Ngọc may đều có một chi tiết riêng — tên của người mặc được thêu bên trong gấu áo, bằng chỉ cùng màu vải, nhìn bên ngoài không thấy nhưng người mặc biết nó ở đó. "Áo biết tên người," Ngọc gọi đó là triết lý thiết kế.</p>

<p>Một năm sau khi mở xưởng, Ngọc có mười thợ may, năm trăm khách hàng, và doanh thu hai tỷ đồng — nhỏ theo tiêu chuẩn công nghiệp, nhưng lớn đối với cô bé cổng chùa.</p>

<p>Và điều quan trọng nhất: trong mười thợ may, có ba người là trẻ mồ côi trưởng thành — Ngọc tuyển họ đầu tiên, đào tạo miễn phí, trả lương bằng thợ lành nghề.</p>

<p>"Không ai nên bị từ chối việc vì không có gia đình," Ngọc viết trong quy chế tuyển dụng. "Ở đây, nguồn gốc duy nhất được hỏi là: em có yêu nghề may không?"</p>"""
}

ch9 = {
    "title": "Chương 9: Gặp Lại Mẹ",
    "content": """<p>Hai mươi lăm tuổi. NGỌC Fashion đã có ba xưởng, năm mươi thợ may, doanh thu mười lăm tỷ đồng mỗi năm.</p>

<p>Một ngày tháng Tám, một phụ nữ trung niên đến xưởng may chính ở Thanh Khê. Bà khoảng bốn mươi hai tuổi, gầy, tóc ngắn, mặc áo bà ba xanh lá, tay xách túi nylon. Bà đứng ngoài cửa nhìn vào, không dám bước qua ngưỡng cửa, mắt đỏ hoe.</p>

<p>"Cô tìm ai ạ?"</p>

<p>Người phụ nữ khóc. Nước mắt chảy dọc hai gò má xạm nắng.</p>

<p>"Con... con là Ngọc phải không? Con chùa Linh Ứng..."</p>

<p>Ngọc đứng im. Tim cô đập nhanh — nhưng không phải vì vui mừng. Mà vì cô đã tưởng tượng khoảnh khắc này một triệu lần trong đầu, và mỗi lần cô tưởng tượng, cô không biết mình sẽ nói gì.</p>

<p>"Cô là ai?"</p>

<p>"Con ơi, mẹ là Lê Thị Hồng. Mẹ... mẹ đã bỏ con ở cổng chùa hai mươi lăm năm trước."</p>

<p>Im lặng. Tiếng máy may Singer ở phòng trong chạy đều đều.</p>

<p>"Tại sao?" Ngọc hỏi. Giọng cô bình tĩnh — nhưng tay cô nắm chặt mép bàn, móng tay trắng bệch, như hồi lớp 1.</p>

<p>Bà Hồng kể. Năm ấy bà mười bảy tuổi, con nhà nông ở Quảng Nam, lỡ có bầu với một thanh niên trong xóm. Bố mẹ bà đuổi ra khỏi nhà. Bà lên Đà Nẵng, sống trong phòng trọ, làm rửa bát cho quán ăn, đẻ con một mình trong phòng trọ. Không tiền, không sữa, không ai giúp.</p>

<p>"Mẹ không bỏ con vì mẹ không thương con. Mẹ bỏ con vì mẹ sợ con chết. Trong chùa, ít nhất con có người nuôi."</p>

<p>Ngọc nghe xong, im lặng rất lâu.</p>

<p>Cô không ôm bà Hồng. Cô không gọi "mẹ." Cô chỉ ngồi đối diện người phụ nữ đã bỏ cô dưới mưa hai mươi lăm năm trước, và cảm nhận một thứ gì đó phức tạp hơn cả giận — đó là cơn đau thương xót cho một bà mẹ mười bảy tuổi không có lựa chọn nào tốt.</p>

<p>"Con không trách mẹ. Nhưng con cần thời gian."</p>

<p>Bà Hồng gật đầu, khóc, rồi ra đi.</p>

<p>Trên bàn xưởng, tấm khăn len đỏ cũ kỹ nằm im lặng — giữ hơi ấm của hai người phụ nữ: một người đã buông, một người không bao giờ quên.</p>"""
}

ch10 = {
    "title": "Chương 10: Đế Chế NGỌC Fashion",
    "content": """<p>Ba năm sau cuộc gặp với mẹ ruột.</p>

<p>NGỌC Fashion tăng trưởng nhanh hơn bất kỳ ai dự đoán. Hai mươi tám tuổi, Ngọc có ba trăm thợ may, năm xưởng sản xuất tại Đà Nẵng, Quảng Nam, và TP.HCM. Doanh thu hàng năm: hai trăm tỷ đồng. Sản phẩm xuất khẩu sang mười hai quốc gia: Nhật Bản, Hàn Quốc, Đức, Pháp, Mỹ, Úc, Singapore, Thái Lan, Malaysia, và ba nước Bắc Âu.</p>

<p>Đặc biệt: NGỌC Fashion là thương hiệu thời trang bền vững (sustainable fashion) đầu tiên của Việt Nam đạt chứng nhận B Corp — chứng nhận quốc tế cho doanh nghiệp vì xã hội.</p>

<p>Nguyên liệu: một trăm phần trăm vải tự nhiên Việt Nam — tơ tằm Quảng Nam, lụa Hà Đông, thổ cẩm Tây Nguyên, lanh Sa Pa, cotton hữu cơ An Giang. Ngọc ký hợp đồng trực tiếp với làng nghề, trả giá cao hơn thị trường hai mươi phần trăm để đảm bảo nghệ nhân có thu nhập ổn định.</p>

<p>Nhân sự: trong ba trăm thợ may, năm mươi người là trẻ mồ côi trưởng thành — được đào tạo nghề miễn phí, cung cấp nhà ở, và trả lương bằng thợ lành nghề. Ngọc cũng tuyển dụng ưu tiên phụ nữ đơn thân, nạn nhân bạo lực gia đình, và người khuyết tật.</p>

<p>Forbes Vietnam đưa Ngọc vào danh sách 30 Under 30 Việt Nam năm 2022. Fortune gọi NGỌC Fashion "the most ethical fashion brand in Southeast Asia."</p>

<p>Nhưng Ngọc không quan tâm danh hiệu. Cô quan tâm một thứ: sáng nào đến xưởng, cô cũng đi qua từng bàn may, chào từng thợ, hỏi có ai gặp khó khăn gì không. Và khi có thợ may mới — đặc biệt là trẻ mồ côi — cô tự tay đào tạo ngày đầu tiên, như cô Lương đã đào tạo cô ngày xưa.</p>

<p>"Em có biết tại sao chị tuyển em không?" Ngọc hỏi một thợ may mới, mười tám tuổi, mồ côi từ Quảng Ngãi.</p>

<p>"Dạ, không."</p>

<p>"Vì hai mươi năm trước, chị cũng như em. Không ai nhận chị. Nên chị mở xưởng — để không đứa trẻ nào phải nghe câu 'chúng tôi rất tiếc' chỉ vì nó không có gia đình."</p>"""
}

ch11 = {
    "title": "Chương 11: Tha Thứ",
    "content": """<p>Một năm sau lần gặp đầu tiên, Ngọc gọi mẹ ruột đến xưởng.</p>

<p>Bà Hồng đến — vẫn áo bà ba, vẫn túi nylon, vẫn đôi mắt đỏ hoe. Bà ngồi đối diện Ngọc trong phòng làm việc nhỏ, tay đặt trên đùi, lưng thẳng, như thể đang chờ tuyên án.</p>

<p>"Mẹ, con đã suy nghĩ một năm."</p>

<p>Bà Hồng nín thở.</p>

<p>"Con không tha thứ được — vì con không có quyền tha thứ cho nỗi đau của đứa bé bị bỏ rơi dưới mưa. Đứa bé đó đã lớn lên, nhưng nỗi đau thì không mất đi — nó chỉ thay đổi hình dạng."</p>

<p>Bà Hồng cúi đầu, nước mắt rơi xuống tay.</p>

<p>"Nhưng con hiểu. Con hiểu mẹ mười bảy tuổi, không tiền, không nhà, không ai giúp. Con hiểu mẹ bỏ con vì mẹ không có lựa chọn nào khác. Và con hiểu rằng hai mươi lăm năm qua, mẹ đã sống với nỗi đau không kém gì con."</p>

<p>Bà Hồng khóc — không phải khóc vì buồn, mà khóc vì nhẹ. Như thể hai mươi lăm năm bà ôm một tảng đá trong ngực, và bây giờ ai đó cho phép bà đặt nó xuống.</p>

<p>"Con mời mẹ đến làm việc tại NGỌC Fashion — nếu mẹ muốn. Không phải vì mẹ là mẹ con — mà vì xưởng con đang cần người kiểm hàng, và mẹ cần một nơi để thuộc về."</p>

<p>"Con cho mẹ cơ hội thật không?"</p>

<p>"Dạ thật. Nhưng mẹ sẽ làm như mọi nhân viên khác — lương bình đẳng, quy tắc bình đẳng."</p>

<p>Bà Hồng bắt đầu làm. Cô kiểm hàng, đóng gói, dán nhãn — những công việc nhỏ nhưng cô làm cẩn thận như thể đang chuộc lại hai mươi lăm năm vắng mặt. Bà không đòi hỏi đặc quyền, không kể với ai mình là mẹ của giám đốc.</p>

<p>Và Ngọc không gọi bà là "mẹ" — chỉ gọi "cô Hồng." Nhưng mỗi sáng, khi đến xưởng sớm, Ngọc để trên bàn làm việc của bà Hồng một ly trà nóng — loại trà sen Tây Hồ mà Ngọc mua ở Hà Nội mỗi lần đi công tác.</p>

<p>Bà Hồng không hỏi ai để ly trà. Bà chỉ uống, và mỉm cười.</p>

<p>Sáu tháng sau, trong buổi liên hoan cuối năm của xưởng, Ngọc đứng lên phát biểu trước ba trăm nhân viên:</p>

<p>"Năm nay NGỌC Fashion đạt doanh thu hai trăm tỷ. Nhưng thành tựu lớn nhất của năm không phải con số — mà là tôi đã học được rằng tha thứ không phải quên, mà là chọn sống tiếp mà không mang theo hận."</p>

<p>Bà Hồng ngồi ở bàn cuối, khóc không thành tiếng.</p>"""
}

ch12 = {
    "title": "Chương 12: Tấm Khăn Len Đỏ",
    "content": """<p>Show thời trang lớn nhất năm của NGỌC Fashion — "Origins" (Nguồn Gốc) — tại Nhà hát Lớn Hà Nội. Hai nghìn khán giả. Tường thuật trực tiếp trên truyền hình và YouTube.</p>

<p>Mười bộ sưu tập, năm mươi bộ trang phục — tất cả đều bằng vải tự nhiên Việt Nam: tơ tằm Quảng Nam, lụa Hà Đông, thổ cẩm Cơ Tu, lanh Sa Pa, cotton An Giang. Người mẫu không phải siêu mẫu quốc tế — mà là nhân viên NGỌC Fashion: thợ may, kế toán, quản lý kho — những người bình thường mặc quần áo do chính tay mình may.</p>

<p>Bộ sưu tập cuối cùng: chỉ một bộ, chỉ một người mặc.</p>

<p>Trần Thị Ngọc bước ra sân khấu.</p>

<p>Cô mặc chiếc áo dài trắng — giống chiếc áo cô may năm mười tám tuổi, trong đêm bị mọi nhà tuyển dụng từ chối. Thắt eo bằng tấm khăn len đỏ cũ kỹ — tấm khăn đêm mưa cổng chùa hai mươi tám năm trước, màu đỏ đã phai thành hồng nhạt, hai mép đã sờn.</p>

<p>"Đây là tấm khăn tôi được quấn khi bị bỏ rơi," Ngọc nói trước hai nghìn khán giả, giọng rõ ràng, không run.</p>

<p>"Ai đó đã bỏ tôi trước cổng chùa lúc ba giờ sáng, dưới mưa, chỉ với tấm khăn này. Tôi không biết mẹ tôi là ai — cho đến ba năm trước. Tôi không biết mình thuộc về đâu — cho đến khi tôi tự tạo ra nơi mình thuộc về."</p>

<p>Khán phòng im phắc.</p>

<p>"Tấm khăn này không có thương hiệu. Không có tag giá. Nhưng đây là tấm vải quý nhất tôi từng chạm — vì nó là thứ duy nhất mẹ tôi để lại."</p>

<p>"Tôi xây NGỌC Fashion không phải vì tôi muốn giàu. Tôi xây nó vì tôi muốn mọi đứa trẻ mồ côi đều có một nơi để đến khi lớn lên — một nơi không hỏi em từ đâu đến, chỉ hỏi em muốn đi đến đâu."</p>

<p>Vỗ tay. Đứng dậy. Hai nghìn người đứng dậy vỗ tay.</p>

<p>Ngọc bước xuống sân khấu, đi thẳng đến hàng ghế cuối — nơi sư thầy Thích Minh Đức ngồi, tám mươi tuổi, tóc bạc, lưng còng, tay lần tràng hạt. Bên cạnh thầy là cô Lương, bác Tám, dì Năm — những người đã giúp cô từ khi cô còn là đứa bé không tên.</p>

<p>Và ở cuối hàng, ngồi im lặng, tay ôm chiếc túi nylon — bà Hồng.</p>

<p>Ngọc quỳ xuống trước mặt sư thầy.</p>

<p>"Con cảm ơn thầy đã nhặt con lên."</p>

<p>Sư thầy cười, tay đặt lên đầu cô — bàn tay gầy guộc, run run, nhưng ấm.</p>

<p>"Ngọc ơi, không phải thầy nhặt con. Là con đã tự đứng dậy."</p>

<p>Ngọc đứng lên, quay sang bà Hồng. Hai người nhìn nhau — không nói gì.</p>

<p>Rồi Ngọc tháo tấm khăn len đỏ từ eo áo dài, gấp gọn, đặt vào tay bà Hồng.</p>

<p>"Con trả mẹ cái này. Mẹ đã giữ nó trong lòng hai mươi tám năm — giờ mẹ giữ nó trên tay."</p>

<p>Bà Hồng ôm tấm khăn, khóc. Ngọc không ôm bà — nhưng cô đứng cạnh bà, và lần đầu tiên, cô gọi:</p>

<p>"Mẹ."</p>

<p>Một tiếng. Nhẹ như gió. Nhưng hai mươi tám năm mới gọi được.</p>

<p>Nhà hát Lớn Hà Nội tối hôm đó không có đứa trẻ nào bị bỏ rơi. Chỉ có một người phụ nữ ba mươi tuổi, mặc áo dài trắng, đứng giữa hai nghìn người, và cuối cùng — cuối cùng — tìm thấy nguồn gốc của mình.</p>

<p>Không phải ở cổng chùa. Không phải ở New York. Không phải trong doanh thu bốn nghìn tỷ.</p>

<p>Mà ở trong tấm khăn len đỏ — thứ mà ai đó đã quấn quanh cô lúc ba giờ sáng, dưới mưa, với tình yêu duy nhất mà người mẹ mười bảy tuổi có thể cho.</p>"""
}

story_data["chapters"] = [ch1, ch2, ch3, ch4, ch5, ch6, ch7, ch8, ch9, ch10, ch11, ch12]

# Save
with open("scratch/story_5618_rewrite.json", "w", encoding="utf-8") as f:
    json.dump(story_data, f, ensure_ascii=False, indent=2)

print("✅ Truyện 5618 viết xong — 12 CHƯƠNG!")
for i, ch in enumerate(story_data['chapters']):
    print(f"  Ch{i+1}: {ch['title']} — {len(ch['content'])} chars")

# Upload
print("\n📤 Uploading to WordPress...")
upload_url = "https://doctieuthuyet.com/overwrite_story_v13.php"
payload = json.dumps(story_data).encode('utf-8')
req = urllib.request.Request(upload_url, data=payload, headers={'Content-Type': 'application/json'})
try:
    resp = urllib.request.urlopen(req, timeout=120)
    result = json.loads(resp.read().decode('utf-8'))
    print(f"  ✅ Success! ID={result['story_id']}, Chapters={result['chapters_count']}")
except Exception as e:
    print(f"  ❌ Error: {e}")

# Update Excel
print("\n📊 Updating Excel...")
wb = openpyxl.load_workbook("danh_sach_truyen_doctieuthuyet.xlsx")
ws = wb.active
for r in range(5, ws.max_row+1):
    if ws.cell(row=r, column=2).value and str(ws.cell(row=r, column=2).value).strip() == "5618":
        ws.cell(row=r, column=3).value = story_data["title"]
        ws.cell(row=r, column=4).value = story_data["author"]
        ws.cell(row=r, column=5).value = "Sảng Văn"
        ws.cell(row=r, column=6).value = 12
        ws.cell(row=r, column=12).value = "Cô bé bị bỏ rơi ở cổng chùa Linh Ứng, lớn lên bằng nghề may, du học Parsons New York, về VN lập tập đoàn thời trang bền vững. Gặp lại mẹ ruột, tha thứ."
        ws.cell(row=r, column=13).value = "REWRITE TOÀN BỘ: 10 chương quá ngắn (300-955 chars) → VIẾT LẠI 12 CHƯƠNG ĐẦY ĐẶN"
        ws.cell(row=r, column=14).value = "☑️ Đã sửa"
        print(f"  ✅ Excel updated row {r}")
        break
wb.save("danh_sach_truyen_doctieuthuyet.xlsx")
print("\n✅ HOÀN THÀNH TRUYỆN 5618!")
