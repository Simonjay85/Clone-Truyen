#!/usr/bin/env python3
"""Batch publish V13 rewrites to WordPress via FTP+HTTP"""
import json, os, sys, requests, ftplib, time
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
WP_URL = "https://doctieuthuyet.com"
SECRET_TOKEN = "ZEN_TRUYEN_2026_BYPASS"
FTP_HOST = "51.79.53.190"
FTP_USER = "alotoinghe"
FTP_PASS = "Nghia234!"
PHP_FILE = "overwrite_story_v13.php"
EXCEL_FILE = os.path.join(BASE_DIR, "danh_sach_truyen_doctieuthuyet.xlsx")

# Story ID → Excel row mapping
EXCEL_ROWS = {
    5838: 8,
    2023: 184, 2020: 185, 2013: 186, 2007: 187,
    2001: 188, 1984: 189, 1968: 190, 1948: 191,
    1933: 192, 1927: 193, 1921: 194
}

def upload_php():
    print("[FTP] Uploading PHP helper...")
    local_path = os.path.join(BASE_DIR, PHP_FILE)
    ftp = ftplib.FTP(FTP_HOST)
    ftp.login(FTP_USER, FTP_PASS)
    with open(local_path, 'rb') as f:
        ftp.storbinary(f'STOR {PHP_FILE}', f)
    ftp.quit()
    print("[FTP] Upload done.")

def cleanup_php():
    print("[FTP] Cleaning up PHP helper...")
    try:
        ftp = ftplib.FTP(FTP_HOST)
        ftp.login(FTP_USER, FTP_PASS)
        ftp.delete(PHP_FILE)
        ftp.quit()
        print("[FTP] Cleanup done.")
    except Exception as e:
        print(f"[FTP] Cleanup warning: {e}")

def publish_story(json_path):
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    story_id = data['story_id']
    print(f"\n{'='*60}")
    print(f"[PUBLISH] Story #{story_id}: {data['title']}")
    print(f"[PUBLISH] Chapters: {len(data['chapters'])}")

    intro = data.get('intro', '')
    if not intro:
        # Extract V13 intro from chapter 1 opening strong tags
        ch1 = data['chapters'][0]['content']
        import re
        strong_ps = re.findall(r'<p><strong>.*?</strong></p>', ch1)
        if strong_ps:
            intro = '\n'.join(strong_ps)

    payload = {
        "secret_token": SECRET_TOKEN,
        "story_id": story_id,
        "title": data['title'],
        "intro": intro,
        "author": data.get('author', 'Đang cập nhật'),
        "chapters": [
            {"title": ch['title'], "content": ch['content']}
            for ch in data['chapters']
        ],
        "seo": data.get('seo', {})
    }

    url = f"{WP_URL}/{PHP_FILE}"
    resp = requests.post(url, json=payload, timeout=120)

    if resp.status_code == 200:
        result = resp.json()
        print(f"[OK] Published! Deleted {result.get('deleted_old_chapters',0)} old chapters, created {result.get('chapters_count',0)} new.")
        return True
    else:
        print(f"[ERROR] HTTP {resp.status_code}: {resp.text[:300]}")
        return False

def mark_excel(story_id):
    row = EXCEL_ROWS.get(story_id)
    if not row:
        print(f"[EXCEL] No row mapping for story #{story_id}")
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.cell(row=row, column=14, value="☑️ Đã sửa")
    wb.save(EXCEL_FILE)
    print(f"[EXCEL] Marked story #{story_id} as done (row {row})")

def main():
    story_ids = sys.argv[1:] if len(sys.argv) > 1 else []
    if not story_ids:
        print("Usage: python batch_publish_v13.py 5838 2023 ...")
        return

    json_files = []
    for sid in story_ids:
        path = os.path.join(BASE_DIR, 'scratch', f'rewrite_{sid}_v13_full.json')
        if os.path.exists(path):
            json_files.append((int(sid), path))
        else:
            print(f"[SKIP] No rewrite file for story #{sid}")

    if not json_files:
        print("No stories to publish.")
        return

    upload_php()
    time.sleep(2)

    for story_id, path in json_files:
        ok = publish_story(path)
        if ok:
            mark_excel(story_id)
        time.sleep(1)

    cleanup_php()
    print(f"\n{'='*60}")
    print(f"Done! Published {len(json_files)} stories.")

if __name__ == '__main__':
    main()
