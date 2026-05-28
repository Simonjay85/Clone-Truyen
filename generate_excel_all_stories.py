import ftplib
import urllib.request
import json
import os
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FTP_HOST = "51.79.53.190"
FTP_USER = "alotoinghe"
FTP_PASS = "Nghia234!"
PHP_FILE = "get_all_stories_detailed.php"
JSON_OUTPUT = "all_stories_detailed.json"
EXCEL_OUTPUT = "danh_sach_truyen_doctieuthuyet.xlsx"
BYPASS_TOKEN = "ZEN_TRUYEN_2026_BYPASS"
AUDIT_FILE = "de_xuat_sua_truyen_dat_10_10_2026-05-24.md"

def fetch_data():
    print("Step 1: Uploading PHP script via FTP...")
    try:
        ftp = ftplib.FTP(FTP_HOST, timeout=30)
        ftp.login(FTP_USER, FTP_PASS)
        
        with open(PHP_FILE, "rb") as f:
            ftp.storbinary(f"STOR {PHP_FILE}", f)
        print("✓ PHP script uploaded successfully.")
        
        print("\nStep 2: Calling API to fetch all stories metadata...")
        url = f"https://doctieuthuyet.com/{PHP_FILE}?token={BYPASS_TOKEN}"
        
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=120) as response:
            resp_text = response.read().decode('utf-8')
            
        print("✓ Data retrieved successfully. Cleaning up PHP script from server...")
        try:
            ftp.delete(PHP_FILE)
            print("✓ PHP script removed from server.")
        except Exception as de:
            print(f"⚠️ Warning: Could not delete PHP script: {de}")
            
        ftp.quit()
        
        try:
            data = json.loads(resp_text)
            print(f"✓ Parsed {len(data)} stories from website.")
            
            with open(JSON_OUTPUT, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print(f"✓ Saved raw JSON data to {JSON_OUTPUT}")
            return data
        except Exception as je:
            print("❌ Error: Response was not valid JSON.")
            print("Snippet of response:")
            print(resp_text[:1000])
            return None
            
    except Exception as e:
        print(f"❌ Error during data fetch: {e}")
        try:
            ftp = ftplib.FTP(FTP_HOST, timeout=10)
            ftp.login(FTP_USER, FTP_PASS)
            ftp.delete(PHP_FILE)
            ftp.quit()
            print("✓ Cleaned up PHP script after error.")
        except:
            pass
        return None

def parse_audit_file():
    audit_data = {}
    if not os.path.exists(AUDIT_FILE):
        print(f"⚠️ Warning: Audit file {AUDIT_FILE} not found. Using dynamic rules only.")
        return audit_data
        
    print(f"✓ Reading audit file {AUDIT_FILE} to map recommendations...")
    try:
        with open(AUDIT_FILE, "r", encoding="utf-8") as f:
            content = f.read()
            
        blocks = content.split("### ")
        for block in blocks[1:]:
            lines = block.split("\n")
            if not lines:
                continue
            slug = None
            points_to_fix = ""
            for line in lines:
                line_str = line.strip()
                if line_str.startswith("- Link:"):
                    match = re.search(r'/truyen/([^/]+)/?', line_str)
                    if match:
                        slug = match.group(1)
                elif line_str.startswith("- Cần sửa:"):
                    points_to_fix = line_str.replace("- Cần sửa:", "").strip()
            if slug:
                audit_data[slug] = points_to_fix
        print(f"✓ Parsed {len(audit_data)} stories from audit file.")
    except Exception as e:
        print(f"⚠️ Error parsing audit file: {e}")
    return audit_data

def build_excel(data):
    if not data:
        print("❌ No data to export.")
        return
        
    # Parse audit recommendations
    audit_recommendations = parse_audit_file()
        
    print("\nStep 3: Creating beautiful styled Excel file with Audit columns...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách truyện"
    
    ws.views.sheetView[0].showGridLines = True
    
    # Styles definition
    font_family = "Segoe UI"
    
    title_font = Font(name=font_family, size=16, bold=True, color="1A365D")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="555555")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10)
    bold_data_font = Font(name=font_family, size=10, bold=True)
    link_font = Font(name=font_family, size=10, color="0000FF", underline="single")
    
    # Checkbox specific styles
    red_fill = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")
    red_font = Font(name=font_family, size=10, bold=True, color="9B2C2C")
    green_fill = PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")
    green_font = Font(name=font_family, size=10, bold=True, color="22543D")
    
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    odd_row_fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="E2E8F0")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Title Block
    ws['A1'] = "DANH SÁCH TẤT CẢ CÁC TRUYỆN TRÊN WEBSITE DOCTIEUTHUYET.COM"
    ws['A1'].font = title_font
    ws.merge_cells('A1:N1')
    ws.row_dimensions[1].height = 30
    
    ws['A2'] = f"Tổng hợp tự động & Audit chất lượng truyện - Số lượng: {len(data)} truyện - Ngày cập nhật: 2026-05-27"
    ws['A2'].font = subtitle_font
    ws.merge_cells('A2:N2')
    ws.row_dimensions[2].height = 18
    
    ws.row_dimensions[3].height = 10
    
    # Headers
    headers = [
        "STT", "ID Truyện", "Tên Truyện", "Tác Giả", "Thể Loại", 
        "Số Chương", "Lượt Xem", "Trạng Thái", "Ngày Đăng", "Đường Dẫn", "Ảnh Bìa", "Tóm Tắt",
        "Điểm Cần Sửa", "Trạng Thái Sửa"
    ]
    
    header_row = 4
    ws.row_dimensions[header_row].height = 28
    
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    # Setup Data Validation (Dropdown Box) for Column 14 (Trạng Thái Sửa)
    dv = DataValidation(type="list", formula1='"⬜ Chưa sửa,☑️ Đã sửa"', allow_blank=True)
    dv.error = 'Vui lòng chọn giá trị trong danh sách'
    dv.errorTitle = 'Lỗi nhập liệu'
    dv.prompt = 'Chọn trạng thái sửa cho truyện'
    dv.promptTitle = 'Trạng thái sửa'
    ws.add_data_validation(dv)
        
    # Data Rows
    current_row = 5
    for idx, item in enumerate(data, 1):
        ws.row_dimensions[current_row].height = 24
        
        row_fill = odd_row_fill if idx % 2 != 0 else even_row_fill
        
        # 1. STT
        c1 = ws.cell(row=current_row, column=1, value=idx)
        c1.alignment = align_center
        
        # 2. ID
        c2 = ws.cell(row=current_row, column=2, value=item['id'])
        c2.alignment = align_center
        
        # 3. Tên Truyện
        c3 = ws.cell(row=current_row, column=3, value=item['title'])
        c3.font = bold_data_font
        c3.alignment = align_left
        
        # 4. Tác Giả
        c4 = ws.cell(row=current_row, column=4, value=item['author'])
        c4.alignment = align_left
        
        # 5. Thể Loại
        c5 = ws.cell(row=current_row, column=5, value=item['genre'])
        c5.alignment = align_left
        
        # 6. Số Chương
        c6 = ws.cell(row=current_row, column=6, value=item['chapters'])
        c6.number_format = "#,##0"
        c6.alignment = align_center
        
        # 7. Lượt Xem
        c7 = ws.cell(row=current_row, column=7, value=item['views'])
        c7.number_format = "#,##0"
        c7.alignment = align_right
        
        # 8. Trạng Thái
        status_vi = "Đã đăng" if item['status'] == 'publish' else item['status']
        c8 = ws.cell(row=current_row, column=8, value=status_vi)
        c8.alignment = align_center
        
        # 9. Ngày Đăng
        c9 = ws.cell(row=current_row, column=9, value=item['date'])
        c9.alignment = align_center
        
        # 10. Đường Dẫn
        c10 = ws.cell(row=current_row, column=10, value="Xem website")
        c10.hyperlink = item['url']
        c10.font = link_font
        c10.alignment = align_center
        
        # 11. Ảnh Bìa
        c11 = ws.cell(row=current_row, column=11, value="Mở ảnh bìa" if item['cover'] else "")
        if item['cover']:
            c11.hyperlink = item['cover']
            c11.font = link_font
        c11.alignment = align_center
        
        # 12. Tóm Tắt
        c12 = ws.cell(row=current_row, column=12, value=item['intro'])
        c12.alignment = align_wrap
        
        # Determine "Điểm cần sửa" and "Trạng thái sửa"
        slug = item['slug']
        chapters = item['chapters']
        title = item['title']
        intro = item['intro']
        
        points_to_fix = ""
        status_fix = "☑️ Đã sửa"
        
        if slug in audit_recommendations:
            points_to_fix = audit_recommendations[slug]
            if points_to_fix and "Ổn" not in points_to_fix and "Đạt" not in points_to_fix:
                status_fix = "⬜ Chưa sửa"
            else:
                points_to_fix = "Ổn ở mức tổng quan (Đạt 10/10)."
                status_fix = "☑️ Đã sửa"
        else:
            # Dynamic rules for newer stories
            issues = []
            if len(title) > 75:
                issues.append("Tên truyện dài (dễ vỡ layout/SEO)")
            if chapters < 8:
                issues.append(f"Số chương ít ({chapters}/8 chương, cần bồi thêm chương)")
            if len(intro) < 100:
                issues.append("Mô tả truyện hơi ngắn, cần tăng độ hấp dẫn")
                
            if issues:
                points_to_fix = "; ".join(issues) + "."
                status_fix = "⬜ Chưa sửa"
            else:
                points_to_fix = "Đã tối ưu theo tiêu chuẩn Gold V13 (Đạt 10/10)."
                status_fix = "☑️ Đã sửa"
                
        # USER OVERRIDE V2: From the third story onwards (idx >= 3), force status_fix to "⬜ Chưa sửa"
        # Otherwise (idx == 1 or 2), force to "☑️ Đã sửa"
        if idx >= 3:
            status_fix = "⬜ Chưa sửa"
        else:
            status_fix = "☑️ Đã sửa"
                
        # 13. Điểm Cần Sửa
        c13 = ws.cell(row=current_row, column=13, value=points_to_fix)
        c13.alignment = align_wrap
        
        # 14. Trạng Thái Sửa (Tick Box / Dropdown)
        c14 = ws.cell(row=current_row, column=14, value=status_fix)
        c14.alignment = align_center
        dv.add(c14) # register dropdown validation
        
        # Apply borders and backgrounds to all cells in the row
        for col_idx in range(1, 15):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = thin_border
            cell.fill = row_fill
            
            # Text fonts
            if col_idx not in [3, 10, 11, 14]:
                cell.font = data_font
            elif col_idx == 14:
                # Color code status column
                if status_fix == "⬜ Chưa sửa":
                    cell.fill = red_fill
                    cell.font = red_font
                else:
                    cell.fill = green_fill
                    cell.font = green_font
                
        current_row += 1
        
    # Auto-adjust column widths
    col_widths = {
        1: 6,    # STT
        2: 12,   # ID
        3: 40,   # Tên Truyện
        4: 20,   # Tác Giả
        5: 25,   # Thể Loại
        6: 12,   # Số Chương
        7: 15,   # Lượt Xem
        8: 15,   # Trạng thái
        9: 22,   # Ngày Đăng
        10: 15,  # Đường Dẫn
        11: 15,  # Ảnh Bìa
        12: 50,  # Tóm Tắt
        13: 50,  # Điểm Cần Sửa
        14: 18   # Trạng Thái Sửa
    }
    
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        
    # Apply Auto Filter across all columns (A4:N{current_row-1})
    ws.auto_filter.ref = f"A4:N{current_row - 1}"
    
    wb.save(EXCEL_OUTPUT)
    print(f"✓ Styled Excel sheet generated successfully: {EXCEL_OUTPUT}")
    
def main():
    print("=== START STORY COMPILATION AND EXCEL EXPORT ===")
    data = fetch_data()
    if data:
        build_excel(data)
        print("\n=== SUCCESS ===")
        print(f"Found total: {len(data)} stories.")
        print(f"Output files ready:")
        print(f"1. Excel format: {EXCEL_OUTPUT}")
        print(f"2. JSON data format: {JSON_OUTPUT}")
    else:
        print("\n❌ Failed to compile stories.")

if __name__ == "__main__":
    main()
