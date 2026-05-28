"""
story_fixer_utils.py
Utility functions to fetch, analyze and update stories on doctieuthuyet.com
"""
import ftplib
import urllib.request
import json
import os

FTP_HOST = "51.79.53.190"
FTP_USER = "alotoinghe"
FTP_PASS = "Nghia234!"
TOKEN = "ZEN_TRUYEN_2026_BYPASS"
WP_URL = "https://doctieuthuyet.com"

def upload_php(local_file, remote_file=None):
    """Upload a PHP file to server via FTP."""
    if remote_file is None:
        remote_file = os.path.basename(local_file)
    ftp = ftplib.FTP(FTP_HOST, timeout=30)
    ftp.login(FTP_USER, FTP_PASS)
    with open(local_file, "rb") as f:
        ftp.storbinary(f"STOR {remote_file}", f)
    ftp.quit()
    return True

def delete_php(remote_file):
    """Delete a PHP file from server via FTP."""
    try:
        ftp = ftplib.FTP(FTP_HOST, timeout=15)
        ftp.login(FTP_USER, FTP_PASS)
        ftp.delete(remote_file)
        ftp.quit()
    except:
        pass

def call_php(remote_file, payload):
    """Call a PHP endpoint on the server with JSON payload."""
    url = f"{WP_URL}/{remote_file}"
    payload["secret_token"] = TOKEN
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers={
        "Content-Type": "application/json",
        "User-Agent": "Mozilla/5.0"
    })
    with urllib.request.urlopen(req, timeout=120) as resp:
        return json.loads(resp.read().decode("utf-8"))

def fetch_story(story_id):
    """Fetch complete story with all chapters from server."""
    upload_php("fetch_story_full.php")
    try:
        result = call_php("fetch_story_full.php", {"story_id": story_id})
        return result
    finally:
        delete_php("fetch_story_full.php")

def update_story(story_id, title, intro, author, chapters):
    """Overwrite story on server with new content."""
    upload_php("overwrite_story_v13.php")
    try:
        payload = {
            "story_id": story_id,
            "title": title,
            "intro": intro,
            "author": author,
            "chapters": chapters
        }
        result = call_php("overwrite_story_v13.php", payload)
        return result
    finally:
        delete_php("overwrite_story_v13.php")

def update_excel_status(story_idx, status="☑️ Đã sửa"):
    """Update Excel file to mark a story as fixed."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    
    excel_file = "danh_sach_truyen_doctieuthuyet.xlsx"
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # story_idx is 1-based STT, data starts at row 5
    row = 4 + story_idx
    cell = ws.cell(row=row, column=14)
    cell.value = status
    
    if status == "☑️ Đã sửa":
        cell.fill = PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="22543D")
    else:
        cell.fill = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="9B2C2C")
    
    wb.save(excel_file)
    return True

if __name__ == "__main__":
    # Quick test
    print("story_fixer_utils loaded successfully")
    print(f"FTP Host: {FTP_HOST}")
    print(f"WordPress URL: {WP_URL}")
